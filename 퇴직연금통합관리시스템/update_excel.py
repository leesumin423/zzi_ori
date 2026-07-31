import os
import shutil
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from datetime import datetime, date

# Paths setup
DATA_DIR = r"C:\Users\tyinc\Desktop\업무리스트 내역 정리\퇴직연금 관련\퇴직연금 관련"
BACKUP_DIR = os.path.join(DATA_DIR, "백업")

# Create backup directory if not exists
if not os.path.exists(BACKUP_DIR):
    os.makedirs(BACKUP_DIR)
    print(f"Created backup directory: {BACKUP_DIR}")

# Base dates mapping
BASE_DATES = {
    "2021. 12월말": date(2021, 12, 31),
    "2022. 12월 말": date(2022, 12, 31),
    "2023. 12월 말": date(2023, 12, 31),
    "2024.12월 말": date(2024, 12, 31),
    "2025.12월 말": date(2025, 12, 31),
    "1. 퇴직연금현황(2026.06월 말": date(2026, 6, 30),
    "1. 퇴직연금현황(2026.05월 말": date(2026, 5, 31),
    "1. 퇴직연금현황(2025.12월 말": date(2025, 12, 31),
}

# Files to update
TARGET_FILES = [
    (r"템플릿\퇴직연금 현황양식(주)동양_5개년 데이터.xlsx", ["2021. 12월말", "2022. 12월 말", "2023. 12월 말", "2024.12월 말", "2025.12월 말"]),
    (r"퇴직연금 현황양식(주)동양_26.6월 말기준.xlsx", ["1. 퇴직연금현황(2026.06월 말"]),
    (r"퇴직연금 현황양식(주)동양_26.5월 말기준_유진260622 회신.xlsx", ["1. 퇴직연금현황(2026.05월 말"]),
    (r"퇴직연금 현황양식(주)동양_25.12월 말기준.xlsx", ["1. 퇴직연금현황(2025.12월 말"])
]

def clean_inst_name(name):
    if not name:
        return ""
    name = str(name).strip().replace(" ", "")
    if "농협" in name: return "농협"
    if "산업은행" in name: return "산업은행"
    if "우리은행" in name: return "우리은행"
    if "기업은행" in name: return "기업은행"
    if "국민은행" in name: return "국민은행"
    if "신한" in name: return "신한은행"
    if "im뱅크" in name.lower() or "대구은행" in name: return "iM뱅크"
    if "하나은행" in name: return "하나은행"
    if "현대해상" in name: return "현대해상"
    if "삼성생명" in name: return "삼성생명"
    return name.split("/")[0].split()[0]

def parse_date_value(val):
    if val is None:
        return None
    if isinstance(val, (datetime, date)):
        return val.date() if isinstance(val, datetime) else val
    val_str = str(val).strip()
    for fmt in ["%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%Y.%m.%d", "%y.%m.%d"]:
        try:
            return datetime.strptime(val_str, fmt).date()
        except ValueError:
            pass
    return None

def process_file(rel_path, sheets):
    src_path = os.path.join(DATA_DIR, rel_path)
    if not os.path.exists(src_path):
        print(f"File not found: {src_path}")
        return
        
    # Backup
    dest_name = os.path.basename(src_path)
    backup_path = os.path.join(BACKUP_DIR, dest_name)
    shutil.copy2(src_path, backup_path)
    print(f"Backed up {rel_path} to {backup_path}")
    
    # Load workbook preserving formulas
    wb = openpyxl.load_workbook(src_path, data_only=False)
    
    for sheet_name in sheets:
        if sheet_name not in wb.sheetnames:
            print(f"  Sheet {sheet_name} not found in workbook.")
            continue
            
        ws = wb[sheet_name]
        base_date = BASE_DATES.get(sheet_name)
        if not base_date:
            print(f"  Base date for {sheet_name} not defined. Skipping.")
            continue
            
        print(f"  Processing sheet: {sheet_name} (Base Date: {base_date})")
        
        # Determine column letters/indices by searching header row (Row 4 or Row 3)
        col_headers = {}
        for r in [3, 4]:
            for c in range(1, 15):
                val = ws.cell(r, c).value
                if val:
                    col_headers[str(val).strip().replace(" ", "")] = c
                    
        rate_col = col_headers.get("금리", 5)
        amount_col = col_headers.get("금액", 8)
        if "평가금액" in col_headers:
            amount_col = col_headers["평가금액"]
        start_col = col_headers.get("최초가입", 6)
        end_col = col_headers.get("만기", 7)
        
        amount_col_letter = get_column_letter(amount_col)
        
        # We need to trace the groups of institutions to unmerge & merge them
        inst_groups = [] # List of tuples: (inst_name, start_row, end_row)
        current_inst = None
        current_start_row = None
        last_processed_row = 5
        
        # Categories list of rows
        cat_rows = {
            "1년 이내": [],
            "1년이상 2년미만": [],
            "2년이상 3년미만": []
        }
        
        # First pass: Identify rows and construct institution groups
        for r in range(5, 100):
            val_b = ws.cell(r, 2).value
            val_c = ws.cell(r, 3).value
            val_d = ws.cell(r, 4).value
            
            # Stop if Grand Total row
            if val_b and any(x in str(val_b).replace(" ", "") for x in ["합계"]):
                if current_inst and current_start_row:
                    inst_groups.append((current_inst, current_start_row, r - 1))
                    current_inst = None
                break
            if val_c and any(x in str(val_c).replace(" ", "") for x in ["합계"]):
                if current_inst and current_start_row:
                    inst_groups.append((current_inst, current_start_row, r - 1))
                    current_inst = None
                break
                
            # Close group if Subtotal row
            is_subtotal = False
            if val_c and any(x in str(val_c).replace(" ", "") for x in ["소계"]):
                is_subtotal = True
            if val_b and any(x in str(val_b).replace(" ", "") for x in ["소계"]):
                is_subtotal = True
                
            if is_subtotal:
                if current_inst and current_start_row:
                    inst_groups.append((current_inst, current_start_row, r - 1))
                    current_inst = None
                continue
                
            # If a new institution name appears
            if val_c:
                if current_inst and current_start_row:
                    inst_groups.append((current_inst, current_start_row, r - 1))
                current_inst = val_c
                current_start_row = r
                
            last_processed_row = r
            
            # Classify maturity for this row
            amount_val = ws.cell(r, amount_col).value
            end_date_val = ws.cell(r, end_col).value
            
            if val_d and amount_val is not None:
                end_date = parse_date_value(end_date_val)
                
                if end_date is None:
                    maturity_cat = "1년 이내"
                else:
                    diff_days = (end_date - base_date).days
                    if diff_days <= 365:
                        maturity_cat = "1년 이내"
                    elif diff_days <= 730:
                        maturity_cat = "1년이상 2년미만"
                    else:
                        maturity_cat = "2년이상 3년미만"
                        
                cat_rows[maturity_cat].append(r)
                
        # Close the last group if it was open
        if current_inst and current_start_row:
            inst_groups.append((current_inst, current_start_row, last_processed_row))
            
        # Unmerge any merged cells in Column B that overlap with our rows
        merged_ranges_to_remove = []
        for r_range in list(ws.merged_cells.ranges):
            # Check if range intersects with Column B (Col 2) and is in data rows (>= 5)
            if r_range.min_col <= 2 <= r_range.max_col and r_range.min_row >= 5:
                merged_ranges_to_remove.append(r_range)
                
        for r_range in merged_ranges_to_remove:
            ws.unmerge_cells(start_row=r_range.min_row, start_column=r_range.min_col,
                             end_row=r_range.max_row, end_column=r_range.max_col)
                             
        # Apply new merging and write clean names
        for inst_name, start_r, end_r in inst_groups:
            clean_name = clean_inst_name(inst_name)
            if not clean_name:
                continue
                
            # Merge if more than one row in the group
            if start_r < end_r:
                ws.merge_cells(start_row=start_r, start_column=2, end_row=end_r, end_column=2)
                
            # Write value to top-left cell
            top_cell = ws.cell(start_r, 2)
            top_cell.value = clean_name
            top_cell.alignment = Alignment(horizontal="center", vertical="center")
            
        print(f"    Column B updated with clean institution names.")

        # Second pass: Find and update Maturity Table
        table_row, table_col = None, None
        for r in range(2, 5):
            for c in range(12, 23):
                val = ws.cell(r, c).value
                if val and str(val).strip().replace(" ", "") in ["구분", "구분(만기)"]:
                    next_val = ws.cell(r, c+1).value
                    if next_val and str(next_val).strip().replace(" ", "") in ["금액", "평가금액"]:
                        table_row = r
                        table_col = c
                        break
            if table_row:
                break
                
        if not table_row:
            print(f"    Maturity Table not found. Creating a new one in Col S, T, U starting at Row 3.")
            table_row = 3
            table_col = 19
            ws.cell(table_row, table_col).value = "구분"
            ws.cell(table_row, table_col+1).value = "금액"
            ws.cell(table_row, table_col+2).value = "비율"
            ws.cell(table_row+1, table_col).value = "1년 이내"
            ws.cell(table_row+2, table_col).value = "1년이상 2년미만"
            ws.cell(table_row+3, table_col).value = "2년이상 3년미만"
            ws.cell(table_row+4, table_col).value = "계"
            
        lbl_col = table_col
        amt_col = table_col + 1
        pct_col = table_col + 2
        
        lbl_col_letter = get_column_letter(lbl_col)
        amt_col_letter = get_column_letter(amt_col)
        pct_col_letter = get_column_letter(pct_col)
        
        category_cell_mapping = {}
        total_cell = None
        
        for r in range(table_row + 1, table_row + 10):
            val = ws.cell(r, lbl_col).value
            if val:
                val_str = str(val).strip().replace(" ", "")
                if val_str in ["1년이내", "1년", "1년미만"]:
                    category_cell_mapping["1년 이내"] = r
                elif val_str in ["1년이상2년미만", "2년", "1년이상"]:
                    category_cell_mapping["1년이상 2년미만"] = r
                elif val_str in ["2년이상3년미만", "3년", "2년이상"]:
                    category_cell_mapping["2년이상 3년미만"] = r
                elif val_str in ["계", "합계", "합  계"]:
                    total_cell = r
                    break
                    
        for cat, r_indices in cat_rows.items():
            tbl_r = category_cell_mapping.get(cat)
            if tbl_r:
                if r_indices:
                    formula_amt = f"=SUM(" + ",".join(f"{amount_col_letter}{ri}" for ri in r_indices) + ")"
                else:
                    formula_amt = 0
                ws.cell(tbl_r, amt_col).value = formula_amt
                
                if total_cell:
                    ws.cell(tbl_r, pct_col).value = f"={amt_col_letter}{tbl_r}/${amt_col_letter}${total_cell}"
                    
        if total_cell:
            r_1yr = category_cell_mapping.get("1년 이내")
            r_2yr = category_cell_mapping.get("1년이상 2년미만")
            r_3yr = category_cell_mapping.get("2년이상 3년미만")
            
            if r_1yr and r_3yr:
                ws.cell(total_cell, amt_col).value = f"=SUM({amt_col_letter}{r_1yr}:{amt_col_letter}{r_3yr})"
            ws.cell(total_cell, pct_col).value = f"={amt_col_letter}{total_cell}/${amt_col_letter}${total_cell}"
            
        print(f"    Maturity Table updated successfully in {sheet_name}.")
        
        # Write Col O and Col P formulas if they exist (Col 15 & 16)
        if sheet_name in ["1. 퇴직연금현황(2026.06월 말", "1. 퇴직연금현황(2026.05월 말", "1. 퇴직연금현황(2025.12월 말", "2025.12월 말"]:
            for cat, r_indices in cat_rows.items():
                tbl_r = category_cell_mapping.get(cat)
                if tbl_r:
                    if r_indices:
                        formula_int = f"=SUM(" + ",".join(f"K{ri}" for ri in r_indices) + ")"
                    else:
                        formula_int = 0
                    ws.cell(tbl_r, 16).value = formula_int # Col P
                    ws.cell(tbl_r, 15).value = f"=P{tbl_r}/T{tbl_r}" # Col O
            print(f"    Expected interest and weighted yield formulas written in Col O & P.")

    # Save workbook
    wb.save(src_path)
    print(f"Saved modified workbook to {src_path}\n")

# Run processing
for rel_path, sheets in TARGET_FILES:
    process_file(rel_path, sheets)

print("Excel modification complete.")
