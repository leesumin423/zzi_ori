# -*- coding: utf-8 -*-
"""퇴직연금 대시보드 서버(Flask). 예전 run_dashboard.py는 파이썬 내장 http.server로
정적 파일만 서빙하고 data.json은 서버 시작 시 딱 한 번만 만들었는데, 그러면 관리자가
기준일을 추가하려면 parse_data.py의 BASE_DATES를 직접 고치고 서버를 재시작해야 했다.
이제 pension_db.py(SQLite)에 등록한 기준일은 /data.json 요청마다 즉석에서 반영된다
(관리자 화면은 /admin — 비밀번호는 pension_config.py)."""
import io
import os
import time
from datetime import datetime

from flask import (
    Flask, request, session, redirect, url_for, flash,
    send_from_directory, send_file, jsonify, render_template, get_flashed_messages,
)

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

import parse_data
import pension_db
import pension_config

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

app = Flask(__name__, template_folder=os.path.join(BASE_DIR, 'templates'))
app.secret_key = 'pension-dashboard-local-secret'  # 이 PC/LAN 내부용 — 외부 노출 안 됨

_DATA_CACHE_TTL = 300  # 초 — 레거시 파일이 그새 바뀌었을 가능성 대비, 업로드 시엔 즉시 무효화
_data_cache = {"data": None, "ts": 0.0}


def _get_data():
    now = time.time()
    if _data_cache["data"] is None or now - _data_cache["ts"] > _DATA_CACHE_TTL:
        _data_cache["data"] = parse_data.build_snapshots_and_yoy()
        _data_cache["ts"] = now
    return _data_cache["data"]


def _invalidate_cache():
    _data_cache["data"] = None


def _is_admin():
    return session.get('pension_admin') is True


@app.route('/data.json')
def data_json():
    return jsonify(_get_data())


_EXPORT_HEADER_FILL = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
_EXPORT_HEADER_FONT = Font(color='FFFFFF', bold=True)
_EXPORT_SUBTOTAL_FILL = PatternFill(start_color='DCE6F1', end_color='DCE6F1', fill_type='solid')
_EXPORT_TOTAL_FILL = PatternFill(start_color='9DC3E6', end_color='9DC3E6', fill_type='solid')
_EXPORT_BOLD = Font(bold=True)


def _export_write_sheet(ws, headers, rows):
    """헤더 행(진한 배경 + 흰 글씨)과 데이터 행을 쓰고, 각 열 너비를 내용에 맞춰
    자동으로 넓힌다 — 시트마다 반복되는 서식이라 함수로 뺐다."""
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = _EXPORT_HEADER_FILL
        cell.font = _EXPORT_HEADER_FONT
        cell.alignment = Alignment(horizontal='center')
    for row in rows:
        ws.append(row)
    widths = [len(str(h)) for h in headers]
    for row in rows:
        for i, v in enumerate(row):
            widths[i] = max(widths[i], len(str(v)) if v is not None else 0)
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = min(max(w + 2, 10), 45)
    ws.freeze_panes = 'A2'


def _export_write_items_sheet(ws, items, institution_order):
    """상세내역 시트 전용 — 기관별로 묶어 나열하고, 기관이 바뀔 때마다 소계 행(연한
    파란 배경), 맨 끝에 전체 합계 행(진한 파란 배경)을 추가한다. 표시 순서는
    institution_distribution과 같은 순서(금액 큰 기관부터)로 맞춘다 — 그 목록에
    없는 기관이 혹시 있으면(데이터 불일치) 누락 없이 뒤에 이어 붙인다."""
    # 컬럼명은 대시보드 표(index.html)ㆍ원본 엑셀 서식과 동일하게 "최초가입일"/"신규일"로 맞춘다.
    headers = ['사업자(기관)', '상품명', '상품유형', '실적배당여부', '금리', '최초가입일', '신규일', '만기일', '금액(원)']
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = _EXPORT_HEADER_FILL
        cell.font = _EXPORT_HEADER_FONT
        cell.alignment = Alignment(horizontal='center')

    by_inst = {}
    for it in items:
        by_inst.setdefault(it.get('institution', ''), []).append(it)
    ordered_insts = [i for i in institution_order if i in by_inst] + [i for i in by_inst if i not in institution_order]

    widths = [len(h) for h in headers]

    def _append_row(row, bold=False, fill=None):
        ws.append(row)
        r = ws.max_row
        for i, v in enumerate(row, start=1):
            cell = ws.cell(row=r, column=i)
            if bold:
                cell.font = _EXPORT_BOLD
            if fill:
                cell.fill = fill
            widths[i - 1] = max(widths[i - 1], len(str(v)) if v is not None else 0)

    grand_total = 0
    for inst in ordered_insts:
        inst_sum = 0
        for it in by_inst[inst]:
            amount = it.get('amount') or 0
            _append_row([
                it.get('institution', ''), it.get('product', ''), it.get('product_type', ''),
                '실적배당' if it.get('is_performance') else '원리금보장',
                f"{it.get('rate', 0) * 100:.2f}%" if it.get('rate') is not None else '-',
                it.get('start') or '-', it.get('new_date') or '-', it.get('end') or '-',
                amount,
            ])
            inst_sum += amount
        _append_row([f'{inst} 소계', '', '', '', '', '', '', '', inst_sum], bold=True, fill=_EXPORT_SUBTOTAL_FILL)
        grand_total += inst_sum

    _append_row(['합계', '', '', '', '', '', '', '', grand_total], bold=True, fill=_EXPORT_TOTAL_FILL)

    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = min(max(w + 2, 10), 45)
    ws.freeze_panes = 'A2'


@app.route('/export.xlsx')
def export_excel():
    """조회 중인 기준일의 퇴직연금 자료를 엑셀로 내려받는다 — 요약ㆍ상세내역(계약별
    기관ㆍ상품ㆍ금리ㆍ만기)ㆍ기관별/상품별/만기별 분포를 시트별로 나눠 담는다.
    base_date를 안 주면 등록된 것 중 가장 최근 기준일을 쓴다."""
    data = _get_data()
    snapshots = data.get('snapshots', {})
    if not snapshots:
        return "등록된 데이터가 없습니다.", 404

    base_date = request.args.get('base_date', '').strip()
    if base_date not in snapshots:
        base_date = sorted(snapshots.keys())[-1]
    snap = snapshots[base_date]

    wb = openpyxl.Workbook()

    ws_summary = wb.active
    ws_summary.title = '요약'
    _export_write_sheet(ws_summary, ['항목', '값'], [
        ['기준일', base_date],
        ['라벨', snap.get('label', '')],
        ['총 적립금(원)', snap.get('total_amount')],
        ['가중평균 수익률', f"{snap.get('weighted_yield', 0) * 100:.2f}%" if snap.get('weighted_yield') is not None else '-'],
        ['예상 연 이자(원)', round(snap.get('expected_interest', 0)) if snap.get('expected_interest') is not None else '-'],
        ['계약 건수', len(snap.get('items', []))],
    ])

    ws_items = wb.create_sheet('상세내역')
    institution_order = [d.get('institution', '') for d in snap.get('institution_distribution', [])]
    _export_write_items_sheet(ws_items, snap.get('items', []), institution_order)

    ws_inst = wb.create_sheet('기관별분포')
    _export_write_sheet(
        ws_inst, ['사업자(기관)', '금액(원)', '비중'],
        [[d.get('institution', ''), d.get('amount'), f"{d.get('ratio', 0) * 100:.2f}%"] for d in snap.get('institution_distribution', [])],
    )

    ws_prod = wb.create_sheet('상품유형별분포')
    _export_write_sheet(
        ws_prod, ['상품유형', '금액(원)', '비중'],
        [[d.get('product_type', ''), d.get('amount'), f"{d.get('ratio', 0) * 100:.2f}%"] for d in snap.get('product_distribution', [])],
    )

    ws_mat = wb.create_sheet('만기별분포')
    _export_write_sheet(
        ws_mat, ['만기 구간', '금액(원)', '비중', '수익률', '예상이자(원)'],
        [
            [
                d.get('category', ''), d.get('amount'), f"{d.get('ratio', 0) * 100:.2f}%",
                f"{d.get('yield', 0) * 100:.2f}%" if d.get('yield') is not None else '-',
                round(d.get('interest', 0)) if d.get('interest') is not None else '-',
            ]
            for d in snap.get('maturity_distribution', [])
        ],
    )

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf, as_attachment=True, download_name=f"퇴직연금_{base_date}.xlsx",
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


@app.route('/admin', methods=['GET', 'POST'])
def admin():
    if request.method == 'POST':
        if request.form.get('password') == pension_config.ADMIN_PASSWORD:
            session['pension_admin'] = True
        else:
            flash('비밀번호가 올바르지 않습니다.', 'error')
        return redirect(url_for('admin'))
    reports = pension_db.list_reports()
    return render_template(
        'pension_admin.html', authed=_is_admin(), reports=reports,
        messages=get_flashed_messages(with_categories=True),
    )


@app.route('/admin/upload', methods=['POST'])
def admin_upload():
    if not _is_admin():
        return redirect(url_for('admin'))

    base_date = request.form.get('base_date', '').strip()
    label = request.form.get('label', '').strip()
    sheet_hint = request.form.get('sheet_hint', '').strip()
    file = request.files.get('excel_file')

    if not base_date or not label or not sheet_hint or not file or not file.filename:
        flash('기준일ㆍ라벨ㆍ시트명 힌트ㆍ엑셀 파일을 모두 입력해주세요.', 'error')
        return redirect(url_for('admin'))
    try:
        datetime.strptime(base_date, '%Y-%m-%d')
    except ValueError:
        flash('기준일 형식이 올바르지 않습니다.', 'error')
        return redirect(url_for('admin'))

    file_bytes = file.read()
    # 저장하기 전에 실제로 파싱되는지 먼저 확인한다 — 이 서식은 시트명/열 위치가
    # 조금만 달라져도 깨지기 쉬워서, 잘못된 파일을 저장해뒀다가 대시보드 전체가
    # 조용히 깨지는 것보다 업로드 시점에 바로 알려주는 편이 낫다.
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        parse_data.parse_sheet_data_from_workbook(wb, sheet_hint, base_date)
    except Exception as e:
        flash(f'파싱 실패: {e} — 시트명 힌트를 확인해주세요.', 'error')
        return redirect(url_for('admin'))

    pension_db.save_report(base_date, label, sheet_hint, file.filename, file_bytes)
    _invalidate_cache()
    flash(f'{label} ({base_date}) 데이터를 저장했습니다.', 'success')
    return redirect(url_for('admin'))


@app.route('/admin/delete/<int:report_id>', methods=['POST'])
def admin_delete(report_id):
    if not _is_admin():
        return redirect(url_for('admin'))
    pension_db.delete_report(report_id)
    _invalidate_cache()
    flash('삭제했습니다.', 'info')
    return redirect(url_for('admin'))


@app.route('/admin/logout', methods=['POST'])
def admin_logout():
    session.pop('pension_admin', None)
    return redirect(url_for('admin'))


@app.route('/')
def index():
    return send_from_directory(BASE_DIR, 'index.html')


@app.route('/<path:filename>')
def static_files(filename):
    return send_from_directory(BASE_DIR, filename)


@app.after_request
def _no_cache_static(response):
    # index.html/app.js/style.css는 자주 고쳐지는데, 브라우저(특히 통합포털 iframe
    # 안에서 열릴 때)가 오래 캐시해서 서버를 재시작해도 화면이 안 바뀌어 보이는
    # 문제가 있었다 — 항상 서버에 다시 확인하도록 캐시를 꺼둔다.
    response.headers['Cache-Control'] = 'no-cache, must-revalidate'
    return response
