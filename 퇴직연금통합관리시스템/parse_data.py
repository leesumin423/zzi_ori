import io
import os
import openpyxl
import json
from datetime import datetime, date

import pension_db

# Paths setup
DATA_DIR = r"C:\Users\tyinc\Desktop\업무리스트 내역 정리\퇴직연금 관련\퇴직연금 관련"
OUTPUT_JSON = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data.json")

# Base dates mapping
BASE_DATES = {
    "2021-12-31": (r"템플릿\퇴직연금 현황양식(주)동양_5개년 데이터.xlsx", "2021. 12월말", "2021년 12월말 (기말)"),
    "2022-12-31": (r"템플릿\퇴직연금 현황양식(주)동양_5개년 데이터.xlsx", "2022. 12월 말", "2022년 12월말 (기말)"),
    "2023-12-31": (r"템플릿\퇴직연금 현황양식(주)동양_5개년 데이터.xlsx", "2023. 12월 말", "2023년 12월말 (기말)"),
    "2024-09-30": (r"과거자료\2024 퇴직연금 현황양식(주)동양_24.10.31 최종.xlsx", "1. 퇴직연금현황(2024.09)", "2024년 9월말 (3분기)"),
    "2024-12-31": (r"템플릿\퇴직연금 현황양식(주)동양_5개년 데이터.xlsx", "2024.12월 말", "2024년 12월말 (기말)"),
    "2025-05-31": (r"과거자료\2024 퇴직연금 현황양식(주)동양_25. 5월 말기준.xlsx", "1. 퇴직연금현황(2025.5월 말)", "2025년 5월말"),
    "2025-09-30": (r"과거자료\퇴직연금 현황양식(주)동양_25. 9월 말기준.xlsx", "1. 퇴직연금현황(2025.9월 말)", "2025년 9월말 (3분기)"),
    "2025-10-31": (r"과거자료\퇴직연금 현황양식(주)동양_25.10월 말기준.xlsx", "1. 퇴직연금현황(2025.10월 말)", "2025년 10월말"),
    "2025-12-31": (r"템플릿\퇴직연금 현황양식(주)동양_5개년 데이터.xlsx", "2025.12월 말", "2025년 12월말 (기말)"),
    "2026-05-31": (r"퇴직연금 현황양식(주)동양_26.5월 말기준_유진260622 회신.xlsx", "1. 퇴직연금현황(2026.05월 말", "2026년 5월말"),
    "2026-06-30": (r"퇴직연금 현황양식(주)동양_26.6월 말기준.xlsx", "1. 퇴직연금현황(2026.06월 말", "2026년 6월말 (반기)"),
}

def clean_inst_name(name):
    if not name:
        return ""
    name = str(name).strip().replace(" ", "")
    if "농협" in name: return "농협"
    if "산업" in name: return "산업은행"
    if "우리" in name: return "우리은행"
    if "기업" in name: return "기업은행"
    if "국민" in name or "KB" in name: return "국민은행"
    if "신한" in name: return "신한은행"
    if "im뱅크" in name.lower() or "대구은행" in name: return "iM뱅크"
    if "하나" in name: return "하나은행"
    if "현대" in name: return "현대해상"
    if "삼성생명" in name or ("삼성" in name and "생명" in name): return "삼성생명"
    return name.split("/")[0].split()[0]

def classify_product(name):
    """상품명을 기반으로 상품 유형 분류"""
    if not name:
        return "현금/기타"
    name_clean = str(name).strip().replace(" ", "")
    # 실적배당형 (펀드/OCIO)
    if "펀드" in name_clean or "OCIO" in name_clean or "자산운용" in name_clean or "실적배당" in name_clean:
        return "실적배당형(펀드)"
    if "정기예금" in name_clean or "예금" in name_clean:
        return "정기예금"
    if "GIC" in name_clean or "이율보증" in name_clean or "이율보장" in name_clean or "보증보험" in name_clean:
        return "GIC"
    if "ELB" in name_clean or "DLB" in name_clean or "파생결합" in name_clean:
        return "ELB/DLB"
    if "고유계정" in name_clean or "현금" in name_clean or "계정대" in name_clean:
        return "현금/기타"
    return "현금/기타"

def is_performance_based(name):
    """실적배당형 상품 여부 확인"""
    if not name:
        return False
    name_clean = str(name).strip().replace(" ", "")
    return any(k in name_clean for k in ["펀드", "OCIO", "자산운용", "실적배당"])

def parse_date_value(val):
    if val is None:
        return None
    if isinstance(val, (datetime, date)):
        return val.date() if isinstance(val, datetime) else val
    val_str = str(val).strip()
    for fmt in ["%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%Y.%m.%d", "%y.%m.%d"]:
        try:
            return datetime.strptime(val_str, fmt).date()
        except ValueError:
            pass
    return None

def parse_sheet_data(file_path, sheet_name, base_date_str):
    """파일 경로에서 워크북을 열어 파싱한다 — BASE_DATES(레거시 하드코딩 파일)용."""
    wb = openpyxl.load_workbook(file_path, data_only=True)
    return parse_sheet_data_from_workbook(wb, sheet_name, base_date_str)

def parse_sheet_data_from_workbook(wb, sheet_name, base_date_str):
    """이미 열린 워크북에서 파싱한다 — pension_db에 등록된 업로드 파일(bytes)용으로
    분리했다. 실제 파싱 로직은 레거시 파일이든 업로드 파일이든 완전히 동일하다."""
    # 시트 이름 부분 일치 허용
    actual_sheet = None
    for sn in wb.sheetnames:
        if sheet_name in sn or sn in sheet_name:
            actual_sheet = sn
            break
    if actual_sheet is None:
        # 정확히 일치하는 시트 없을 때 첫 번째 시트 시도
        raise KeyError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
    
    ws = wb[actual_sheet]
    base_date = datetime.strptime(base_date_str, "%Y-%m-%d").date()
    
    col_headers = {}
    for r in [3, 4]:
        for c in range(1, 15):
            val = ws.cell(r, c).value
            if val:
                col_headers[str(val).strip().replace(" ", "")] = c
                
    rate_col = col_headers.get("금리", 5)
    amount_col = col_headers.get("금액", 8)
    if "평가금액" in col_headers:
        amount_col = col_headers["평가금액"]
    start_col = col_headers.get("최초가입", 6)
    new_date_col = col_headers.get("신규일", None)
    end_col = col_headers.get("만기", 7)
    
    current_cat = ""
    current_inst = ""
    current_prod = ""
    
    items = []
    
    for r in range(5, 120):
        val_b = ws.cell(r, 2).value
        val_c = ws.cell(r, 3).value
        val_d = ws.cell(r, 4).value
        
        # Stop at Grand Total
        if val_b and any(x in str(val_b).replace(" ", "") for x in ["합계"]):
            break
        if val_c and any(x in str(val_c).replace(" ", "") for x in ["합계"]):
            break
            
        # Skip Subtotal
        if val_c and any(x in str(val_c).replace(" ", "") for x in ["소계"]):
            continue
        if val_b and any(x in str(val_b).replace(" ", "") for x in ["소계"]):
            continue
            
        if val_b:
            current_cat = str(val_b).strip()
        if val_c:
            current_inst = str(val_c).strip()
        if val_d:
            current_prod = str(val_d).strip()
            
        rate = ws.cell(r, rate_col).value
        amount = ws.cell(r, amount_col).value
        start_date_val = ws.cell(r, start_col).value
        new_date_val = ws.cell(r, new_date_col).value if new_date_col else None
        end_date_val = ws.cell(r, end_col).value
        
        if amount is not None and isinstance(amount, (int, float)) and amount > 0:
            # 실적배당형 상품: 금리가 없는 경우 분기 수익률 기반 추정
            prod_type = classify_product(current_prod)
            if is_performance_based(current_prod):
                # 실적배당형: 금리 셀 값이 있으면 그대로 사용 (분기 수익률로 입력된 값)
                # 분기 수익률이 입력된 경우 연환산 (×4) 적용
                if rate is not None and isinstance(rate, (int, float)) and rate > 0:
                    # 값이 연율인지 분기율인지 판단: 0.1 초과이면 분기율로 판단
                    if rate > 0.1:
                        rate_float = rate * 4  # 분기 수익률 → 연환산
                    else:
                        rate_float = float(rate)  # 이미 연율
                else:
                    rate_float = 0.0
            else:
                rate_float = float(rate) if rate is not None and isinstance(rate, (int, float)) else 0.0
            
            start_date = parse_date_value(start_date_val)
            new_date = parse_date_value(new_date_val)
            end_date = parse_date_value(end_date_val)
            
            items.append({
                "category": current_cat,
                "institution": clean_inst_name(current_inst),
                "product": current_prod,
                "product_type": prod_type,
                "is_performance": is_performance_based(current_prod),
                "rate": rate_float,
                "start": start_date.strftime("%Y-%m-%d") if start_date else None,
                "new_date": new_date.strftime("%Y-%m-%d") if new_date else None,
                "end": end_date.strftime("%Y-%m-%d") if end_date else None,
                "amount": int(round(amount))
            })
            
    return items

def analyze_snapshot(items, base_date_str, label):
    base_date = datetime.strptime(base_date_str, "%Y-%m-%d").date()
    total_amount = sum(x["amount"] for x in items)
    
    # 연 이자: 금리 × 금액, 정수로 반올림
    expected_interest_raw = sum(x["rate"] * x["amount"] for x in items)
    expected_interest = int(round(expected_interest_raw))
    
    weighted_yield = expected_interest_raw / total_amount if total_amount > 0 else 0.0
    
    # 1. Maturity Distribution
    maturity_items = {
        "1년 이내": [],
        "1년이상 2년미만": [],
        "2년이상 3년미만": []
    }
    
    for x in items:
        end_date = parse_date_value(x["end"])
        if end_date is None:
            cat = "1년 이내"
        else:
            diff_days = (end_date - base_date).days
            if diff_days <= 365:
                cat = "1년 이내"
            elif diff_days <= 730:
                cat = "1년이상 2년미만"
            else:
                cat = "2년이상 3년미만"
        maturity_items[cat].append(x)
        
    maturity_distribution = []
    for cat in ["1년 이내", "1년이상 2년미만", "2년이상 3년미만"]:
        cat_list = maturity_items[cat]
        cat_amt = sum(x["amount"] for x in cat_list)
        cat_int = sum(x["rate"] * x["amount"] for x in cat_list)
        cat_yield = cat_int / cat_amt if cat_amt > 0 else 0.0
        maturity_distribution.append({
            "category": cat,
            "amount": int(round(cat_amt)),
            "ratio": cat_amt / total_amount if total_amount > 0 else 0.0,
            "yield": cat_yield,
            "interest": int(round(cat_int))
        })
        
    # 2. Institution Distribution
    inst_sums = {}
    for x in items:
        inst = x["institution"]
        inst_sums[inst] = inst_sums.get(inst, 0) + x["amount"]
        
    institution_distribution = []
    for inst, amt in sorted(inst_sums.items(), key=lambda v: v[1], reverse=True):
        institution_distribution.append({
            "institution": inst,
            "amount": int(round(amt)),
            "ratio": amt / total_amount if total_amount > 0 else 0.0
        })
        
    # 3. Product Distribution (updated categories)
    prod_sums = {}
    for x in items:
        p_type = x.get("product_type", classify_product(x["product"]))
        prod_sums[p_type] = prod_sums.get(p_type, 0) + x["amount"]
        
    product_distribution = []
    for p_type in ["GIC", "정기예금", "ELB/DLB", "실적배당형(펀드)", "현금/기타"]:
        amt = prod_sums.get(p_type, 0)
        if amt > 0:
            product_distribution.append({
                "product_type": p_type,
                "amount": int(round(amt)),
                "ratio": amt / total_amount if total_amount > 0 else 0.0
            })
        
    return {
        "base_date": base_date_str,
        "label": label,
        "total_amount": int(round(total_amount)),
        "weighted_yield": weighted_yield,
        "expected_interest": expected_interest,
        "maturity_distribution": maturity_distribution,
        "institution_distribution": institution_distribution,
        "product_distribution": product_distribution,
        "items": items
    }

def _parse_legacy_snapshots():
    """parse_data.py에 하드코딩된 BASE_DATES(과거 연도 등, 관리자가 다시 업로드할
    일이 거의 없는 자료)를 파싱한다."""
    snapshots_data = {}
    for base_date_str, (rel_path, sheet_name, label) in BASE_DATES.items():
        file_path = os.path.join(DATA_DIR, rel_path)
        if not os.path.exists(file_path):
            print(f"  Warning: File not found {file_path}. Skipping.")
            continue
        try:
            items = parse_sheet_data(file_path, sheet_name, base_date_str)
            analysis = analyze_snapshot(items, base_date_str, label)
            snapshots_data[base_date_str] = analysis
            print(f"  Parsed {base_date_str} OK | 총액: {analysis['total_amount']:,}원 | 수익률: {analysis['weighted_yield']*100:.3f}% | 예상이자: {analysis['expected_interest']:,}원")
        except Exception as e:
            print(f"  Error parsing {base_date_str}: {e}")
    return snapshots_data


def _parse_db_snapshots():
    """pension_db에 등록된(관리자가 화면에서 업로드한) 기준일 엑셀을 파싱한다 —
    같은 기준일이 BASE_DATES에도 있으면 이쪽이 우선한다(코드 수정 없이 화면에서
    갱신할 수 있게)."""
    snapshots_data = {}
    for row in pension_db.list_reports():
        report_id, base_date_str, label, sheet_hint, filename, _uploaded_at, _size = row
        try:
            file_bytes = pension_db.get_report_bytes(report_id)
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
            items = parse_sheet_data_from_workbook(wb, sheet_hint, base_date_str)
            analysis = analyze_snapshot(items, base_date_str, label)
            snapshots_data[base_date_str] = analysis
            print(f"  Parsed (DB) {base_date_str} OK | 총액: {analysis['total_amount']:,}원 | 수익률: {analysis['weighted_yield']*100:.3f}% | 예상이자: {analysis['expected_interest']:,}원")
        except Exception as e:
            print(f"  Error parsing DB report {base_date_str} ({filename}): {e}")
    return snapshots_data


def _compute_yoy_years(all_dates):
    """연말 기준일(YYYY-12-31)이 있는 해는 전부 넣고, 그중 없는 가장 최근 기준일(예:
    반기말)도 마지막에 하나 추가한다 — 새 기준일이 계속 추가돼도 코드를 손대지
    않아도 되도록 하드코딩 목록 대신 실제 존재하는 날짜에서 계산한다."""
    if not all_dates:
        return []
    years = sorted({int(d.split('-')[0]) for d in all_dates})
    yoy = [f"{y}-12-31" for y in years if f"{y}-12-31" in all_dates]
    latest = max(all_dates)
    if latest not in yoy:
        yoy.append(latest)
    return yoy


def build_snapshots_and_yoy():
    """레거시 BASE_DATES + pension_db 업로드분을 합쳐 최종 data.json 내용을 만든다.
    같은 기준일이 둘 다에 있으면 DB(화면에서 업로드한 것)가 이긴다."""
    print("Parsing all snapshots...")
    snapshots_data = _parse_legacy_snapshots()
    snapshots_data.update(_parse_db_snapshots())

    yoy_years = _compute_yoy_years(list(snapshots_data.keys()))
    yoy_comparison = []
    
    for i, date_str in enumerate(yoy_years):
        snap = snapshots_data.get(date_str)
        if not snap:
            continue
            
        year = int(date_str.split("-")[0])
        total_amount = snap["total_amount"]
        weighted_yield = snap["weighted_yield"]
        
        growth_amount = 0
        growth_rate = 0.0
        yield_change = 0.0
        
        if i > 0:
            prev_date_str = yoy_years[i-1]
            prev_snap = snapshots_data.get(prev_date_str)
            if prev_snap:
                prev_amount = prev_snap["total_amount"]
                prev_yield = prev_snap["weighted_yield"]
                growth_amount = total_amount - prev_amount
                growth_rate = growth_amount / prev_amount if prev_amount > 0 else 0.0
                yield_change = (weighted_yield - prev_yield) * 100
                
        yoy_comparison.append({
            "year": year,
            "date": date_str,
            "label": snap["label"],
            "total_amount": total_amount,
            "weighted_yield": weighted_yield,
            "expected_interest": snap["expected_interest"],
            "growth_amount": int(round(growth_amount)),
            "growth_rate": growth_rate,
            "yield_change": yield_change
        })
        
    return {
        "snapshots": snapshots_data,
        "yoy_comparison": yoy_comparison
    }


def main():
    """CLI 진입점 — data.json을 한 번 만들어 디스크에 써둔다(수동 재생성용). 서버
    (pension_server.py)는 이 함수를 쓰지 않고 build_snapshots_and_yoy()를 요청마다
    직접 호출해 업로드가 즉시 반영되게 한다."""
    output_data = build_snapshots_and_yoy()
    with open(OUTPUT_JSON, "w", encoding="utf-8") as f:
        json.dump(output_data, f, ensure_ascii=False, indent=2)
    print(f"\nSuccessfully wrote unified pension data to {OUTPUT_JSON}")
    print(f"Total snapshots: {len(output_data['snapshots'])}")

if __name__ == "__main__":
    main()
