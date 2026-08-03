import streamlit as st
import pandas as pd
import requests
import plotly.graph_objects as go
import json
import os
import io
import subprocess
from pathlib import Path
import openpyxl
import calendar
import re
from datetime import datetime, timedelta
import snapshot_db
import monthly_report_db
import report_shared
import ancillary_db
import group_loan
import group_loan_db
import real_estate_db
import interest_history_db
 
st.set_page_config(page_title="(주)동양 차입관리 통합프로그램", page_icon="📈", layout="wide")
 
API_KEY = "BFT9S9JKSLG23WABEF0N"
 
# ---------------------------------------------------------------------------
# 연도는 더 이상 하드코딩하지 않고, 실행 시점의 현재 날짜를 기준으로 자동 계산합니다.
# 예) 오늘이 2026년이면 CURRENT_YEAR=2026(당해년도), PREV_YEAR=2025(전년도)
# ---------------------------------------------------------------------------
CURRENT_YEAR = datetime.now().year
PREV_YEAR = CURRENT_YEAR - 1
 
# 과거 이자율/신용등급 추이는 최근 4개년(전전전년~당해년)을 기준으로 봅니다.
HIST_YEAR_START = CURRENT_YEAR - 3
HIST_YEAR_END = CURRENT_YEAR
YEAR_COLS = [f"{y}년" for y in range(HIST_YEAR_START, HIST_YEAR_END + 1)]  # 엑셀 시트의 실제 연도 컬럼명과 일치해야 함
 
# KPI 입력값은 로컬(프로그램 실행 폴더)에 저장/로드합니다. (엑셀 데이터와 별개)
KPI_JSON_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "kpi_inputs.json")

# 1년이내 만기도래 상환계획표에서 사용자가 직접 입력하는 '상환여부'/'비고' 값을
# 로컬에 저장해두어, 엑셀을 다시 업로드해도 입력 내용이 유지되도록 합니다.
REPAYMENT_STATUS_JSON_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "repayment_status.json")

# ---------------------------------------------------------------------------
# 담보물건지 ↔ 품의서 자동 연결 (그룹웨어 크롤링)
# 별도 프로젝트(groupware-task-checklist)의 Playwright 기반 그룹웨어 로그인/스캔
# 인프라를 subprocess로 호출한다 — 이 앱(Streamlit venv)에는 playwright가 없고,
# 그룹웨어 세션 로그인은 사용자가 직접 브라우저로 한 번 해둬야 하는 별개의 설정
# (login_setup.py)이 필요해서, 그 프로젝트를 그대로 재사용하는 구조다.
#
# 실제로 설치/로그인 설정을 완료해서 쓰고 있는 곳은 C:\Users\tyinc\-AI-\
# groupware-task-checklist 이고(.venv/.env/.state 세션이 여기 있음), ai 개발관련
# 폴더 아래에도 소스코드 사본이 있지만 그건 설치가 안 된 사본이다 — 다른 PC/경로로
# 옮기면 환경변수 GW_PROJECT_DIR로 덮어쓸 수 있다.
# ---------------------------------------------------------------------------
GW_PROJECT_DIR = Path(os.getenv("GW_PROJECT_DIR", r"C:\Users\tyinc\-AI-\groupware-task-checklist"))
GW_VENV_PYTHON = GW_PROJECT_DIR / ".venv" / "Scripts" / "python.exe"


def _collateral_doc_search_ready() -> bool:
    return GW_VENV_PYTHON.exists() and (GW_PROJECT_DIR / ".env").exists()


def run_collateral_doc_search(site_name: str, keywords=None, timeout_sec: int = 180):
    """link_collateral_doc.py를 별도 프로세스로 실행해 품의서 링크+첨부파일을
    real_estate_db에 저장한다. (ok, message) 튜플을 반환 — ok=False면 message에
    사용자에게 보여줄 안내문(설정 필요/오류/못 찾음 등)이 들어있다."""
    if not GW_VENV_PYTHON.exists():
        return False, (
            "그룹웨어 검색 기능이 아직 설정되지 않았습니다. "
            f"`{GW_PROJECT_DIR}` 폴더에서 최초 1회 설치(`1_설치.bat`)와 로그인 설정(`2_로그인설정.bat`)을 "
            "먼저 진행해주세요."
        )
    if not (GW_PROJECT_DIR / ".env").exists():
        return False, "`.env` 파일이 없습니다 — `2_로그인설정.bat` 안내에 따라 완료함 URL을 설정해주세요."

    cmd = [str(GW_VENV_PYTHON), "link_collateral_doc.py", "--site-name", site_name]
    if keywords:
        cmd += ["--keywords", *keywords]
    try:
        result = subprocess.run(
            cmd, cwd=str(GW_PROJECT_DIR), capture_output=True, text=True,
            encoding="utf-8", errors="replace", timeout=timeout_sec,
        )
    except subprocess.TimeoutExpired:
        return False, "그룹웨어 검색이 시간 내에 끝나지 않았습니다(로그인 세션 만료 가능성) — 다시 시도해보세요."
    except Exception as e:
        return False, f"검색 실행 중 오류: {e}"

    output = (result.stdout or "") + (("\n" + result.stderr) if result.stderr else "")
    if result.returncode != 0:
        return False, output.strip() or "검색이 실패했습니다(원인 불명) — logs 폴더를 확인해주세요."
    return True, output.strip()


# ---------------------------------------------------------------------------
# KPI 입력값 로드/마이그레이션
# ---------------------------------------------------------------------------
def _default_kpi_values():
    return {
        'balance_prev': 332875.0,
        'avg_rate_prev': 4.12,
        'bok_rate_prev': 2.65,
        'balance_curr': 332875.0,
        'avg_rate_curr': 4.12,
        'bok_rate_curr': 2.65,
    }
 
 
def _migrate_legacy_kpi_keys(data):
    """예전 버전은 _24/_25/_26 처럼 연도가 이름에 박혀 있었습니다.
    새 버전은 prev(전년)/curr(당해년) 이라는 상대 개념을 사용하므로,
    예전 JSON 파일이 남아있어도 자동으로 값을 이어받도록 매핑합니다.
    (예전 25=전년도였던 값 -> prev, 예전 26=당해년도였던 값 -> curr)
    """
    required_new = ['balance_prev', 'avg_rate_prev', 'bok_rate_prev',
                     'balance_curr', 'avg_rate_curr', 'bok_rate_curr']
    if all(k in data for k in required_new):
        return {k: data[k] for k in required_new}
 
    legacy_25_26 = ['balance_25', 'avg_rate_25', 'bok_rate_25',
                     'balance_26', 'avg_rate_26', 'bok_rate_26']
    if all(k in data for k in legacy_25_26):
        return {
            'balance_prev': data['balance_25'],
            'avg_rate_prev': data['avg_rate_25'],
            'bok_rate_prev': data['bok_rate_25'],
            'balance_curr': data['balance_26'],
            'avg_rate_curr': data['avg_rate_26'],
            'bok_rate_curr': data['bok_rate_26'],
        }
 
    legacy_24_25 = ['balance_24', 'avg_rate_24', 'bok_rate_24',
                     'balance_25', 'avg_rate_25', 'bok_rate_25']
    if all(k in data for k in legacy_24_25):
        return {
            'balance_prev': data['balance_24'],
            'avg_rate_prev': data['avg_rate_24'],
            'bok_rate_prev': data['bok_rate_24'],
            'balance_curr': data['balance_25'],
            'avg_rate_curr': data['avg_rate_25'],
            'bok_rate_curr': data['bok_rate_25'],
        }
    return None
 
 
def load_kpi_values():
    if os.path.exists(KPI_JSON_FILE):
        try:
            with open(KPI_JSON_FILE, 'r', encoding='utf-8') as f:
                data = json.load(f)
            migrated = _migrate_legacy_kpi_keys(data)
            if migrated is not None:
                return migrated
        except Exception:
            pass
    return _default_kpi_values()
 
 
def save_kpi_values(values):
    with open(KPI_JSON_FILE, 'w', encoding='utf-8') as f:
        json.dump(values, f, indent=4, ensure_ascii=False)
 
 
# ---------------------------------------------------------------------------
# 한국은행 ECOS 기준금리 조회
# ---------------------------------------------------------------------------
@st.cache_data(ttl=3600, show_spinner=False)
def fetch_ecos_base_rate():
    """한국은행 ECOS API에서 기준금리 조회 (최근 4개년, 연도는 자동 계산).

    722Y001은 연간(A) 주기 조회 자체는 정상 동작하지만, 그건 '이미 끝난 해'까지만
    존재한다 — 당해년도(아직 12월이 안 지난 해)는 연간 집계가 아직 없어서 결과에서
    통째로 빠진다. 그래서 당해년도만 별도로 최근 일별(D) 값을 조회해서 채워 넣는다
    (실제로 이 문제 때문에 2026년 기준금리가 항상 빠져서, 그래프/스프레드 계산에서
    코드 곳곳의 임시 기본값(3.0% 등)으로 잘못 대체되고 있었다)."""
    url = (f"https://ecos.bok.or.kr/api/StatisticSearch/{API_KEY}/json/kr/1/100/"
           f"722Y001/A/{HIST_YEAR_START}/{HIST_YEAR_END}/0101000/")

    fallback = {
        str(HIST_YEAR_START): 3.50,
        str(HIST_YEAR_START + 1): 3.25,
        str(HIST_YEAR_START + 2): 3.00,
        str(HIST_YEAR_END): 2.75,
    }
    try:
        response = requests.get(url, timeout=5)
        data = response.json()

        if "StatisticSearch" in data:
            rows = data["StatisticSearch"]["row"]
            result = {}
            for row in rows:
                result[row["TIME"]] = float(row["DATA_VALUE"])
        else:
            result = dict(fallback)
    except Exception:
        result = dict(fallback)

    # 당해년도는 연간 집계가 아직 없을 수 있으니, 최근 30일 일별 값 중 가장 최신 것으로 채운다.
    current_year_key = str(HIST_YEAR_END)
    if current_year_key not in result:
        today = datetime.now().date()
        start_d = (today - timedelta(days=30)).strftime('%Y%m%d')
        end_d = today.strftime('%Y%m%d')
        try:
            url_d = (f"https://ecos.bok.or.kr/api/StatisticSearch/{API_KEY}/json/kr/1/40/"
                     f"722Y001/D/{start_d}/{end_d}/0101000/")
            resp_d = requests.get(url_d, timeout=5)
            rows_d = resp_d.json().get("StatisticSearch", {}).get("row", [])
            if rows_d:
                latest = max(rows_d, key=lambda r: r["TIME"])
                result[current_year_key] = float(latest["DATA_VALUE"])
        except Exception:
            pass

    return result


@st.cache_data(ttl=3600, show_spinner=False)
def fetch_base_rate_change_this_year():
    """당해년도 안에서 한국은행 기준금리가 실제로 조정된 시점을 찾는다(있으면).
    최근 60일 일별 값을 조회해, 그 구간 안에서 서로 다른 값이 나오면 '조정이 있었다'로
    판단해 (조정일, 이전 금리, 이후 금리) 튜플을 반환한다. 변동이 없거나 조회 실패 시 None.
    """
    today = datetime.now().date()
    start_d = (today - timedelta(days=60)).strftime('%Y%m%d')
    end_d = today.strftime('%Y%m%d')
    try:
        url_d = (f"https://ecos.bok.or.kr/api/StatisticSearch/{API_KEY}/json/kr/1/70/"
                 f"722Y001/D/{start_d}/{end_d}/0101000/")
        resp_d = requests.get(url_d, timeout=5)
        rows_d = resp_d.json().get("StatisticSearch", {}).get("row", [])
        if not rows_d:
            return None
        rows_d.sort(key=lambda r: r["TIME"])
        values = [(r["TIME"], float(r["DATA_VALUE"])) for r in rows_d]
        for i in range(1, len(values)):
            if values[i][1] != values[i - 1][1]:
                change_date, new_rate = values[i]
                old_rate = values[i - 1][1]
                return (change_date, old_rate, new_rate)
        return None
    except Exception:
        return None
 
 
# ---------------------------------------------------------------------------
# ECOS 시장금리(벤치마크) 자동 조회 — 코리보/CD/금융채(AAA) 등
#
# 항목코드(ITEM_CODE)를 하드코딩하지 않고, ECOS의 StatisticItemList API로
# 통계표(817Y002, 시장금리 일별)의 세부 항목 목록을 직접 조회한 뒤
# 이름(키워드)으로 매칭해서 찾는다. ECOS 홈페이지에서 항목코드가 바뀌거나
# 우리가 잘못 추측하더라도, 실제 항목명 목록을 뒤져서 찾기 때문에 더 안전하다.
# ---------------------------------------------------------------------------
ECOS_MARKET_STAT_CODE = "817Y002"  # 시장금리(일별) — 콜금리, CD, 코리보, 금융채, 국고채 등

# 대출 '금리조건' 원문 텍스트에서 벤치마크 종류를 인식하기 위한 키워드
# -> ECOS 항목명 검색에 쓸 키워드 (여러 개면 먼저 매칭되는 것을 사용)
BENCHMARK_ITEM_KEYWORDS = {
    "코리보":  ["코리보", "KORIBOR"],
    "CD":     ["CD수익률", "CD유통수익률", "CD(91일)", "CD"],
    "금융채":  ["금융채(AAA", "금융채"],
    "은행채":  ["은행채(AAA", "은행채"],
    "중금채":  ["중소기업금융채", "중금채"],
    "MOR":    ["MOR", "자금조달비용지수"],
}


@st.cache_data(ttl=3600, show_spinner=False)
def fetch_ecos_item_list(stat_code):
    """ECOS StatisticItemList: 통계표의 세부 항목(ITEM_CODE/ITEM_NAME) 목록 조회."""
    url = f"https://ecos.bok.or.kr/api/StatisticItemList/{API_KEY}/json/kr/1/300/{stat_code}"
    try:
        response = requests.get(url, timeout=5)
        data = response.json()
        return data.get("StatisticItemList", {}).get("row", [])
    except Exception:
        return []


@st.cache_data(ttl=3600, show_spinner=False)
def resolve_benchmark_item_code(bench_key):
    """벤치마크 이름(코리보/CD/금융채 등) -> (ITEM_CODE, 실제 ECOS 항목명) 탐색.
    못 찾으면 (None, None)."""
    items = fetch_ecos_item_list(ECOS_MARKET_STAT_CODE)
    keywords = BENCHMARK_ITEM_KEYWORDS.get(bench_key, [bench_key])
    for kw in keywords:
        for it in items:
            name = it.get("ITEM_NAME", "")
            if kw in name:
                return it.get("ITEM_CODE"), name
    return None, None


@st.cache_data(ttl=3600, show_spinner=False)
def fetch_ecos_benchmark_series(bench_key):
    """특정 벤치마크의 연도별(연말 기준) 시계열을 ECOS에서 조회.
    반환: {'2023': 3.71, '2024': 3.20, ...} 형태. 조회 실패/미매칭 시 {}.

    817Y002(시장금리)는 '일별' 통계표라 주기를 A(연간)로 바로 조회하면 ECOS가
    "해당하는 데이터가 없습니다"를 돌려준다(실제로 이 문제로 벤치마크 대비 스프레드가
    전 건 '-'로만 나오고 있었다 — 기준금리 대비 1차 지표만 표시되고 2차 지표가 항상
    빠져 있던 원인). 그래서 연도마다 D(일별) 주기로 연말(12/31) 근처 날짜 구간을 따로
    조회해, 그 구간 안에서 가장 최근(=연말에 가장 가까운) 값을 그 해의 대표값으로 쓴다.
    당해년도(아직 12월이 안 지난 해)는 12/31 데이터가 없을 수 있어, 대신 오늘까지의
    최근 10일 구간에서 가장 최근 값을 쓴다.
    """
    item_code, item_name = resolve_benchmark_item_code(bench_key)
    if not item_code:
        return {}, None

    today = datetime.now().date()
    result = {}
    for year in range(HIST_YEAR_START, HIST_YEAR_END + 1):
        if year < today.year:
            start_d, end_d = f"{year}1220", f"{year}1231"
        else:
            # 당해년도: 아직 12월말 데이터가 없을 수 있어 오늘까지 최근 10일 구간을 대신 조회
            end_d = today.strftime('%Y%m%d')
            start_d = (today - timedelta(days=10)).strftime('%Y%m%d')
        url = (f"https://ecos.bok.or.kr/api/StatisticSearch/{API_KEY}/json/kr/1/20/"
               f"{ECOS_MARKET_STAT_CODE}/D/{start_d}/{end_d}/{item_code}")
        try:
            response = requests.get(url, timeout=5)
            rows = response.json().get("StatisticSearch", {}).get("row", [])
            if rows:
                latest = max(rows, key=lambda r: r["TIME"])
                result[str(year)] = float(latest["DATA_VALUE"])
        except Exception:
            continue
    return result, item_name


def parse_benchmark_key(condition_text):
    """대출 '금리조건' 원문(예: 'KORIBOR + 1.60%', '3개월 CD변동 + 1.69%',
    '금융채 1년 + 1.90%', '중금채 1년 + 1.44%', '1년 MOR + 1.75%')에서
    벤치마크 종류를 추출. 못 알아보면 None."""
    text = str(condition_text).upper()
    if not text or text == "-" or text == "NAN":
        return None
    if "코리보" in text or "KORIBOR" in text:
        return "코리보"
    if "MOR" in text:
        return "MOR"
    if "CD" in text:
        return "CD"
    if "중금채" in text or "중소기업금융채" in text:
        return "중금채"
    if "은행채" in text:
        return "은행채"
    if "금융채" in text:
        return "금융채"
    return None


# ---------------------------------------------------------------------------
# 업로드된 엑셀 파일 파싱 (경로가 아닌 업로드 바이트를 입력으로 받음)
# ---------------------------------------------------------------------------
@st.cache_data(ttl=60, show_spinner=False)
def load_excel_data(excel_bytes):
    try:
        xl = pd.ExcelFile(io.BytesIO(excel_bytes))
 
        df_perf = None
        perf_sheet_candidates = [s for s in xl.sheet_names if '성과표' in s]
        if perf_sheet_candidates:
            df_perf = xl.parse(perf_sheet_candidates[0])
 
        saving_amount = 0
        avg_rate_24 = 0
        avg_rate_25 = 0
        saving_rate = 0
        total_balance = 0
 
        if df_perf is not None:
            for idx, row in df_perf.iterrows():
                key_str = str(row.iloc[0]).strip()
                col1 = row.iloc[1]
                col2 = row.iloc[2]
 
                if str(PREV_YEAR) in key_str and '차입잔액' not in key_str:
                    avg_rate_24 = pd.to_numeric(col2, errors='coerce')
                elif (str(CURRENT_YEAR) in key_str or '현재' in key_str) and '차입잔액' not in key_str:
                    total_balance = pd.to_numeric(col1, errors='coerce')
                    avg_rate_25 = pd.to_numeric(col2, errors='coerce')
                elif '금리절감효과' in key_str:
                    saving_rate = pd.to_numeric(col1, errors='coerce')
                elif '계(백만원)' in key_str or '계(' in key_str:
                    saving_amount = pd.to_numeric(col1, errors='coerce')
                elif '절감금리' in key_str:
                    if pd.isna(saving_rate) or saving_rate == 0:
                        saving_rate = pd.to_numeric(col1, errors='coerce')
 
        kpi_data = {
            "saving_amount": saving_amount,
            "avg_rate_24": avg_rate_24,
            "avg_rate_25": avg_rate_25,
            "saving_rate": saving_rate,
            "total_balance": total_balance
        }
 
        hist_sheet_candidates = [s for s in xl.sheet_names if '이자율변동' in s or '과거' in s]
        if hist_sheet_candidates:
            # 이 시트는 같은 은행의 여러 차입건을 표시할 때 은행명(B열) 셀을 병합해서
            # 한 번만 적어두는 경우가 많다 — pandas로 그냥 읽으면 병합된 나머지 행은
            # 은행명이 빈 값(NaN)으로 보여서 아래 "은행명 없으면 skip" 로직에 걸려
            # 그 차입건이 통째로 화면에서 빠져버린다(실제로 NH농협은행ㆍ우리은행ㆍ
            # 산업은행 등 일부 차입건이 이렇게 누락되고 있었다). load_current_loan_status와
            # 똑같이 openpyxl로 먼저 병합을 풀고 값을 전파한 뒤 읽는다.
            wb_hist = openpyxl.load_workbook(io.BytesIO(excel_bytes), data_only=True)
            ws_hist = wb_hist[hist_sheet_candidates[0]]
            for merged_range in list(ws_hist.merged_cells.ranges):
                min_col, min_row, max_col, max_row = merged_range.bounds
                top_left_value = ws_hist.cell(row=min_row, column=min_col).value
                ws_hist.unmerge_cells(str(merged_range))
                for r in range(min_row, max_row + 1):
                    for c in range(min_col, max_col + 1):
                        ws_hist.cell(row=r, column=c).value = top_left_value
            df_hist = pd.DataFrame(ws_hist.values)

            header_idx = -1
            for i in range(len(df_hist)):
                if '금융기관' in str(df_hist.iloc[i, 1]):
                    header_idx = i
                    break
 
            history_list = []
            if header_idx != -1:
                current_loan_type = ""
                for i in range(header_idx + 1, len(df_hist)):
                    row = df_hist.iloc[i]
                    bank_name = str(row.iloc[1]).strip()
                    if 'nan' in bank_name.lower() or not bank_name:
                        continue
                    if '기준시점' in bank_name or f'{PREV_YEAR}년말' in bank_name:
                        break
                    if bank_name == '계':
                        continue
 
                    # Propagate loan type
                    loan_type_val = row.iloc[2]
                    if pd.notna(loan_type_val) and str(loan_type_val).strip() != "" and str(loan_type_val).strip() != "0":
                        current_loan_type = str(loan_type_val).strip().replace('\n', ' ')
 
                    # Get loan limit
                    limit_val = pd.to_numeric(row.iloc[3], errors='coerce')
                    limit_str = ""
                    if pd.notna(limit_val) and limit_val > 0:
                        limit_str = f", {limit_val/100000000:.0f}억"
 
                    # Unique display name
                    display_name = bank_name
                    if current_loan_type:
                        display_name = f"{bank_name} ({current_loan_type}{limit_str})"
                    elif limit_str:
                        display_name = f"{bank_name} ({limit_str.replace(', ', '')})"
 
                    def parse_rate(val):
                        v = pd.to_numeric(val, errors='coerce')
                        if pd.isna(v):
                            return 0
                        return v * 100 if v < 1 else v
 
                    def cond_text(val):
                        return str(val).strip().replace('\n', ' ') if pd.notna(val) else ""

                    history_list.append({
                        "금융기관": display_name.replace('\n', ' '),
                        f"{HIST_YEAR_START}년": parse_rate(row.iloc[6]),
                        f"{HIST_YEAR_START + 1}년": parse_rate(row.iloc[8]),
                        f"{HIST_YEAR_START + 2}년": parse_rate(row.iloc[11]),
                        f"{HIST_YEAR_END}년": parse_rate(row.iloc[14]),
                        # 금리조건(벤치마크) 원문 텍스트 — ECOS 벤치마크 매칭용
                        f"{HIST_YEAR_START}년_조건": cond_text(row.iloc[5]),
                        f"{HIST_YEAR_START + 1}년_조건": cond_text(row.iloc[7]),
                        f"{HIST_YEAR_START + 2}년_조건": cond_text(row.iloc[10]),
                        f"{HIST_YEAR_END}년_조건": cond_text(row.iloc[13]),
                        "신용등급변동": str(row.iloc[12]) if pd.notna(row.iloc[12]) else "-"
                    })
            hist_data = pd.DataFrame(history_list)
        else:
            hist_data = pd.DataFrame()
 
        return kpi_data, hist_data
 
    except Exception as e:
        st.error(f"엑셀 파일 처리 중 오류가 발생했습니다: {e}")
        return None, None
 
 
def get_normalized_bank_name(name):
    name = str(name).strip().replace('\n', ' ').replace(' ', '')
    if '국민' in name or 'KB' in name:
        return 'KB'
    if '우리' in name:
        return '우리'
    if '농협' in name or 'NH' in name:
        return 'NH'
    if '산업' in name:
        return '산업'
    if '대구' in name or 'IM' in name or '아이엠' in name or 'im' in name.lower():
        return 'IM뱅크'
    if '신한' in name:
        return '신한'
    if '하나' in name:
        return '하나'
    if '기업' in name:
        return '기업'
    return None


def normalize_relation_name(raw_name):
    """예금·카드 거래처 이름을 정규화한다. '우리투자(증권)'처럼 우리은행과 구분해야 하는
    관계사는 은행 코드로 합쳐지지 않도록 별도 처리한다."""
    n = str(raw_name).strip()
    if '우리투자' in n or ('우리' in n and '증권' in n):
        return '우리투자(증권)'
    norm = get_normalized_bank_name(n)
    return norm or n

 
 
def rating_to_numeric(rating_str):
    import re
    r = str(rating_str).strip().upper()
    if r == '-' or r == 'NAN' or not r or '전년등급' in r:
        return None
    match = re.search(r'\(([^)]+)\)', r)
    if match:
        r = match.group(1).strip()
    r = r.replace('0', '')
 
    mapping = {
        'AAA': 10,
        'AA+': 9,
        'AA': 8,
        'AA-': 7,
        'A+': 6,
        'A': 5,
        'A-': 4,
        'BBB+': 3,
        'BBB': 2,
        'BBB-': 1
    }
    for k in sorted(mapping.keys(), key=len, reverse=True):
        if k in r:
            return mapping[k]
 
    if '3등급' in r or 'A5' in r or 'A-2' in r or 'A-3' in r:
        return 4
    if '5등급' in r or 'A6' in r:
        return 2
    return None
 
 
@st.cache_data(ttl=60, show_spinner=False)
def check_credit_downgrade(excel_bytes):
    downgrades = {}
    try:
        xl = pd.ExcelFile(io.BytesIO(excel_bytes))
        credit_sheets = [s for s in xl.sheet_names if '신용등급' in s]
        if credit_sheets:
            df = xl.parse(credit_sheets[0], header=None)
            for r in range(6, 14):
                row = df.iloc[r]
                bank_name = str(row.iloc[1]).strip()
                norm_name = get_normalized_bank_name(bank_name)
                if not norm_name:
                    continue
                r_24 = str(row.iloc[9]).strip() if pd.notna(row.iloc[9]) else ""
                r_25 = str(row.iloc[10]).strip() if pd.notna(row.iloc[10]) else ""
 
                is_downgrade = False
                if "하락" in r_25 or "하락" in r_24:
                    is_downgrade = True
                elif norm_name == '산업':
                    is_downgrade = True
                elif norm_name == 'IM뱅크':
                    is_downgrade = True
 
                downgrades[norm_name] = {
                    "r_24": r_24,
                    "r_25": r_25,
                    "is_downgrade": is_downgrade
                }
    except Exception:
        pass
    return downgrades
 
 
@st.cache_data(ttl=60, show_spinner=False)
def load_deposit_contribution(excel_bytes):
    """'금융기관별_예금_기여도_및_수익률_분석' 양식에서 '{연도}년_연간 기여도 분석' 시트들을 모두 읽어
    {연도(int): DataFrame(금융기관, 잔액, 비중, 금리, 비고)} 형태로 반환한다. '합계/전체 평균' 행은 제외."""
    import re
    result = {}
    try:
        xl = pd.ExcelFile(io.BytesIO(excel_bytes))
        target_sheets = [s for s in xl.sheet_names if '연간 기여도 분석' in s]
        for sheet in target_sheets:
            m = re.search(r'(\d{4})', sheet)
            if not m:
                continue
            year = int(m.group(1))
            df_raw = xl.parse(sheet, header=2)  # 3번째 행이 실제 헤더('금융기관명' 등)
            df_raw = df_raw.rename(columns=lambda c: str(c).strip())
            name_col = next((c for c in df_raw.columns if '금융기관' in c), None)
            bal_col = next((c for c in df_raw.columns if '잔액' in c), None)
            weight_col = next((c for c in df_raw.columns if '비중' in c or '기여도' in c), None)
            rate_col = next((c for c in df_raw.columns if '금리' in c), None)
            note_col = next((c for c in df_raw.columns if '비고' in c), None)
            if not all([name_col, bal_col, weight_col, rate_col]):
                continue
            df_clean = df_raw[[name_col, bal_col, weight_col, rate_col] + ([note_col] if note_col else [])].copy()
            df_clean.columns = ['금융기관', '잔액', '비중', '금리'] + (['비고'] if note_col else [])
            df_clean = df_clean[df_clean['금융기관'].notna()]
            df_clean = df_clean[~df_clean['금융기관'].astype(str).str.contains('합계|전체 평균', na=False)]
            df_clean['잔액'] = pd.to_numeric(df_clean['잔액'], errors='coerce').fillna(0)
            df_clean['비중'] = pd.to_numeric(df_clean['비중'], errors='coerce').fillna(0)
            df_clean['금리'] = pd.to_numeric(df_clean['금리'], errors='coerce').fillna(0)
            result[year] = df_clean.reset_index(drop=True)
    except Exception:
        pass
    return result


@st.cache_data(ttl=60, show_spinner=False)
def load_card_usage(excel_bytes):
    """'법인카드_사용_금액' 양식을 읽어 은행별-연도별 사용금액 long-format DataFrame으로 반환한다.
    컬럼: 은행, 연도(문자열), 금액"""
    try:
        xl = pd.ExcelFile(io.BytesIO(excel_bytes))
        sheet = xl.sheet_names[0]
        df_raw = xl.parse(sheet, header=None)

        header_row_idx = None
        for i in range(min(10, len(df_raw))):
            row_vals = [str(v).strip() for v in df_raw.iloc[i] if pd.notna(v)]
            if any('은행' in v for v in row_vals):
                header_row_idx = i
                break
        if header_row_idx is None:
            return pd.DataFrame()

        header = df_raw.iloc[header_row_idx]
        bank_col_idx = next(i for i, v in enumerate(header) if pd.notna(v) and '은행' in str(v))
        year_cols = [(i, str(v).strip()) for i, v in enumerate(header)
                     if pd.notna(v) and i != bank_col_idx and str(v).strip() != ""]

        records = []
        for r in range(header_row_idx + 1, len(df_raw)):
            row = df_raw.iloc[r]
            bank_name = row.iloc[bank_col_idx]
            if pd.isna(bank_name) or str(bank_name).strip().startswith('*'):
                continue
            bank_name = str(bank_name).strip()
            for i, yr_label in year_cols:
                amt = pd.to_numeric(row.iloc[i], errors='coerce')
                if pd.notna(amt):
                    records.append({'은행': bank_name, '연도': yr_label, '금액': amt})
        return pd.DataFrame(records)
    except Exception:
        return pd.DataFrame()


PENSION_SERVER_URL = "http://localhost:8000/data.json"


@st.cache_data(ttl=300, show_spinner=False)
def fetch_pension_institution_amounts():
    """퇴직연금통합관리시스템(별도 프로그램, 포트 8000)에서 기관별 적립금 현황을
    가져온다. 예금·법인카드와 마찬가지로 "기여도" 축 중 하나로, 부수거래 화면에서
    같은 기관 기준으로 나란히 보여주기 위함이다(2026년 KPI 'CSF: 금융기관별
    기여도(퇴직연금ㆍ수신예금 등) 히스토리 관리'에 대응).
    반환: (base_date, {정규화된 기관명: 적립금(원)}) — 서버가 꺼져있거나 데이터가
    없으면 (None, {})."""
    try:
        resp = requests.get(PENSION_SERVER_URL, timeout=5)
        resp.raise_for_status()
        snapshots = resp.json().get('snapshots', {})
        if not snapshots:
            return None, {}
        latest_date = max(snapshots.keys())
        dist = snapshots[latest_date].get('institution_distribution', [])
        amounts = {}
        for row in dist:
            norm = normalize_relation_name(row.get('institution', ''))
            amounts[norm] = amounts.get(norm, 0) + float(row.get('amount', 0))
        return latest_date, amounts
    except Exception:
        return None, {}


@st.cache_data(ttl=60, show_spinner=False)
def list_loan_snapshot_sheets(excel_bytes):
    """엑셀 안의 '차입금 관리내역(현재)' 계열 시트를 모두 찾아, 시트명 끝부분에 박혀있는
    날짜(예: '26.07.31', '~26.0716')를 파싱해 (시트명, 날짜 또는 None) 리스트로 반환합니다.
    최신 날짜가 먼저 오도록 정렬합니다(날짜를 못 찾은 시트는 맨 뒤).
    여러 스냅샷을 남겨둔 파일에서 분기별/시점별로 골라볼 수 있게 하기 위함입니다."""
    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes), read_only=True)
    candidates = [
        s for s in wb.sheetnames
        if '차입금' in s and '이자율' not in s and '신용등급' not in s and '성과' not in s
    ]
    current_candidates = [s for s in candidates if '현재' in s]
    if not current_candidates:
        current_candidates = candidates[:1] if candidates else (wb.sheetnames[:1] if wb.sheetnames else [])

    results = []
    for s in current_candidates:
        digits = re.sub(r'\D', '', s)
        parsed = None
        if len(digits) >= 6:
            yy, mm, dd = digits[-6:-4], digits[-4:-2], digits[-2:]
            try:
                parsed = pd.Timestamp(year=2000 + int(yy), month=int(mm), day=int(dd))
            except ValueError:
                parsed = None
        results.append((s, parsed))

    results.sort(key=lambda t: (t[1] is None, -(t[1].value if t[1] is not None else 0)))
    return results


@st.cache_data(ttl=60, show_spinner=False)
def load_current_loan_status(excel_bytes, sheet_name_override=None):
    """'★ 차입금 관리내역(현재)' 시트를 그대로 파싱해서 실시간 차입금 현황표로 사용합니다.

    이 시트는 은행명/대출과목/금리구분 등 여러 열이 병합 셀(merged cell)로
    구성되어 있어, 반드시 openpyxl로 병합 범위를 먼저 풀어(unmerge) 값을
    모든 하위 셀에 채워 넣은 뒤 파싱해야 데이터 누락이 없습니다.
    (과거에는 pandas로 바로 읽어 병합된 셀들이 NaN으로 남아 행이 빠졌고,
     '금융기관'(A열)이 아닌 '대출과목'(B열) 기준으로 유효 행을 판정하는
     버그가 있었습니다.)

    sheet_name_override가 주어지면(분기별 스냅샷 선택 등) 그 시트를 그대로 사용하고,
    없으면 기존처럼 가장 최근에 추가된 '현재' 시트를 자동 선택합니다.
    """
    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes), data_only=True)

    if sheet_name_override and sheet_name_override in wb.sheetnames:
        sheet_name = sheet_name_override
    else:
        # 1) 시트 후보 탐색: '차입금'이 들어가되 과거이력/신용등급/성과표 시트는 제외
        candidates = [
            s for s in wb.sheetnames
            if '차입금' in s and '이자율' not in s and '신용등급' not in s and '성과' not in s
        ]
        if not candidates:
            candidates = [wb.sheetnames[0]]
        # '현재'가 포함된 시트를 우선 사용, 없으면 첫 후보 (시트명 앞뒤 공백 허용)
        # '현재' 시트가 여러 개면(예: '~26.0716', '26.07.31'처럼 날짜별 스냅샷을 남겨둔 경우)
        # 통상 나중에 추가된 시트가 더 최신 데이터이므로 마지막 후보를 사용한다.
        current_candidates = [s for s in candidates if '현재' in s]
        sheet_name = current_candidates[-1] if current_candidates else candidates[0]
    ws = wb[sheet_name]

    # 2) 병합 셀 해제 + 값 전파: 병합 범위의 좌상단 값을 병합된 모든 셀에 채워넣는다.
    merged_ranges = list(ws.merged_cells.ranges)
    for merged_range in merged_ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds
        top_left_value = ws.cell(row=min_row, column=min_col).value
        ws.unmerge_cells(str(merged_range))
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                ws.cell(row=r, column=c).value = top_left_value

    max_row = ws.max_row
    max_col = ws.max_column

    # 3) '금융기관'이 있는 헤더 행 탐색 (A열 기준)
    header_idx = None
    for r in range(1, min(max_row, 30) + 1):
        v = ws.cell(row=r, column=1).value
        if v is not None and '금융기관' in str(v):
            header_idx = r
            break

    if header_idx is None:
        return pd.DataFrame(), sheet_name

    # 4) 메인 표(좌측 A~R열 등)의 오른쪽 경계 찾기: 헤더가 시작된 후
    #    처음으로 빈 헤더 셀이 나오는 지점을 경계로 삼는다.
    #    (그 오른쪽엔 은행별 이율/이자/잔액 비교용 별도 피벗표가 있어 제외해야 함)
    main_cols = []
    started = False
    for c in range(1, max_col + 1):
        v = ws.cell(row=header_idx, column=c).value
        if v is not None and str(v).strip() != '':
            main_cols.append(c)
            started = True
        elif started:
            break

    # 5) 헤더명 정리 (줄바꿈 제거, 중복명 처리)
    columns = []
    seen = {}
    for c in main_cols:
        v = ws.cell(row=header_idx, column=c).value
        name = str(v).strip().replace('\n', ' ') if v is not None else ''
        if name == '' or name.lower() == 'nan':
            name = f"항목{c}"
        if name in seen:
            seen[name] += 1
            name = f"{name}_{seen[name]}"
        else:
            seen[name] = 0
        columns.append(name)

    # 6) 데이터 행 수집: '금융기관'(A열, main_cols[0])을 기준으로 유효 행 판정
    bank_col = main_cols[0]
    data_rows = []
    for r in range(header_idx + 1, max_row + 1):
        bank_val = ws.cell(row=r, column=bank_col).value
        bank_val_str = str(bank_val).strip() if bank_val is not None else ""
        if bank_val_str == '' or bank_val_str.lower() == 'nan':
            continue
        row_vals = [ws.cell(row=r, column=c).value for c in main_cols]
        data_rows.append(row_vals)

    if not data_rows:
        return pd.DataFrame(columns=columns), sheet_name

    df = pd.DataFrame(data_rows, columns=columns)

    # 날짜형 컬럼은 문자열로 통일 (표시 일관성)
    for col in df.columns:
        if pd.api.types.is_datetime64_any_dtype(df[col]):
            df[col] = df[col].dt.strftime('%Y-%m-%d')

    return df, sheet_name


def compute_current_bank_rate_summary(df_current):
    """'🏢 금융기관별 금리 비교' 화면의 당해년도 막대는 원래 '과거 이자율변동내역' 시트에
    사람이 따로 타이핑해둔 셀 값을 썼는데, 그 셀이 '★ 차입금 관리내역(현재)'가 바뀔 때
    같이 갱신된다는 보장이 없어(실제로 아이엠뱅크ㆍ한국증권금융 등에서 어긋난 값이
    확인됨) 기본 데이터를 고쳐도 그래프가 그대로였다. 당해년도만큼은 이렇게 별도로
    타이핑된 값 대신, '★ 차입금 관리내역(현재)' 표를 은행별로 직접 집계해 항상 최신
    상태를 그대로 반영하게 한다(과거 연도는 현재 시점 데이터로 재현할 수 없어 그대로
    '과거 이자율변동내역' 시트를 쓴다).

    은행 하나에 차입건이 여러 개면(한도대+일반대 등) 잔액 가중평균을, 잔액이 없는(미사용
    한도 등) 건은 차입한도로 대신 가중치를 준다. 반환: {금융기관: 가중평균금리(%)}.
    """
    if df_current is None or df_current.empty:
        return {}
    is_total_row = df_current['금융기관'].astype(str).str.strip().isin(['계', '합계'])
    df = df_current[~is_total_row].copy()

    balance_col = next((c for c in df.columns if '잔액' in c), None)
    limit_col = next((c for c in df.columns if '차입한도' in c), None)
    rate_col = next((c for c in df.columns if c.strip() == '금리'), None)
    if rate_col is None:
        return {}

    df['_rate'] = pd.to_numeric(df[rate_col], errors='coerce')
    df['_balance'] = pd.to_numeric(df[balance_col], errors='coerce') if balance_col else pd.Series(dtype=float)
    df['_limit'] = pd.to_numeric(df[limit_col], errors='coerce') if limit_col else pd.Series(dtype=float)
    # 가중치: 잔액이 있으면 잔액, 없으면(미사용 한도대 등) 차입한도, 둘 다 없으면 이 건은 제외
    df['_weight'] = df['_balance'].where(df['_balance'].fillna(0) > 0, df['_limit'])

    summary = {}
    for bank, g in df.groupby(df['금융기관'].astype(str).str.strip()):
        g = g[g['_rate'].notna() & (g['_weight'].fillna(0) > 0)]
        if g.empty:
            continue
        weighted_rate = (g['_rate'] * g['_weight']).sum() / g['_weight'].sum()
        summary[bank] = weighted_rate * 100 if weighted_rate < 1 else weighted_rate
    return summary


def build_loan_status_workbook(df, sheet_name="★ 차입금 관리내역(현재)"):
    """직접입력 표(DataFrame)를 load_current_loan_status()가 그대로 다시 읽을 수 있는
    엑셀 바이트로 변환합니다. 1행에 헤더, 2행부터 데이터를 채운 단순한 워크북이며,
    A열이 '금융기관' 헤더/값 기준으로 유효 행을 판정하는 기존 파싱 로직과 호환됩니다."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    headers = [str(c) for c in df.columns]
    for j, h in enumerate(headers, start=1):
        ws.cell(row=1, column=j, value=h)
    for i, row in enumerate(df.itertuples(index=False), start=2):
        for j, val in enumerate(row, start=1):
            if val is None or (isinstance(val, float) and pd.isna(val)):
                continue
            ws.cell(row=i, column=j, value=val)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


render_sheet_html = report_shared.render_sheet_html


def flatten_sheet_to_df(ws, min_header_cells=3, search_rows=15):
    """단순한(단일 헤더행) 표 형태 시트를 편집 가능한 DataFrame으로 변환합니다.
    병합 셀은 좌상단 값을 전체 범위에 채워 넣어 풀고, 제목/여백 행을 건너뛴 뒤
    '실제 값이 여러 개 채워진 첫 행'을 헤더로 자동 인식합니다(이 보고서들은 1행이
    바로 헤더가 아니라 "N. OO 현황" 같은 제목 행으로 시작하는 경우가 많음).
    2단 헤더(헤더 바로 아래 줄에 소제목이 더 있는 경우)나 표지처럼 표 형태가 아닌
    시트에는 완벽히 맞지 않을 수 있어, 그런 시트는 직접입력 대상에서 제외합니다."""
    merged_ranges = list(ws.merged_cells.ranges)
    for merged_range in merged_ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds
        top_left_value = ws.cell(row=min_row, column=min_col).value
        ws.unmerge_cells(str(merged_range))
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                ws.cell(row=r, column=c).value = top_left_value

    header_row = 1
    for r in range(1, min(ws.max_row, search_rows) + 1):
        non_empty = sum(
            1 for c in range(1, ws.max_column + 1)
            if ws.cell(row=r, column=c).value is not None and str(ws.cell(row=r, column=c).value).strip() != ''
        )
        if non_empty >= min_header_cells:
            header_row = r
            break

    def _cell_has_value(r, c):
        v = ws.cell(row=r, column=c).value
        return v is not None and str(v).strip() != ''

    # 실제 값이 있는 마지막 열까지만 사용(끝쪽의 빈 '항목N' 열이 잔뜩 남는 것을 방지).
    used_max_col = 1
    for c in range(1, ws.max_column + 1):
        if _cell_has_value(header_row, c) or any(
            _cell_has_value(r, c) for r in range(header_row + 1, ws.max_row + 1)
        ):
            used_max_col = c

    headers = []
    seen = {}
    for c in range(1, used_max_col + 1):
        v = ws.cell(row=header_row, column=c).value
        name = str(v).strip().replace('\n', ' ') if v is not None and str(v).strip() != '' else f"항목{c}"
        if name in seen:
            seen[name] += 1
            name = f"{name}_{seen[name]}"
        else:
            seen[name] = 0
        headers.append(name)

    data_rows = []
    for r in range(header_row + 1, ws.max_row + 1):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, used_max_col + 1)]
        if all(v is None or str(v).strip() == '' for v in row_vals):
            continue
        data_rows.append(row_vals)

    df = pd.DataFrame(data_rows, columns=headers)
    for col in df.columns:
        if pd.api.types.is_datetime64_any_dtype(df[col]):
            df[col] = df[col].dt.strftime('%Y-%m-%d')
    return df


def write_df_as_sheet(wb, sheet_name, df):
    """DataFrame을 워크북의 시트로 씁니다(같은 이름의 시트가 있으면 삭제 후 재생성)."""
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    for j, h in enumerate(df.columns, start=1):
        ws.cell(row=1, column=j, value=str(h))
    for i, row in enumerate(df.itertuples(index=False), start=2):
        for j, val in enumerate(row, start=1):
            if val is None or (isinstance(val, float) and pd.isna(val)):
                continue
            ws.cell(row=i, column=j, value=val)


@st.cache_data(ttl=60, show_spinner=False)
def get_active_loan_bank_names(excel_bytes):
    """'★ 차입금 관리내역(현재)' 시트를 기준으로, 현재 차입금 잔액(잔액 ≠ 0)이 있는 금융기관의
    정규화된 이름 집합을 반환한다. 예금 기여도·법인카드 등 부수거래 화면에서 차입금 현황이
    없는(잔액 0 또는 아예 거래이력이 없는) 기관을 표에서 제외하는 데 사용한다."""
    try:
        df_current, _ = load_current_loan_status(excel_bytes)
        if df_current.empty or '금융기관' not in df_current.columns:
            return set()
        balance_col = next((c for c in df_current.columns if '잔액' in c), None)
        is_total_row = df_current['금융기관'].astype(str).str.strip().isin(['계', '합계'])
        df_banks = df_current[~is_total_row]
        if balance_col is not None:
            bal_numeric = pd.to_numeric(df_banks[balance_col], errors='coerce').fillna(0)
            df_banks = df_banks[bal_numeric != 0]
        return {normalize_relation_name(v) for v in df_banks['금융기관'] if str(v).strip()}
    except Exception:
        return set()


def get_loan_balance_by_bank(excel_bytes):
    """'★ 차입금 관리내역(현재)' 시트를 기준으로 정규화된 기관명별 차입금 잔액 합계를
    반환한다. 부수거래(예금ㆍ카드ㆍ퇴직연금) 화면에서 "이 기관이 전체 차입금 중 몇 %를
    차지하는가"를 계산해, 반대편(퇴직연금 적립금 등)의 비중과 나란히 비교하기 위함이다
    — 거래관계 규모가 커도 그게 우리 전체 포트폴리오에서 큰 비중인지 작은 비중인지는
    실제 %로 봐야 알 수 있다."""
    try:
        df_current, _ = load_current_loan_status(excel_bytes)
        if df_current.empty or '금융기관' not in df_current.columns:
            return {}
        balance_col = next((c for c in df_current.columns if '잔액' in c), None)
        if balance_col is None:
            return {}
        is_total_row = df_current['금융기관'].astype(str).str.strip().isin(['계', '합계'])
        df_banks = df_current[~is_total_row].copy()
        df_banks['_잔액숫자'] = pd.to_numeric(df_banks[balance_col], errors='coerce').fillna(0)
        df_banks = df_banks[df_banks['_잔액숫자'] != 0]
        result = {}
        for _, row in df_banks.iterrows():
            norm = normalize_relation_name(row['금융기관'])
            result[norm] = result.get(norm, 0.0) + float(row['_잔액숫자'])
        return result
    except Exception:
        return {}


# ---------------------------------------------------------------------------
# 1년이내 만기도래 및 상환계획표
# ---------------------------------------------------------------------------
def load_repayment_status():
    """사용자가 입력한 '상환여부'/'비고' 값을 로컬 JSON에서 불러옵니다."""
    if os.path.exists(REPAYMENT_STATUS_JSON_FILE):
        try:
            with open(REPAYMENT_STATUS_JSON_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception:
            return {}
    return {}


def save_repayment_status(data):
    """사용자가 입력한 '상환여부'/'비고' 값을 로컬 JSON에 저장합니다."""
    try:
        with open(REPAYMENT_STATUS_JSON_FILE, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
    except Exception:
        pass


def build_maturity_schedule(df_current, as_of_date=None):
    """'차입금 관리내역(현재)' 표에서 조회일자(현재일) 기준 1년 이내 만기도래
    예정인 대출 건만 뽑아 '1년이내 만기도래 및 상환계획표' 형태로 재구성합니다.

    - 기준일(as_of_date) ~ 기준일+1년 사이에 만기일이 있는 건만 포함
    - 금액은 잔액(원)을 우선 사용하고, 잔액이 없으면 차입한도(원)을 사용 (백만원 단위)
    - 과목은 '신용/담보' 구분을 담보대출/신용대출로 표기
    - 동일 차입기관에 2건 이상 있으면 만기일 순으로 정렬
    - 마지막에 '계'(합계) 행을 추가
    반환: (표시용 DataFrame, 각 행의 고유 key 리스트) — key는 상환여부/비고 저장용
    """
    if as_of_date is None:
        as_of_date = datetime.now().date()
    cutoff_date = pd.Timestamp(as_of_date) + pd.DateOffset(years=1)

    is_total_row = df_current['금융기관'].astype(str).str.strip().isin(['계', '합계'])
    df = df_current[~is_total_row].copy()

    maturity_col = next((c for c in df.columns if '만기일' in c), None)
    balance_col = next((c for c in df.columns if '잔액' in c), None)
    limit_col = next((c for c in df.columns if '차입한도' in c or ('한도' in c and '금액' in c)), None)
    subj_col = next((c for c in df.columns if '신용' in c and '담보' in c), None)  # 신용/담보
    loan_type_col = '대출과목' if '대출과목' in df.columns else None
    collateral_col = next((c for c in df.columns if '담보내역' in c), None)

    if maturity_col is None:
        return pd.DataFrame(columns=['차입기관명', '과목', '금액(백만원)', '만기일', '담보종류', '상환여부', '비고']), []

    df['_만기일_dt'] = pd.to_datetime(df[maturity_col], errors='coerce')
    df = df.dropna(subset=['_만기일_dt'])
    mask = (df['_만기일_dt'] >= pd.Timestamp(as_of_date)) & (df['_만기일_dt'] <= cutoff_date)
    df = df[mask].copy()

    if df.empty:
        return pd.DataFrame(columns=['차입기관명', '과목', '금액(백만원)', '만기일', '담보종류', '상환여부', '비고']), []

    def pick_amount(row):
        bal = pd.to_numeric(row.get(balance_col), errors='coerce') if balance_col else None
        if pd.notna(bal) and bal not in (0, 0.0):
            return bal
        return None  # 잔액(실제 차입액)이 없으면 상환계획표에 포함하지 않음 (미사용 한도 등)

    def pick_subject(row):
        subj = str(row.get(subj_col, '')).strip() if subj_col else ''
        if '담보' in subj:
            return '담보대출'
        if '신용' in subj:
            return '신용대출'
        # '신용/담보' 표기가 비어있으면 담보내역 유무로 추정
        collateral = str(row.get(collateral_col, '')).strip() if collateral_col else ''
        if collateral and collateral.lower() != 'nan':
            return '담보대출'
        if loan_type_col:
            v = str(row.get(loan_type_col, '')).strip()
            return v if v and v.lower() != 'nan' else '-'
        return '-'

    def pick_collateral(row):
        v = str(row.get(collateral_col, '')).strip() if collateral_col else ''
        if v == '' or v.lower() == 'nan':
            return '-'
        return v

    records = []
    keys = []
    df = df.sort_values(['금융기관', '_만기일_dt'])
    for _, row in df.iterrows():
        amount = pick_amount(row)
        if amount is None:
            continue  # 실제 차입잔액이 없는 건(미사용 한도 등)은 상환계획표에서 제외
        amount_mm = round(amount / 1_000_000)
        maturity_str = row['_만기일_dt'].strftime('%Y-%m-%d')
        subject = pick_subject(row)
        collateral = pick_collateral(row)
        bank = str(row['금융기관']).strip()
        key = f"{bank}|{subject}|{maturity_str}|{amount_mm}"
        records.append({
            '차입기관명': bank,
            '과목': subject,
            '금액(백만원)': amount_mm,
            '만기일': maturity_str,
            '담보종류': collateral,
            '상환여부': '',
            '비고': '',
        })
        keys.append(key)

    result_df = pd.DataFrame(records)

    # 합계 행 추가
    total_amount = result_df['금액(백만원)'].sum(skipna=True)
    total_row = pd.DataFrame([{
        '차입기관명': '계', '과목': '', '금액(백만원)': total_amount,
        '만기일': '', '담보종류': '', '상환여부': '', '비고': ''
    }])
    result_df = pd.concat([result_df, total_row], ignore_index=True)
    keys.append('__TOTAL__')

    return result_df, keys


# ---------------------------------------------------------------------------
# 담보현황: 현재 차입금 중 '담보내역'이 기재된 건을 물건지별로 정리
# ---------------------------------------------------------------------------
def build_collateral_overview(df_current):
    """차입금현황(현재) 표에서 '담보내역'이 채워진 대출만 뽑아, 대출별 상세와
    담보 물건지별(같은 물건지가 여러 대출을 담보하는 경우 한눈에 보이도록) 합계를 반환합니다.

    반환: (대출별 상세 DataFrame, 물건지별 합계 DataFrame)
    """
    is_total_row = df_current['금융기관'].astype(str).str.strip().isin(['계', '합계'])
    df = df_current[~is_total_row].copy()

    balance_col = next((c for c in df.columns if '잔액' in c), None)
    collateral_col = next((c for c in df.columns if '담보내역' in c), None)
    subj_col = next((c for c in df.columns if '신용' in c and '담보' in c), None)
    loan_type_col = '대출과목' if '대출과목' in df.columns else None
    maturity_col = next((c for c in df.columns if '만기일' in c), None)

    if collateral_col is None:
        return pd.DataFrame(), pd.DataFrame()

    df['_담보내역'] = df[collateral_col].astype(str).str.strip()
    # 빈 셀(None)이 pandas의 astype(str)을 거치면 "nan"이 아니라 "None" 문자열이 되는
    # 경우가 있어(openpyxl이 빈 셀을 np.nan이 아니라 파이썬 None으로 주는 컬럼에서 발생),
    # "nan"만 걸러내던 기존 필터로는 못 잡혀 표에 "None" 글자가 그대로 노출되고, 담보
    # 없는 대출(예: 신용보증기금 보증 대출처럼 물건지가 없는 건)들이 "None"이라는
    # 가짜 물건지 하나로 잘못 묶이는 문제가 있었다(실제 발견: 685억ㆍ5건이 이렇게 묶여
    # 있었음 — 담보내역이 비어있는 신용보증형 대출들).
    df = df[(df['_담보내역'] != '') & (~df['_담보내역'].str.lower().isin(('nan', 'none')))]
    if df.empty:
        return pd.DataFrame(), pd.DataFrame()

    df['_잔액숫자'] = pd.to_numeric(df[balance_col], errors='coerce').fillna(0) if balance_col else 0.0

    def _cell_or_dash(row, col):
        # row.get(col, '-')는 컬럼 자체가 없을 때만 기본값을 주고, 컬럼은 있는데 그 칸이
        # 비어있으면(NaN/None) 그대로 None을 돌려줘 str()을 거치면 "None" 글자가 그대로
        # 표에 찍히는 문제가 있었다 — 값이 없으면 항상 "-"가 나오도록 명시적으로 확인한다.
        if not col:
            return '-'
        v = row.get(col)
        if v is None or (isinstance(v, float) and pd.isna(v)) or str(v).strip().lower() in ('', 'nan', 'none'):
            return '-'
        return str(v).strip()

    detail_rows = []
    for _, row in df.iterrows():
        detail_rows.append({
            '담보물건지': row['_담보내역'] or '-',
            '금융기관': str(row['금융기관']).strip(),
            '대출과목': _cell_or_dash(row, loan_type_col),
            '구분': _cell_or_dash(row, subj_col),
            '잔액(원)': row['_잔액숫자'],
            '만기일': _cell_or_dash(row, maturity_col),
        })

    detail_df = pd.DataFrame(detail_rows).sort_values(['담보물건지', '금융기관']).reset_index(drop=True)

    summary_df = (
        detail_df.groupby('담보물건지')
        .agg(
            잔액합계=('잔액(원)', 'sum'),
            담보대출건수=('잔액(원)', 'size'),
            담보제공기관=('금융기관', lambda s: ', '.join(sorted(set(s)))),
        )
        .reset_index()
        .sort_values('잔액합계', ascending=False)
        .reset_index(drop=True)
    )

    return detail_df, summary_df


def _norm_site_name(s):
    return re.sub(r'[\s()（）]', '', str(s or ''))


def find_real_estate_loan_diffs(real_estate_rows, coll_summary_df):
    """부동산현황(은행이 기재되어 '담보 제공중'인 사업장)과 차입금현황에서 뽑은
    담보물건지별 합계(build_collateral_overview)를 사업장명 부분일치로 대조합니다.
    이름 표기가 서로 다를 수 있어 완벽한 자동 매칭은 아니며, 사람이 한 번 더
    확인해야 하는 후보를 걸러내는 용도입니다.
    반환: {real_estate_site_id: "차이 설명"} — 차이가 없으면 그 id는 없음."""
    coll_bank_by_site = {}
    if coll_summary_df is not None and not coll_summary_df.empty:
        for _, r in coll_summary_df.iterrows():
            coll_bank_by_site[r['담보물건지']] = r['담보제공기관']

    diffs = {}
    for row in real_estate_rows:
        bank = str(row.get('bank') or '').strip()
        if not bank:
            continue  # 담보로 제공되지 않은 부동산은 대조 대상이 아님
        site_norm = _norm_site_name(row['site_name'])
        matched_bank = None
        for coll_site, coll_bank in coll_bank_by_site.items():
            coll_site_norm = _norm_site_name(coll_site)
            if site_norm and coll_site_norm and (site_norm in coll_site_norm or coll_site_norm in site_norm):
                matched_bank = coll_bank
                break
        if matched_bank is None:
            diffs[row['id']] = f"차입금현황에서 이 담보물건을 찾지 못함(부동산현황 은행: {bank})"
        elif bank not in matched_bank:
            diffs[row['id']] = f"은행 불일치 — 부동산현황: {bank} / 차입금현황: {matched_bank}"
    return diffs


# ---------------------------------------------------------------------------
# 예상이자 계산 (금리변경 예정일 체크 + 이자지급일 기준 예상이자)
# ---------------------------------------------------------------------------
def parse_interest_pay_schedule(text):
    """'이자지급일' 텍스트를 파싱해 (지급일(1~31), 지급주기(개월), 비고)를 반환합니다.
    예) '28일' → (28, 1, ''), '3개월단위 - 26일' → (26, 3, ''),
        비어있으면 → (None, None, '') 로 반환해 호출부에서 기본값(28일/매월)을 적용하게 합니다."""
    if text is None:
        return None, None, ''
    s = str(text).strip()
    if not s or s.lower() == 'nan':
        return None, None, ''

    day_match = re.search(r'(\d{1,2})\s*일', s)
    day = int(day_match.group(1)) if day_match else None
    if day is not None and not (1 <= day <= 31):
        day = None

    cadence_match = re.search(r'(\d{1,2})\s*개월\s*단위', s)
    if cadence_match:
        cadence = int(cadence_match.group(1))
    elif '분기' in s:
        cadence = 3
    else:
        cadence = 1 if day is not None else None

    return day, cadence, s


def _snap_day(d, target_day):
    """date/Timestamp d의 연-월은 유지한 채, day만 target_day로 맞춥니다(말일 초과 시 말일로 보정)."""
    last_day = calendar.monthrange(d.year, d.month)[1]
    return pd.Timestamp(year=d.year, month=d.month, day=min(target_day, last_day))


def bracket_pay_dates(anchor, cadence_months, target_day, as_of):
    """anchor(기표일/차입일)를 시작점으로 cadence_months 간격, 매번 target_day에 맞춘 지급일들 중,
    as_of를 감싸는 (직전 지급일, 다음 지급일)을 반환합니다."""
    anchor = pd.to_datetime(anchor, errors='coerce')
    as_of_ts = pd.Timestamp(as_of)
    if pd.isna(anchor):
        anchor = as_of_ts

    cur = _snap_day(anchor, target_day)
    guard = 0
    while cur <= as_of_ts and guard < 2000:
        nxt = _snap_day(cur + pd.DateOffset(months=cadence_months), target_day)
        if nxt <= cur:
            nxt = cur + pd.DateOffset(days=1)
        if nxt > as_of_ts:
            return cur, nxt
        cur = nxt
        guard += 1

    # anchor 자체가 as_of보다 미래인 경우: 한 주기 이전으로 되돌려 직전 지급일을 추정
    prev = _snap_day(cur - pd.DateOffset(months=cadence_months), target_day)
    return prev, cur


def next_rate_change_date(cycle_months, last_change_raw, origin_raw, as_of):
    """변동금리 대출의 다음 금리변경 예정일을 계산합니다.
    반환: (예정일 Timestamp 또는 None, 산출근거 문자열)
    - cycle_months가 없거나(고정금리/'-') 0 이하면 정기 스케줄이 없는 것으로 보고 (None, 'no_cycle')
    - '최근금리변경일' 필드에 이미 미래(as_of 이후) 날짜가 기재돼 있으면 그 값을 그대로 사용
    - 없으면 기표일/차입일로부터 cycle_months 간격으로 as_of 이후 첫 날짜를 추정
    """
    cycle = pd.to_numeric(cycle_months, errors='coerce')
    if pd.isna(cycle) or cycle <= 0:
        return None, 'no_cycle'

    as_of_ts = pd.Timestamp(as_of)
    last_change = pd.to_datetime(last_change_raw, errors='coerce')
    if pd.notna(last_change) and last_change >= as_of_ts:
        return last_change, 'from_field'

    origin = pd.to_datetime(origin_raw, errors='coerce')
    if pd.isna(origin):
        origin = last_change
    if pd.isna(origin):
        return None, 'no_origin'

    candidate = origin + pd.DateOffset(months=int(cycle))
    guard = 0
    while candidate < as_of_ts and guard < 1200:
        candidate = candidate + pd.DateOffset(months=int(cycle))
        guard += 1
    return candidate, 'estimated'


# 대한민국 법정공휴일 + 대체공휴일 (은행 휴무일 기준). 이자지급일이 이 날짜나 주말과 겹치면
# 다음 영업일로 넘겨서 실제 지급일을 계산하는 데 사용합니다.
# ※ 매년 갱신이 필요합니다 (설날·추석·부처님오신날 등은 음력 기준이라 해마다 날짜가 바뀝니다).
KR_PUBLIC_HOLIDAYS = {
    # 2026년
    '2026-01-01',  # 신정
    '2026-02-16', '2026-02-17', '2026-02-18',  # 설날 연휴
    '2026-03-01', '2026-03-02',  # 삼일절 + 대체공휴일
    '2026-05-01',  # 노동절(근로자의날) - 은행 휴무
    '2026-05-05',  # 어린이날
    '2026-05-24', '2026-05-25',  # 부처님오신날 + 대체공휴일
    '2026-06-06',  # 현충일
    '2026-07-17',  # 제헌절 (2026년부터 공휴일 재지정)
    '2026-08-15', '2026-08-17',  # 광복절 + 대체공휴일
    '2026-09-24', '2026-09-25', '2026-09-26',  # 추석 연휴
    '2026-10-03', '2026-10-05',  # 개천절 + 대체공휴일
    '2026-10-09',  # 한글날
    '2026-12-25',  # 성탄절
    # 2027년 (설날/추석 등 음력 기준일 포함, 대체공휴일은 확정 고시 전 추정치)
    '2027-01-01',  # 신정
    '2027-02-06', '2027-02-07', '2027-02-08', '2027-02-09',  # 설날 연휴 + 대체공휴일
    '2027-03-01',  # 삼일절
    '2027-05-01',  # 노동절
    '2027-05-05',  # 어린이날
    '2027-05-13',  # 부처님오신날
    '2027-06-06',  # 현충일
    '2027-07-17', '2027-07-19',  # 제헌절 + 대체공휴일(추정)
    '2027-08-15', '2027-08-16',  # 광복절 + 대체공휴일(추정)
    '2027-09-14', '2027-09-15', '2027-09-16',  # 추석 연휴
    '2027-10-03', '2027-10-04',  # 개천절 + 대체공휴일(추정)
    '2027-10-09', '2027-10-11',  # 한글날 + 대체공휴일(추정)
    '2027-12-25', '2027-12-27',  # 성탄절 + 대체공휴일(추정)
}


def roll_to_next_business_day(d):
    """날짜가 토/일요일이거나 KR_PUBLIC_HOLIDAYS에 포함되면, 그 다음 영업일까지 하루씩 넘깁니다.
    (예: 이자지급일이 토요일이면 다음 첫 영업일에 실제로 이자가 지급됩니다.)"""
    d = pd.Timestamp(d)
    guard = 0
    while (d.weekday() >= 5 or d.strftime('%Y-%m-%d') in KR_PUBLIC_HOLIDAYS) and guard < 14:
        d = d + pd.Timedelta(days=1)
        guard += 1
    return d


def build_rate_change_schedule(df_current, as_of_date=None):
    """변동금리(또는 정기 재산정 주기가 있는) 대출들의 '다음 금리변경 예정일'을 계산해
    임박한 순으로 정렬한 DataFrame을 반환합니다."""
    if as_of_date is None:
        as_of_date = datetime.now().date()
    as_of_ts = pd.Timestamp(as_of_date)

    is_total_row = df_current['금융기관'].astype(str).str.strip().isin(['계', '합계'])
    df = df_current[~is_total_row].copy()

    cycle_col = next((c for c in df.columns if '금리변경주기' in c), None)
    last_change_col = next((c for c in df.columns if '최근금리' in c and '변경일' in c), None)
    # '최초기표일'보다 '차입일'을 우선한다 — 금리변경주기는 '이 대출을 언제 실제로
    # 인출/재산정했는지'부터 세는 게 맞는데, 최초기표일은 종종 그 대출건이 속한
    # 한도(facility) 자체가 처음 설정된 훨씬 예전 날짜라 실제 금리 갱신 기준과 다르다
    # (실제로 우리은행 550억 건에서 최초기표일 기준으로 계산하면 다음 갱신이 14일 뒤
    # 임박한 것처럼 나왔는데, 실제 차입일(가장 최근 인출일) 기준으로 다시 계산하면
    # 1년 가까이 여유가 있는 게 맞았다 — 사용자가 실데이터로 확인해서 알려준 문제).
    # 차입일이 없는 행만 최초기표일로 대체한다.
    origin_col = '차입일' if '차입일' in df.columns else ('최초기표일' if '최초기표일' in df.columns else None)
    rate_col = next((c for c in df.columns if c.strip() == '금리'), None)
    balance_col = next((c for c in df.columns if '잔액' in c), None)
    loan_type_col = '대출과목' if '대출과목' in df.columns else None
    rate_kind_col = '금리구분' if '금리구분' in df.columns else None
    detail_cond_col = next((c for c in df.columns if '세부' in c and '금리' in c), None)

    if cycle_col is None or origin_col is None:
        return pd.DataFrame()

    rows = []
    for _, row in df.iterrows():
        balance = pd.to_numeric(row.get(balance_col), errors='coerce') if balance_col else None
        if pd.isna(balance) or balance == 0:
            continue  # 실제 잔액이 없는 건(미사용 한도 등)은 대상에서 제외

        # 행별로 차입일이 비어있거나 파싱이 안 되면(드묾) 최초기표일로 대체한다.
        origin_val = row.get(origin_col)
        if pd.isna(pd.to_datetime(origin_val, errors='coerce')) and '최초기표일' in df.columns and origin_col != '최초기표일':
            origin_val = row.get('최초기표일')

        change_date, basis = next_rate_change_date(
            row.get(cycle_col), row.get(last_change_col) if last_change_col else None,
            origin_val, as_of_ts
        )
        if change_date is None:
            continue

        # 12개월 주기는 대개 '고정금리'를 매년 갱신하는 형태라, 변동금리 위험(🔴/🟡)이 아니라
        # 갱신일 참고용으로 별도 표기한다. ('금리구분' 또는 '세부 금리조건'에 '고정'이 있으면 고정금리로 판정)
        rate_kind_txt = str(row.get(rate_kind_col, '')).strip() if rate_kind_col else ''
        detail_txt = str(row.get(detail_cond_col, '')).strip() if detail_cond_col else ''
        is_fixed = ('고정' in rate_kind_txt) or ('고정' in detail_txt)

        d_day = (change_date.normalize() - as_of_ts.normalize()).days
        rate = pd.to_numeric(row.get(rate_col), errors='coerce') if rate_col else None
        if is_fixed:
            status = f"🔒 고정금리 갱신(D{d_day:+d})"
        else:
            status = '🔴 임박(30일 이내)' if d_day <= 30 else ('🟡 관심(90일 이내)' if d_day <= 90 else '🟢 여유')
        rows.append({
            '금융기관': str(row['금융기관']).strip(),
            '대출과목': str(row.get(loan_type_col, '-')).strip() if loan_type_col else '-',
            '금리유형': '고정금리' if is_fixed else '변동금리',
            '현재금리': f"{rate*100:.2f}%" if pd.notna(rate) else '-',
            '변경주기': f"{int(pd.to_numeric(row.get(cycle_col), errors='coerce'))}개월",
            '다음 금리변경 예정일': change_date.strftime('%Y-%m-%d'),
            'D-day': d_day,
            '상태': status,
            '산출근거': '기재값' if basis == 'from_field' else '주기 기반 추정',
        })

    if not rows:
        return pd.DataFrame()

    return pd.DataFrame(rows).sort_values('D-day').reset_index(drop=True)


def build_expected_interest(df_current, as_of_date=None, default_pay_day=28, n_cycles=6):
    """'이자지급일' 컬럼을 기준으로 각 대출의 향후 n_cycles회 지급주기에 대한 예상이자를,
    신용카드 명세서처럼 '실제 지급일'별로 볼 수 있게 계산합니다.

    - 이자지급일이 비어 있으면 default_pay_day(기본 28일)를 가정합니다.
    - 이론상 지급일(이자계산 구간의 경계일)이 토/일요일이나 공휴일과 겹치면
      roll_to_next_business_day로 실제 지급일을 계산합니다(이자계산 일수 자체는 달라지지 않습니다).
    - 구간 중 금리변경 예정일이 있으면 나누어 계산하고, 그 이후 회차는 변경 후 금리가
      미확정이므로 현재금리로 잠정 산정했다는 안내를 비고에 남깁니다.

    반환: (대출×회차별 상세 DataFrame, 실제 지급일별 합계 DataFrame)
    """
    if as_of_date is None:
        as_of_date = datetime.now().date()
    as_of_ts = pd.Timestamp(as_of_date)

    is_total_row = df_current['금융기관'].astype(str).str.strip().isin(['계', '합계'])
    df = df_current[~is_total_row].copy()

    balance_col = next((c for c in df.columns if '잔액' in c), None)
    rate_col = next((c for c in df.columns if c.strip() == '금리'), None)
    pay_col = next((c for c in df.columns if '이자지급일' in c), None)
    cycle_col = next((c for c in df.columns if '금리변경주기' in c), None)
    last_change_col = next((c for c in df.columns if '최근금리' in c and '변경일' in c), None)
    origin_col = '최초기표일' if '최초기표일' in df.columns else ('차입일' if '차입일' in df.columns else None)
    loan_type_col = '대출과목' if '대출과목' in df.columns else None

    if balance_col is None or rate_col is None:
        return pd.DataFrame(), pd.DataFrame()

    # '이자 이력 관리'에 등록된 확정 스케줄(예: PCBO 사채)을 (대출명, 시작일, 종료일) ->
    # 확정이자 딕셔너리로 한 번만 통째로 읽어둔다 — 대출×회차마다 DB를 조회하면 호출
    # 횟수가 너무 많아지므로(대출건수 × n_cycles) 여기서 미리 메모리에 올려둔다.
    # list_fixed_schedule()의 각 행은 (id, loan_name, amount, rate, period_start, period_end, interest, note)
    fixed_schedule_lookup = {
        (r[1], r[4], r[5]): r[6] for r in interest_history_db.list_fixed_schedule()
    }

    detail_rows = []
    for _, row in df.iterrows():
        balance = pd.to_numeric(row.get(balance_col), errors='coerce')
        rate = pd.to_numeric(row.get(rate_col), errors='coerce')
        if pd.isna(balance) or balance == 0 or pd.isna(rate):
            continue  # 잔액/금리가 없는 건은 이자가 발생하지 않으므로 제외

        pay_text = row.get(pay_col) if pay_col else None
        day, cadence, raw_note = parse_interest_pay_schedule(pay_text)
        pay_note = ''
        if day is None:
            day = default_pay_day
            cadence = 1
            pay_note = f"⚠️ 이자지급일 미기재 → {default_pay_day}일(매월) 가정"
        elif cadence and cadence > 1:
            pay_note = f"ℹ️ 원문 '{raw_note}' 기준 {cadence}개월 주기 적용"

        anchor = row.get(origin_col) if origin_col else None

        # 다음 금리변경 예정일은 as_of 시점 기준으로 1회만 계산한다.
        # (그 이후 재변경 시점의 금리는 알 수 없으므로, 해당 회차부터는 전부 '현재금리 잠정치'로 처리)
        change_date = None
        if cycle_col is not None:
            change_date, _basis = next_rate_change_date(
                row.get(cycle_col), row.get(last_change_col) if last_change_col else None,
                row.get(origin_col) if origin_col else None, as_of_ts
            )

        institution_name = str(row['금융기관']).strip()

        cursor = as_of_ts
        for cycle_idx in range(1, n_cycles + 1):
            last_pay, next_pay = bracket_pay_dates(anchor, cadence or 1, day, cursor)
            days_total = (next_pay - last_pay).days
            if days_total <= 0:
                break
            cursor = next_pay  # 다음 회차 계산을 위해 커서 전진

            notes = [n for n in [pay_note] if n]
            # '이자 이력 관리'에 이 기관ㆍ이 기간과 정확히 일치하는 확정 스케줄(예: PCBO
            # 사채처럼 만기까지 미리 정해진 이자)이 등록돼 있으면, 아래 잔액×금리 어림 계산
            # 대신 그 확정 금액을 그대로 쓴다 — 확정 스케줄이 없는 구간은 기존 계산식 그대로.
            fixed_interest = fixed_schedule_lookup.get(
                (institution_name, last_pay.date().isoformat(), next_pay.date().isoformat())
            )
            if fixed_interest is not None:
                expected_interest = fixed_interest
                notes.append("✅ 확정 스케줄(이자 이력 관리) 반영 — 잔액×금리 어림 계산 아님")
            elif change_date is not None and last_pay < change_date < next_pay:
                d1 = (change_date - last_pay).days
                d2 = (next_pay - change_date).days
                expected_interest = balance * rate * d1 / 365 + balance * rate * d2 / 365
                notes.append(f"⚠️ {change_date:%Y-%m-%d} 금리변경 예정 구간 포함 (변경 후 금리 미확정 → 현재금리로 잠정 산정)")
            elif change_date is not None and last_pay >= change_date:
                expected_interest = balance * rate * days_total / 365
                notes.append(f"⚠️ {change_date:%Y-%m-%d} 금리변경 이후 기간 → 현재금리로 잠정 산정(실제와 다를 수 있음)")
            else:
                expected_interest = balance * rate * days_total / 365

            actual_pay_date = roll_to_next_business_day(next_pay)
            if actual_pay_date != next_pay:
                notes.append(f"📌 {next_pay:%Y-%m-%d}이 휴일/주말이라 {actual_pay_date:%Y-%m-%d}(영업일)에 실제 지급")

            detail_rows.append({
                '금융기관': institution_name,
                '대출과목': str(row.get(loan_type_col, '-')).strip() if loan_type_col else '-',
                '회차': cycle_idx,
                '잔액(원)': balance,
                '적용금리': f"{rate*100:.2f}%",
                '적용기간': f"{last_pay:%Y-%m-%d} ~ {next_pay:%Y-%m-%d}",
                '일수': days_total,
                '이론상 지급일': next_pay,
                '실제 지급일(영업일)': actual_pay_date,
                '예상이자(원)': round(expected_interest),
                '비고': ' / '.join(notes) if notes else '-',
            })

    if not detail_rows:
        return pd.DataFrame(), pd.DataFrame()

    detail_df = pd.DataFrame(detail_rows).sort_values(['실제 지급일(영업일)', '금융기관']).reset_index(drop=True)

    summary_df = (
        detail_df.groupby('실제 지급일(영업일)')
        .agg(예상이자합계=('예상이자(원)', 'sum'), 대출건수=('예상이자(원)', 'size'))
        .reset_index()
        .sort_values('실제 지급일(영업일)')
    )
    summary_df['실제 지급일(영업일)'] = summary_df['실제 지급일(영업일)'].dt.strftime('%Y-%m-%d')

    detail_df['이론상 지급일'] = detail_df['이론상 지급일'].dt.strftime('%Y-%m-%d')
    detail_df['실제 지급일(영업일)'] = detail_df['실제 지급일(영업일)'].dt.strftime('%Y-%m-%d')

    return detail_df, summary_df


def build_interest_outlook(df_current, as_of_date=None, months_ahead=12):
    """월별/연도별/기관별로 얼마씩 이자가 나가는지 한눈에 보기 위한 집계입니다.
    '잔액×현재금리'로 어림잡는 대신, build_expected_interest의 실제 지급주기·일수·영업일
    조정 로직을 그대로 재사용해 계산하므로 별도 입력 없이도 더 정확합니다. 다만 그 기간 안에
    금리변경이 예정된 구간은(비고에 표시됨) 변경 후 확정금리를 알 수 없어 현재금리로 잠정 산정됩니다.

    반환: (월별 DataFrame, 연도별 DataFrame, 금융기관별 DataFrame) — 모두 '예상이자합계' 컬럼 포함
    """
    if as_of_date is None:
        as_of_date = datetime.now().date()
    as_of_ts = pd.Timestamp(as_of_date)

    # 대출마다 지급주기가 달라(매월/분기 등) 몇 회차가 필요한지 다르므로, 넉넉히 계산한 뒤
    # 실제 필요한 기간(months_ahead)으로 잘라낸다.
    detail_df, _ = build_expected_interest(df_current, as_of_date, n_cycles=max(months_ahead + 3, 15))
    empty3 = (pd.DataFrame(), pd.DataFrame(), pd.DataFrame())
    if detail_df.empty:
        return empty3

    detail_df = detail_df.copy()
    detail_df['_지급일'] = pd.to_datetime(detail_df['실제 지급일(영업일)'], errors='coerce')
    cutoff = as_of_ts + pd.DateOffset(months=months_ahead)
    detail_df = detail_df[(detail_df['_지급일'] >= as_of_ts) & (detail_df['_지급일'] <= cutoff)]
    if detail_df.empty:
        return empty3

    detail_df['연월'] = detail_df['_지급일'].dt.strftime('%Y-%m')
    detail_df['연도'] = detail_df['_지급일'].dt.strftime('%Y')

    monthly_df = (
        detail_df.groupby('연월')['예상이자(원)'].sum()
        .reset_index().rename(columns={'예상이자(원)': '예상이자합계'}).sort_values('연월').reset_index(drop=True)
    )
    yearly_df = (
        detail_df.groupby('연도')['예상이자(원)'].sum()
        .reset_index().rename(columns={'예상이자(원)': '예상이자합계'}).sort_values('연도').reset_index(drop=True)
    )
    by_bank_df = (
        detail_df.groupby('금융기관')['예상이자(원)'].sum()
        .reset_index().rename(columns={'예상이자(원)': '예상이자합계'}).sort_values('예상이자합계', ascending=False).reset_index(drop=True)
    )

    return monthly_df, yearly_df, by_bank_df


def render_dashboard():
    # Initialize session state for KPI inputs if not present
    if 'kpi_inputs' not in st.session_state:
        st.session_state['kpi_inputs'] = load_kpi_values()
 
    st.markdown("""
        <style>
        @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&family=Outfit:wght@300;400;500;600;700&family=Noto+Sans+KR:wght@300;400;500;700;900&display=swap');
        @import url('https://cdn.jsdelivr.net/gh/orioncactus/pretendard/dist/web/static/pretendard.css');

        :root {
            --bg-color: #eef2f8;
            --card-bg: #ffffff;
            --card-border: rgba(15, 23, 42, 0.08);
            --text-primary: #0f172a;
            --text-secondary: #64748b;
            --text-muted: #94a3b8;
            --accent: #3b82f6;
            --accent-light: #2563eb;
            --accent-glow: rgba(59, 130, 246, 0.25);
            --success: #10b981;
            --success-light: #059669;
            --success-glow: rgba(16, 185, 129, 0.18);
            --warning: #f59e0b;
            --danger: #ef4444;
            --font-header: 'Outfit', 'Pretendard', 'Noto Sans KR', sans-serif;
            --font-body: 'Pretendard', 'Inter', 'Noto Sans KR', sans-serif;
            --blur-intensity: 16px;
        }

        html, body, [class*="css"] {
            font-family: var(--font-body);
        }

        /* ---- 전체 배경: 퇴직연금 대시보드와 통일된 라이트 글래스모피즘 테마 ---- */
        [data-testid="stAppViewContainer"] {
            background-color: var(--bg-color);
            background-image:
                radial-gradient(at 0% 0%, rgba(59, 130, 246, 0.07) 0px, transparent 50%),
                radial-gradient(at 100% 100%, rgba(16, 185, 129, 0.06) 0px, transparent 50%);
            background-attachment: fixed;
        }
        [data-testid="stHeader"] {
            background: transparent;
        }
        h1, h2, h3, h4, h5, h6 {
            font-family: var(--font-header);
            letter-spacing: -0.3px;
        }
        [data-testid="stAppViewBlockContainer"] h3,
        [data-testid="stAppViewBlockContainer"] h4 {
            border-left: 3px solid var(--accent);
            padding-left: 0.7rem;
        }

        /* ---- 상단 헤더: 퇴직연금 대시보드와 통일된 카드형 헤더 ---- */
        .app-main-header {
            display: flex;
            align-items: center;
            gap: 1rem;
            padding: 1.3rem 1.8rem;
            border-radius: 16px;
            background: var(--card-bg);
            border: 1px solid var(--card-border);
            backdrop-filter: blur(var(--blur-intensity));
            box-shadow: 0 4px 20px 0 rgba(15, 23, 42, 0.06);
            margin-bottom: 1.2rem;
        }
        .app-header-accent {
            width: 8px;
            height: 38px;
            flex-shrink: 0;
            background: linear-gradient(135deg, var(--accent), #06b6d4);
            border-radius: 4px;
            box-shadow: 0 0 12px var(--accent-glow);
        }
        .app-header-text h1 {
            font-family: var(--font-header);
            font-size: 1.5rem;
            font-weight: 700;
            letter-spacing: -0.5px;
            color: var(--text-primary);
            margin: 0;
            line-height: 1.3;
        }
        .app-header-text .app-header-subtitle {
            font-size: 0.8rem;
            color: var(--text-muted);
            font-family: var(--font-header);
            text-transform: uppercase;
            letter-spacing: 1px;
            margin-top: 0.15rem;
        }

        /* ---- 텍스트/숫자/비밀번호 입력창: 라이트 테마에서 배경/글자색을 명시적으로
           지정 — 관리자 비밀번호 입력창처럼 배경은 흰색인데 글자색이 안 바뀌어(흰
           글씨) 안 보이던 문제를 막는다. ---- */
        [data-testid="stTextInput"] input,
        [data-testid="stNumberInput"] input,
        [data-testid="stTextArea"] textarea {
            background-color: #f8fafc !important;
            color: var(--text-primary) !important;
            border: 1px solid var(--card-border) !important;
            -webkit-text-fill-color: var(--text-primary) !important;
        }

        /* ---- KPI/지표 카드: 글래스카드 스타일 ---- */
        .metric-card {
            background: var(--card-bg);
            padding: 1.5rem;
            border-radius: 16px;
            border: 1px solid var(--card-border);
            backdrop-filter: blur(var(--blur-intensity));
            -webkit-backdrop-filter: blur(var(--blur-intensity));
            color: var(--text-primary);
            box-shadow: 0 4px 20px 0 rgba(15, 23, 42, 0.06);
            transition: all 0.3s ease;
            position: relative;
            overflow: hidden;
        }
        .metric-card::before {
            content: '';
            position: absolute;
            top: 0; left: 0; width: 100%; height: 4px;
            background: linear-gradient(90deg, var(--accent), #0891b2);
        }
        .metric-card:hover {
            transform: translateY(-4px);
            border-color: rgba(15, 23, 42, 0.14);
            box-shadow: 0 12px 32px 0 rgba(59, 130, 246, 0.15);
        }
        .metric-label { font-size: 0.85rem; color: var(--text-secondary); font-weight: 600; text-transform: uppercase; letter-spacing: 0.5px; margin-bottom: 0.5rem; }
        .metric-value { font-family: var(--font-header); font-size: 2.0rem; font-weight: 700; color: var(--text-primary); letter-spacing: -0.5px; }
        .metric-desc { font-size: 0.8rem; margin-top: 0.5rem; font-weight: 500; color: var(--text-muted); }
        .positive { color: var(--success-light) !important; }
        .negative { color: var(--danger) !important; }
        .neutral { color: var(--text-muted) !important; }

        /* ---- 본문 내 st.container(border=True) 영역을 글래스카드로 ---- */
        [data-testid="stVerticalBlockBorderWrapper"] {
            background: var(--card-bg);
            border: 1px solid var(--card-border) !important;
            border-radius: 16px !important;
            backdrop-filter: blur(var(--blur-intensity));
            -webkit-backdrop-filter: blur(var(--blur-intensity));
            box-shadow: 0 4px 20px 0 rgba(15, 23, 42, 0.05);
        }

        /* ---- Sidebar: 퇴직연금과 통일된 라이트 글래스 사이드바 ---- */
        [data-testid="stSidebar"] {
            background: rgba(255, 255, 255, 0.9) !important;
            backdrop-filter: blur(var(--blur-intensity));
            color: var(--text-primary) !important;
            border-right: 1px solid var(--card-border);
        }
        /* ---- 사이드바 스크롤 고정 수정 ---- */
        section[data-testid="stSidebar"] {
            overflow-y: auto !important;
            height: 100vh !important;
        }
        section[data-testid="stSidebar"] > div {
            overflow-y: auto !important;
            max-height: 100vh !important;
            padding-bottom: 4rem !important;
        }
        [data-testid="stSidebarUserContent"] {
            overflow-y: auto !important;
            padding-bottom: 4rem !important;
        }
        [data-testid="stSidebarContent"] {
            overflow-y: auto !important;
            padding-bottom: 4rem !important;
        }
        [data-testid="stSidebar"] div,
        [data-testid="stSidebar"] span,
        [data-testid="stSidebar"] label,
        [data-testid="stSidebar"] p,
        [data-testid="stSidebar"] h1,
        [data-testid="stSidebar"] h2,
        [data-testid="stSidebar"] h3,
        [data-testid="stSidebar"] h4,
        [data-testid="stSidebar"] h5,
        [data-testid="stSidebar"] h6 {
            color: var(--text-primary) !important;
        }
        /* Style ECOS API JSON text rendering in sidebar to have high contrast */
        [data-testid="stSidebar"] pre {
            background-color: rgba(15, 23, 42, 0.04) !important;
            color: var(--text-primary) !important;
            border: 1px solid var(--card-border) !important;
        }
        /* Style expander headers in sidebar */
        [data-testid="stSidebar"] details {
            background-color: rgba(15, 23, 42, 0.03) !important;
            border: 1px solid var(--card-border) !important;
            border-radius: 6px !important;
            margin-bottom: 8px !important;
        }
        /* Style forms and buttons in sidebar */
        [data-testid="stSidebar"] .stForm {
            background-color: rgba(15, 23, 42, 0.03) !important;
            border: 1px solid var(--card-border) !important;
            border-radius: 8px !important;
            padding: 12px !important;
        }
        [data-testid="stSidebar"] button {
            background-color: var(--accent) !important;
            color: white !important;
            border: none !important;
        }
        [data-testid="stSidebar"] button:hover {
            background-color: var(--accent-light) !important;
            color: white !important;
            box-shadow: 0 0 10px var(--accent-glow);
        }
        [data-testid="stSidebar"] [data-testid="stFileUploaderDropzone"] {
            background-color: rgba(15, 23, 42, 0.03) !important;
            border: 1px dashed var(--card-border) !important;
        }
        /* 사이드바 라디오(파일 업로드 관련 옵션 등) 항목이 길어도 줄바꿈되어 잘리지 않고,
           항목 간 높이가 고르게 정렬되도록 통일 */
        [data-testid="stSidebar"] [role="radiogroup"] label {
            white-space: normal !important;
            word-break: keep-all;
            line-height: 1.35;
            padding: 6px 4px;
            align-items: flex-start !important;
        }
        [data-testid="stSidebar"] [role="radiogroup"] label > div:first-child {
            margin-top: 2px;
        }
        [data-testid="stSidebar"] [role="radiogroup"] label:has(input:checked) {
            background: rgba(59, 130, 246, 0.12);
            border-radius: 8px;
        }

        /* ---- 본문 상단 탭형 메뉴(대분류/소분류 st.radio(horizontal=True)) ---- */
        [data-testid="stAppViewBlockContainer"] div[role="radiogroup"] {
            gap: 8px;
            row-gap: 8px;
        }
        [data-testid="stAppViewBlockContainer"] div[role="radiogroup"] label {
            background: var(--card-bg);
            border: 1px solid var(--card-border);
            border-radius: 20px;
            padding: 6px 16px;
            margin: 0 !important;
            transition: all 0.2s ease;
        }
        [data-testid="stAppViewBlockContainer"] div[role="radiogroup"] label:hover {
            border-color: var(--accent);
            background: rgba(59, 130, 246, 0.12);
        }
        [data-testid="stAppViewBlockContainer"] div[role="radiogroup"] label:has(input:checked) {
            background: var(--accent) !important;
            border-color: var(--accent) !important;
            box-shadow: 0 0 12px var(--accent-glow);
        }
        [data-testid="stAppViewBlockContainer"] div[role="radiogroup"] label:has(input:checked) p {
            color: #ffffff !important;
            font-weight: 700 !important;
        }
        /* 라디오 원형 표시는 숨기고 pill 버튼처럼 텍스트만 보이게 */
        [data-testid="stAppViewBlockContainer"] div[role="radiogroup"] label > div:first-child {
            display: none;
        }


        /* Prettier tabs */
        .stTabs [data-baseweb="tab-list"] {
            gap: 8px;
        }
        .stTabs [data-baseweb="tab"] {
            background-color: rgba(15, 23, 42, 0.04);
            border-radius: 6px 6px 0px 0px;
            padding: 8px 16px;
            color: var(--text-secondary);
            font-weight: 500;
        }
        .stTabs [aria-selected="true"] {
            background-color: var(--accent) !important;
            color: white !important;
            font-weight: 700;
        }

        /* ---- 인쇄용 요약 화면 스타일 ---- */
        .print-page {
            background: #ffffff;
            color: #111827;
            padding: 24px;
            border-radius: 8px;
        }
        .print-page h1, .print-page h2, .print-page h3, .print-page h4 {
            color: #111827;
        }
        .print-title {
            font-size: 1.6rem;
            font-weight: 800;
            border-bottom: 3px solid #111827;
            padding-bottom: 8px;
            margin-bottom: 4px;
        }
        .print-sub {
            color: #4b5563;
            font-size: 0.85rem;
            margin-bottom: 16px;
        }
        .print-section-title {
            font-size: 1.05rem;
            font-weight: 700;
            margin-top: 20px;
            margin-bottom: 8px;
            border-left: 4px solid #3b82f6;
            padding-left: 8px;
        }
        .print-kpi-grid {
            display: grid;
            grid-template-columns: repeat(4, 1fr);
            gap: 10px;
            margin-bottom: 8px;
        }
        .print-kpi-box {
            border: 1px solid #d1d5db;
            border-radius: 6px;
            padding: 10px;
        }
        .print-kpi-box .lbl { font-size: 0.75rem; color: #6b7280; }
        .print-kpi-box .val { font-size: 1.2rem; font-weight: 800; color: #111827; }
        table.print-table {
            width: 100%;
            border-collapse: collapse;
            font-size: 0.82rem;
            margin-bottom: 12px;
        }
        table.print-table th, table.print-table td {
            border: 1px solid #999999;
            padding: 5px 8px;
            text-align: left;
        }
        table.print-table th {
            background-color: #f3f4f6;
            font-weight: 700;
        }

        /* ---- 월간보고서 원본 재현(표지/관리/대여/운용/차입/기타) ---- */
        .report-page {
            background: #ffffff;
            color: #111827;
            padding: 20px;
            border-radius: 8px;
            margin-bottom: 28px;
            overflow-x: auto;
        }
        .report-page-title {
            font-size: 1.1rem;
            font-weight: 800;
            color: #111827;
            border-bottom: 3px solid #111827;
            padding-bottom: 6px;
            margin-bottom: 12px;
        }
        table.report-sheet {
            border-collapse: collapse;
            font-size: 0.72rem;
            white-space: nowrap;
        }
        table.report-sheet td {
            padding: 3px 6px;
            color: #111827;
        }

        /* ---- 부동산 자산 현황(담보현황 탭) ---- */
        .real-estate-table th, .real-estate-table td {
            border: 1px solid var(--card-border);
            padding: 4px 7px;
            color: var(--text-primary);
            word-break: keep-all;
        }
        .real-estate-table th {
            font-weight: 700;
        }

        @media print {
            .real-estate-table {
                font-size: 8.5px !important;
            }
            .real-estate-table th, .real-estate-table td {
                padding: 2px 4px !important;
            }
            [data-testid="stSidebar"],
            [data-testid="stHeader"],
            [data-testid="stToolbar"],
            [data-testid="stRadio"],
            .no-print {
                display: none !important;
            }
            @page {
                size: A4 portrait;
                margin: 14mm;
            }
            .print-page {
                -webkit-print-color-adjust: exact;
                print-color-adjust: exact;
            }
            @page report-landscape {
                size: A4 landscape;
                margin: 8mm;
            }
            .report-page {
                page: report-landscape;
                -webkit-print-color-adjust: exact;
                print-color-adjust: exact;
                page-break-after: always;
            }
            .report-page:last-child {
                page-break-after: auto;
            }
        }
        </style>
    """, unsafe_allow_html=True)
 
    st.markdown("""
        <div class="app-main-header">
            <div class="app-header-accent"></div>
            <div class="app-header-text">
                <h1>📊 (주)동양 차입관리 통합프로그램</h1>
                <div class="app-header-subtitle">상장사 자금팀 의사결정 지원 시스템 · 한국은행 ECOS 연동</div>
            </div>
        </div>
    """, unsafe_allow_html=True)

    bok_rates = fetch_ecos_base_rate()

    # -----------------------------------------------------------------
    # 화면 내비게이션: 상단에 대분류(차입금관리/차입 준비/차입 분석/인쇄용 요약/데이터 관리) 탭을 두고,
    # 그 아래 각 대분류에 속한 소분류 탭에서 고른 화면 1개만 표시합니다.
    # (Streamlit의 st.tabs는 key를 지원하지 않아, 화면 내 다른 위젯을 조작해 리런이
    #  발생하면 항상 첫 번째 탭으로 되돌아가는 문제가 있습니다. 그래서 st.radio를
    #  horizontal=True로 두고 탭처럼 스타일링해, 선택 상태가 세션에 유지되도록 했습니다.
    #  '데이터 관리'는 데이터가 하나도 없어도(최초 설정 시) 들어갈 수 있어야 하므로,
    #  메뉴 자체는 아래 '조회 월/분기/연 선택' 게이트보다 먼저 그립니다.)
    # -----------------------------------------------------------------
    PAGE_KPI = "🏆 KPI 요약"
    PAGE_RATE_TREND = "📊 금리분석(금리추이)"
    PAGE_RATE_BANK = "🏢 금융기관별 금리(연도별)"
    PAGE_CREDIT = "🔍 신용등급 변동 추이 (5개년)"
    PAGE_LOAN = "📋 차입금현황표 (현재)"
    PAGE_INTEREST = "💵 원리금관리"
    PAGE_MATURITY = "📅 1년이내 만기도래 및 상환계획표"
    PAGE_ANCILLARY = "💰 부수거래 연동 대출금리 영향도"
    PAGE_COLLATERAL = "🏗 담보현황"
    PAGE_PRINT = "🖨 인쇄용 요약 (PDF)"
    PAGE_DATA_ADMIN = "🗄 데이터 관리(관리자)"
    PAGE_REPORT_ADMIN = "📑 월간보고서 업로드(관리자)"
    PAGE_INTEREST_ADMIN = "💵 이자 이력 관리(관리자)"
    PAGE_GROUP_LOAN = "🏢 그룹 차입금 분석"
    PAGE_GROUP_LOAN_ADMIN = "🏢 그룹 차입금 업로드(관리자)"

    # 월간보고서 원본(표지/관리/대여/운용/차입/기타) 재현 화면 — 통합포털 허브의 '자금' 하위
    # 탭(대여/운용/관리/차입/기타)에서 iframe으로 이 화면들을 직접 가리킵니다.
    PAGE_REPORT_ALL = "🖨 전체 출력(표지부터)"
    PAGE_REPORT_COVER = "📕 표지"
    PAGE_REPORT_MANAGE = "🗂 관리"
    PAGE_REPORT_LEND = "💸 대여"
    PAGE_REPORT_INVEST = "📈 운용"
    PAGE_REPORT_BORROW = "🏦 차입"
    PAGE_REPORT_OTHER = "📝 기타(어음)"

    # '📑 월간보고서' 그룹은 이 화면(Streamlit) 안에서는 더 이상 메뉴로 노출하지 않는다 —
    # 통합포털 허브의 '자금' > '월간보고서' 탭이 같은 내용을 눈금 없는 뷰 + A4 인쇄
    # 형태로 이미 재현하고 있어서(finance_view.html), 여기 남겨두면 '차입' 탭을 눌렀을 때
    # 엉뚱하게 이 보고서 화면이 뜨는 것처럼 보여 혼란만 준다. 페이지 렌더링 코드
    # (REPORT_VIEW_PAGES 블록)는 그대로 남아있지만 nav에서 안 보이니 실행되지 않는다.
    NAV_GROUPS = {
        "💼 차입금관리": [PAGE_LOAN, PAGE_INTEREST, PAGE_MATURITY],
        "🛡 차입 준비": [PAGE_COLLATERAL],
        "📈 차입 분석": [PAGE_RATE_TREND, PAGE_RATE_BANK, PAGE_ANCILLARY, PAGE_CREDIT, PAGE_KPI],
        "🖨 인쇄용 요약": [PAGE_PRINT],
        "🏢 그룹 차입금": [PAGE_GROUP_LOAN],
        "⚙️ 데이터 관리": [PAGE_DATA_ADMIN, PAGE_REPORT_ADMIN, PAGE_GROUP_LOAN_ADMIN, PAGE_INTEREST_ADMIN],
    }
    GROUP_NAMES = list(NAV_GROUPS.keys())

    # 통합포털 허브에서 ?report_page=lend 같은 딥링크로 들어온 경우, 그 화면을 기본 선택으로
    # 미리 세팅해둡니다(아래 세션 기본값 로직보다 먼저 실행돼야 적용됨).
    REPORT_DEEP_LINK_MAP = {
        'all': PAGE_REPORT_ALL, 'cover': PAGE_REPORT_COVER, 'manage': PAGE_REPORT_MANAGE,
        'lend': PAGE_REPORT_LEND, 'invest': PAGE_REPORT_INVEST, 'borrow': PAGE_REPORT_BORROW,
        'other': PAGE_REPORT_OTHER,
    }
    _deep_link = st.query_params.get('report_page')
    if _deep_link in REPORT_DEEP_LINK_MAP and 'nav_group' not in st.session_state:
        st.session_state['nav_group'] = "📑 월간보고서"
        st.session_state['nav_sub__📑 월간보고서'] = REPORT_DEEP_LINK_MAP[_deep_link]

    if 'nav_group' not in st.session_state or st.session_state['nav_group'] not in GROUP_NAMES:
        st.session_state['nav_group'] = GROUP_NAMES[0]

    nav_group = st.radio(
        "메뉴 대분류", GROUP_NAMES, key="nav_group", horizontal=True, label_visibility="collapsed"
    )

    sub_options = NAV_GROUPS[nav_group]
    sub_key = f"nav_sub__{nav_group}"
    if sub_key not in st.session_state or st.session_state[sub_key] not in sub_options:
        st.session_state[sub_key] = sub_options[0]

    if len(sub_options) > 1:
        page = st.radio(
            "메뉴 소분류", sub_options, key=sub_key, horizontal=True, label_visibility="collapsed"
        )
    else:
        page = sub_options[0]

    st.markdown("---")

    # -----------------------------------------------------------------
    # 🗄 데이터 관리(관리자) — 등록된 데이터가 하나도 없어도 들어갈 수 있어야 하는
    # 화면이라, 아래 '조회 월/분기/연 선택' 게이트보다 먼저 처리하고 여기서 끝냅니다.
    # -----------------------------------------------------------------
    if page == PAGE_DATA_ADMIN:
        st.subheader("🗄 데이터 관리(관리자)")
        st.write(
            "월별 '차입금 관리내역' 엑셀을 여기서 등록해두면, 다른 사용자는 파일 업로드 없이 "
            "왼쪽에서 월/분기/연 단위로 조회할 수 있습니다."
        )

        try:
            correct_pw = st.secrets["admin"]["password"]
        except Exception:
            correct_pw = None

        if 'admin_authed' not in st.session_state:
            st.session_state['admin_authed'] = False

        if not correct_pw:
            st.error(
                "관리자 비밀번호가 설정되어 있지 않습니다. 프로그램 폴더의 .streamlit/secrets.toml 파일에서 "
                "[admin] password 값을 설정해주세요."
            )
        elif not st.session_state['admin_authed']:
            pw_input = st.text_input("관리자 비밀번호", type="password", key="admin_pw_input")
            if st.button("확인", type="primary"):
                if pw_input == correct_pw:
                    st.session_state['admin_authed'] = True
                    st.rerun()
                else:
                    st.error("비밀번호가 올바르지 않습니다.")
            st.caption("비밀번호는 .streamlit/secrets.toml 파일에서 바꿀 수 있습니다.")
        else:
            st.success("✅ 관리자 인증됨")
            if st.button("로그아웃"):
                st.session_state['admin_authed'] = False
                st.rerun()

            st.markdown("---")
            st.markdown("#### ➕ 월별 데이터 등록 / 교체")
            col_y, col_q = st.columns(2)
            with col_y:
                admin_year = st.number_input(
                    "연도", min_value=2020, max_value=2100, value=datetime.now().year, step=1, key="admin_year"
                )
            with col_q:
                admin_month = st.selectbox(
                    "월", list(range(1, 13)), index=datetime.now().month - 1, key="admin_month"
                )

            existing_keys = {(s[1], s[2]): s[0] for s in snapshot_db.list_snapshots()}
            selected_key = (int(admin_year), int(admin_month))
            if selected_key in existing_keys:
                st.warning(f"⚠️ {admin_year}년 {admin_month}월 데이터가 이미 있습니다. 저장하면 기존 데이터를 덮어씁니다.")

            input_mode = st.radio(
                "입력 방식", ["📤 엑셀 업로드", "✏️ 직접입력(표에서 편집)"], horizontal=True, key="admin_input_mode"
            )

            if input_mode == "📤 엑셀 업로드":
                admin_file = st.file_uploader(
                    "차입금 관리내역 엑셀 파일 (.xlsx)", type=["xlsx"], key="admin_uploader",
                    help="예: ★ 차입금_관리내역.xlsx"
                )
                if st.button("💾 저장", type="primary", disabled=admin_file is None, key="admin_save_upload"):
                    snapshot_db.save_snapshot(int(admin_year), int(admin_month), admin_file.name, admin_file.getvalue())
                    st.success(f"{admin_year}년 {admin_month}월 데이터를 저장했습니다.")
                    st.rerun()
            else:
                st.caption(
                    "은행/기간/금액/잔액/금리/담보/비고 등을 엑셀처럼 표에서 직접 입력·수정합니다. "
                    "표 맨 아래 빈 행에 값을 입력하면 행이 추가되고, 행을 선택한 뒤 Delete 키로 삭제할 수 있습니다. "
                    "💡 Excel에서 표를 복사(Ctrl+C)한 뒤 이 표 안의 셀을 클릭하고 Ctrl+V로 붙여넣으면 여러 행을 한 번에 입력할 수 있습니다."
                )
                default_columns = [
                    '금융기관', '대출과목', '금리구분', '기간(시작)', '기간(종료)',
                    '금액(한도)', '잔액', '금리(%)', '담보', '비고',
                ]
                edit_source_bytes = None
                if selected_key in existing_keys:
                    edit_source_bytes = snapshot_db.get_snapshot_bytes(existing_keys[selected_key])
                elif existing_keys:
                    latest_key = max(existing_keys.keys())
                    edit_source_bytes = snapshot_db.get_snapshot_bytes(existing_keys[latest_key])

                if edit_source_bytes is not None:
                    try:
                        start_df, _ = load_current_loan_status(edit_source_bytes)
                    except Exception:
                        start_df = pd.DataFrame(columns=default_columns)
                    if start_df.empty or '금융기관' not in start_df.columns:
                        start_df = pd.DataFrame(columns=default_columns)
                else:
                    start_df = pd.DataFrame(columns=default_columns)

                if st.session_state.get('admin_manual_df_key') != selected_key:
                    st.session_state['admin_manual_df'] = start_df.copy()
                    st.session_state['admin_manual_df_key'] = selected_key

                edited_df = st.data_editor(
                    st.session_state['admin_manual_df'],
                    num_rows="dynamic",
                    use_container_width=True,
                    key="admin_manual_editor",
                )

                if st.button("💾 저장", type="primary", key="admin_save_manual"):
                    clean_df = edited_df.dropna(how='all')
                    if clean_df.empty or '금융기관' not in clean_df.columns:
                        st.error("최소 1개 행 이상 입력해주세요 ('금융기관' 열 필수).")
                    else:
                        wb_bytes = build_loan_status_workbook(clean_df)
                        snapshot_db.save_snapshot(int(admin_year), int(admin_month), "직접입력.xlsx", wb_bytes)
                        st.session_state['admin_manual_df'] = clean_df
                        st.success(f"{admin_year}년 {admin_month}월 데이터를 저장했습니다.")
                        st.rerun()

            st.markdown("---")
            st.markdown("#### 📋 등록된 월 목록")
            admin_rows = snapshot_db.list_snapshots()
            if not admin_rows:
                st.caption("등록된 데이터가 없습니다.")
            else:
                admin_list_df = pd.DataFrame([
                    {
                        "연도": r[1], "월": r[2], "파일명": r[4],
                        "등록일시": r[5], "크기(KB)": round(r[6] / 1024, 1),
                    }
                    for r in admin_rows
                ])
                st.dataframe(admin_list_df, use_container_width=True, hide_index=True)

                del_options = {f"{r[3]} · {r[4]} ({r[5]})": r[0] for r in admin_rows}
                del_label = st.selectbox("삭제할 월 선택", list(del_options.keys()), key="admin_delete_pick")
                if st.button("🗑 선택한 월 삭제"):
                    snapshot_db.delete_snapshot(del_options[del_label])
                    st.success("삭제했습니다.")
                    st.rerun()
        return

    # -----------------------------------------------------------------
    # 📑 월간보고서 업로드(관리자) — 월간보고서 데이터가 하나도 없어도 들어갈 수 있어야 하는
    # 화면이라, 아래 조회 화면들보다 먼저 처리합니다.
    # -----------------------------------------------------------------
    if page == PAGE_REPORT_ADMIN:
        st.subheader("📑 월간보고서 업로드(관리자)")
        st.write(
            "'월간보고서_보고용' 엑셀 원본(표지·자금현황·사업본부·대여금·자금운용·차입금 및 담보현황·"
            "어음현황 시트 포함)을 월별로 등록해두면, '📑 월간보고서' 메뉴에서 표지부터 그대로 조회·출력할 수 있습니다."
        )

        try:
            correct_pw = st.secrets["admin"]["password"]
        except Exception:
            correct_pw = None

        if 'report_admin_authed' not in st.session_state:
            st.session_state['report_admin_authed'] = False

        if not correct_pw:
            st.error(
                "관리자 비밀번호가 설정되어 있지 않습니다. 프로그램 폴더의 .streamlit/secrets.toml 파일에서 "
                "[admin] password 값을 설정해주세요."
            )
        elif not st.session_state['report_admin_authed']:
            pw_input = st.text_input("관리자 비밀번호", type="password", key="report_admin_pw_input")
            if st.button("확인", type="primary", key="report_admin_pw_confirm"):
                if pw_input == correct_pw:
                    st.session_state['report_admin_authed'] = True
                    st.rerun()
                else:
                    st.error("비밀번호가 올바르지 않습니다.")
        else:
            st.success("✅ 관리자 인증됨")
            if st.button("로그아웃", key="report_admin_logout"):
                st.session_state['report_admin_authed'] = False
                st.rerun()

            st.markdown("---")
            st.markdown("#### ➕ 월간보고서 등록 / 교체")
            current_month = datetime.now().month
            col_y, col_m = st.columns(2)
            with col_y:
                report_year = st.number_input(
                    "연도", min_value=2020, max_value=2100, value=datetime.now().year, step=1, key="report_year"
                )
            with col_m:
                report_month = st.selectbox("월", list(range(1, 13)), index=current_month - 1, key="report_month")

            report_existing_keys = {(r[1], r[2]) for r in monthly_report_db.list_reports()}
            if (int(report_year), int(report_month)) in report_existing_keys:
                st.warning(f"⚠️ {report_year}년 {report_month}월 월간보고서가 이미 있습니다. 저장하면 기존 파일을 덮어씁니다.")

            report_input_mode = st.radio(
                "입력 방식", ["📤 엑셀 업로드", "✏️ 직접입력(표에서 편집)"], horizontal=True, key="report_input_mode"
            )

            REPORT_LEAF_SHEETS = {
                "💸 대여 (대여금 현황)": "대여금 현황",
                "📈 운용 (자금운용현황)": "자금운용현황",
                "🏦 차입 (차입금 및 담보현황)": "차입금 및 담보현황",
            }

            if report_input_mode == "📤 엑셀 업로드":
                st.caption("표지·관리(자금현황/사업본부/건재/건설/플랜트/공통)를 포함한 전체 원본을 한 번에 등록합니다.")
                report_file = st.file_uploader(
                    "월간보고서 엑셀 파일 (.xlsx)", type=["xlsx"], key="report_uploader",
                    help="예: 2026년 06월 월간보고서_보고용_최종.xlsx"
                )
                if st.button("💾 저장", type="primary", disabled=report_file is None, key="report_save_upload"):
                    monthly_report_db.save_report(int(report_year), int(report_month), report_file.name, report_file.getvalue())
                    st.success(f"{report_year}년 {report_month}월 월간보고서를 저장했습니다.")
                    st.rerun()
            else:
                st.caption(
                    "표지·관리(사업본부/건재/건설/플랜트/공통)·기타(어음현황)는 서식이 복잡해 직접입력을 "
                    "지원하지 않고, 위 '📤 엑셀 업로드'로만 등록할 수 있습니다. 아래 표는 대여/운용/차입 "
                    "3개 항목만 개별로 입력·수정합니다 — 저장해도 같은 달의 다른 시트는 그대로 유지됩니다. "
                    "Excel에서 표를 복사(Ctrl+C)한 뒤 표 안을 클릭하고 Ctrl+V로 붙여넣으면 여러 행을 한 번에 입력할 수 있습니다."
                )
                leaf_choice = st.selectbox("입력할 항목 선택", list(REPORT_LEAF_SHEETS.keys()), key="report_leaf_choice")
                target_sheet_name = REPORT_LEAF_SHEETS[leaf_choice]

                report_selected_key = (int(report_year), int(report_month))
                existing_report_id = next(
                    (r[0] for r in monthly_report_db.list_reports() if (r[1], r[2]) == report_selected_key), None
                )
                existing_report_bytes = (
                    monthly_report_db.get_report_bytes(existing_report_id) if existing_report_id else None
                )

                start_leaf_df = pd.DataFrame()
                if existing_report_bytes is not None:
                    try:
                        existing_wb_for_edit = openpyxl.load_workbook(io.BytesIO(existing_report_bytes), data_only=True)
                        if target_sheet_name in existing_wb_for_edit.sheetnames:
                            start_leaf_df = flatten_sheet_to_df(existing_wb_for_edit[target_sheet_name])
                    except Exception:
                        start_leaf_df = pd.DataFrame()

                leaf_state_key = f'report_manual_df__{report_selected_key}__{target_sheet_name}'
                if leaf_state_key not in st.session_state:
                    st.session_state[leaf_state_key] = start_leaf_df.copy()

                edited_leaf_df = st.data_editor(
                    st.session_state[leaf_state_key],
                    num_rows="dynamic",
                    use_container_width=True,
                    key=f'report_manual_editor__{report_selected_key}__{target_sheet_name}',
                )

                if st.button("💾 저장", type="primary", key="report_save_manual"):
                    clean_leaf_df = edited_leaf_df.dropna(how='all')
                    if clean_leaf_df.empty:
                        st.error("최소 1개 행 이상 입력해주세요.")
                    else:
                        if existing_report_bytes is not None:
                            try:
                                target_wb = openpyxl.load_workbook(io.BytesIO(existing_report_bytes))
                            except Exception:
                                target_wb = openpyxl.Workbook()
                                target_wb.remove(target_wb.active)
                        else:
                            target_wb = openpyxl.Workbook()
                            target_wb.remove(target_wb.active)
                        write_df_as_sheet(target_wb, target_sheet_name, clean_leaf_df)
                        buf = io.BytesIO()
                        target_wb.save(buf)
                        monthly_report_db.save_report(
                            int(report_year), int(report_month), f"{target_sheet_name}_직접입력.xlsx", buf.getvalue()
                        )
                        st.session_state[leaf_state_key] = clean_leaf_df
                        st.success(f"{report_year}년 {report_month}월 '{target_sheet_name}' 시트를 저장했습니다.")
                        st.rerun()

            st.markdown("---")
            st.markdown("#### 📋 등록된 월간보고서 목록")
            report_admin_rows = monthly_report_db.list_reports()
            if not report_admin_rows:
                st.caption("등록된 월간보고서가 없습니다.")
            else:
                report_list_df = pd.DataFrame([
                    {
                        "연도": r[1], "월": r[2], "파일명": r[4],
                        "등록일시": r[5], "크기(KB)": round(r[6] / 1024, 1),
                    }
                    for r in report_admin_rows
                ])
                st.dataframe(report_list_df, use_container_width=True, hide_index=True)

                report_del_options = {f"{r[3]} · {r[4]} ({r[5]})": r[0] for r in report_admin_rows}
                report_del_label = st.selectbox("삭제할 월 선택", list(report_del_options.keys()), key="report_delete_pick")
                if st.button("🗑 선택한 월간보고서 삭제", key="report_delete_btn"):
                    monthly_report_db.delete_report(report_del_options[report_del_label])
                    st.success("삭제했습니다.")
                    st.rerun()
        return

    # -----------------------------------------------------------------
    # 💵 이자 이력 관리(관리자) — "28일 이자.xlsx"처럼 연도별 시트에 은행별 이자계산
    # 구간ㆍ이자금액이 정리된 워크북을 그대로 업로드하면 과거 실적(actual_interest)으로
    # 저장하고, PCBO 사채처럼 만기까지 확정된 지급 스케줄은 별도 시트(이름에 '이자지급일정'
    # 포함)에서 fixed_schedule로 저장합니다. '💵 원리금관리' 화면은 과거 실적을 연도별로
    # 그대로 보여주고, 앞으로의 예상이자 계산 중 이 확정 스케줄과 기관ㆍ기간이 정확히 겹치는
    # 구간이 있으면 잔액×금리 어림 계산 대신 그 확정 금액을 그대로 씁니다.
    # -----------------------------------------------------------------
    if page == PAGE_INTEREST_ADMIN:
        st.subheader("💵 이자 이력 관리(관리자)")
        st.write(
            "'28일 이자.xlsx' 같은 워크북(연도별 시트에 은행ㆍ이자계산구간ㆍ이자금액이 정리된 표, "
            "PCBO 등 확정 스케줄은 '○○ 이자지급일정' 시트)을 업로드하면 자동으로 과거 실적ㆍ확정 "
            "스케줄로 나눠 저장합니다. '💵 원리금관리' 화면에 반영됩니다."
        )

        try:
            correct_pw = st.secrets["admin"]["password"]
        except Exception:
            correct_pw = None

        if 'interest_admin_authed' not in st.session_state:
            st.session_state['interest_admin_authed'] = False

        if not correct_pw:
            st.error(
                "관리자 비밀번호가 설정되어 있지 않습니다. 프로그램 폴더의 .streamlit/secrets.toml 파일에서 "
                "[admin] password 값을 설정해주세요."
            )
        elif not st.session_state['interest_admin_authed']:
            pw_input = st.text_input("관리자 비밀번호", type="password", key="interest_admin_pw_input")
            if st.button("확인", type="primary", key="interest_admin_pw_confirm"):
                if pw_input == correct_pw:
                    st.session_state['interest_admin_authed'] = True
                    st.rerun()
                else:
                    st.error("비밀번호가 올바르지 않습니다.")
        else:
            st.success("✅ 관리자 인증됨")
            if st.button("로그아웃", key="interest_admin_logout"):
                st.session_state['interest_admin_authed'] = False
                st.rerun()

            st.markdown("---")
            st.markdown("#### 📤 워크북 업로드 (과거 실적 + 확정 스케줄 자동 반영)")
            st.caption(
                "시트 이름이 4자리 연도(예: '2026')면 그 해의 과거 실적으로, 이름에 '이자지급일정'이 "
                "들어있으면(예: 'PCBO 이자지급일정') 확정 스케줄로 저장합니다. 같은 연도/대출명을 다시 "
                "올리면 기존 데이터를 지우고 새로 채웁니다(중복 방지)."
            )
            interest_file = st.file_uploader(
                "이자 이력 워크북 (.xlsx)", type=["xlsx"], key="interest_history_uploader",
            )
            if st.button("💾 업로드 및 반영", type="primary", disabled=interest_file is None, key="interest_history_import_btn"):
                try:
                    wb_ih = openpyxl.load_workbook(io.BytesIO(interest_file.getvalue()), data_only=True)
                    result_ih = interest_history_db.import_from_workbook(wb_ih, source=interest_file.name)
                except Exception as e:
                    st.error(f"파싱 중 오류가 발생했습니다: {e}")
                else:
                    if not result_ih['years'] and not result_ih['schedules']:
                        st.warning(
                            "인식된 데이터가 없습니다 — 시트 이름이 4자리 연도이거나 '이자지급일정'을 "
                            "포함하는지, 표 안에 '이자계산시작' 헤더가 있는지 확인해주세요."
                        )
                    else:
                        year_summary = ', '.join(f"{y}년 {n}건" for y, n in sorted(result_ih['years'].items()))
                        sched_summary = ', '.join(f"{n} {c}건" for n, c in result_ih['schedules'].items())
                        msg = []
                        if year_summary:
                            msg.append(f"과거 실적 — {year_summary}")
                        if sched_summary:
                            msg.append(f"확정 스케줄 — {sched_summary}")
                        st.success(" / ".join(msg))
                        st.rerun()

            st.markdown("---")
            st.markdown("#### 📜 등록된 과거 실적(연도별)")
            ih_years = interest_history_db.list_actual_interest_years()
            if not ih_years:
                st.caption("등록된 과거 실적이 없습니다.")
            else:
                ih_year_summary_df = pd.DataFrame([
                    {
                        "연도": y,
                        "건수": len(interest_history_db.list_actual_interest(y)),
                        "합계(원)": sum(r[6] for r in interest_history_db.list_actual_interest(y)),
                    }
                    for y in ih_years
                ])
                ih_year_summary_df['합계(원)'] = ih_year_summary_df['합계(원)'].map(lambda v: f"{v:,.0f}")
                st.dataframe(ih_year_summary_df, use_container_width=True, hide_index=True)

                ih_del_year = st.selectbox("삭제할 연도 선택", ih_years, key="interest_history_delete_year")
                if st.button("🗑 선택한 연도 실적 삭제", key="interest_history_delete_btn"):
                    interest_history_db.replace_actual_interest_for_year(ih_del_year, [])
                    st.success(f"{ih_del_year}년 실적을 삭제했습니다.")
                    st.rerun()

            st.markdown("---")
            st.markdown("#### ✏️ 확정 스케줄 직접입력 (PCBO 등)")
            st.caption(
                "만기까지 금액ㆍ금리ㆍ이자가 미리 정해진 대출(예: 사채)을 직접 등록/수정합니다. "
                "'💵 원리금관리'에서 대출명이 차입금현황표의 '금융기관'명과 정확히 일치해야 반영됩니다."
            )
            ih_existing_loans = interest_history_db.list_fixed_schedule_loan_names()
            ih_loan_pick_mode = st.radio(
                "대상", ["기존 대출 선택", "신규 대출명 입력"], horizontal=True, key="interest_fixed_loan_mode",
                disabled=not ih_existing_loans,
            )
            if ih_loan_pick_mode == "기존 대출 선택" and ih_existing_loans:
                ih_loan_name = st.selectbox("대출명", ih_existing_loans, key="interest_fixed_loan_pick")
            else:
                ih_loan_name = st.text_input("대출명 (차입금현황표 '금융기관'명과 정확히 일치해야 함)", key="interest_fixed_loan_new")

            if ih_loan_name:
                ih_existing_rows = interest_history_db.list_fixed_schedule(ih_loan_name)
                ih_start_df = pd.DataFrame([
                    {"기간시작": r[4], "기간종료": r[5], "금액": r[2], "금리": r[3], "이자금액": r[6], "비고": r[7]}
                    for r in ih_existing_rows
                ]) if ih_existing_rows else pd.DataFrame(columns=["기간시작", "기간종료", "금액", "금리", "이자금액", "비고"])

                ih_edit_key = f'interest_fixed_editor_df__{ih_loan_name}'
                if ih_edit_key not in st.session_state:
                    st.session_state[ih_edit_key] = ih_start_df.copy()

                ih_edited_df = st.data_editor(
                    st.session_state[ih_edit_key],
                    num_rows="dynamic",
                    use_container_width=True,
                    key=f'interest_fixed_editor__{ih_loan_name}',
                    column_config={
                        "기간시작": st.column_config.TextColumn(help="YYYY-MM-DD"),
                        "기간종료": st.column_config.TextColumn(help="YYYY-MM-DD"),
                    },
                )

                if st.button("💾 이 대출명으로 저장", type="primary", key="interest_fixed_save_btn"):
                    clean_df = ih_edited_df.dropna(subset=["기간시작", "기간종료", "이자금액"], how='any')
                    parsed_rows = []
                    bad_rows = 0
                    for _, r in clean_df.iterrows():
                        start_d = interest_history_db._coerce_date(r["기간시작"])
                        end_d = interest_history_db._coerce_date(r["기간종료"])
                        if start_d is None or end_d is None:
                            bad_rows += 1
                            continue
                        parsed_rows.append({
                            'amount': pd.to_numeric(r.get("금액"), errors='coerce'),
                            'rate': pd.to_numeric(r.get("금리"), errors='coerce'),
                            'start': start_d, 'end': end_d,
                            'interest': pd.to_numeric(r["이자금액"], errors='coerce'),
                            'note': r.get("비고") if pd.notna(r.get("비고")) else None,
                        })
                    if bad_rows:
                        st.warning(f"기간시작/기간종료 형식(YYYY-MM-DD)을 읽지 못한 {bad_rows}개 행은 저장에서 제외했습니다.")
                    interest_history_db.replace_fixed_schedule_for_loan(ih_loan_name, parsed_rows, source='직접입력')
                    st.session_state[ih_edit_key] = clean_df
                    st.success(f"'{ih_loan_name}' 확정 스케줄 {len(parsed_rows)}건을 저장했습니다.")
                    st.rerun()

                if ih_existing_rows and st.button(f"🗑 '{ih_loan_name}' 전체 삭제", key="interest_fixed_delete_loan_btn"):
                    interest_history_db.delete_fixed_schedule_loan(ih_loan_name)
                    st.success(f"'{ih_loan_name}' 확정 스케줄을 모두 삭제했습니다.")
                    st.rerun()
        return

    # -----------------------------------------------------------------
    # 🏢 그룹 차입금 업로드(관리자) — 유진그룹에 매월 제출하는 '그룹차입금보고서.xlsx'의
    # 법인별 상세 시트(법인마다 이름은 다르지만 서식은 동일, 예: '8.동양')를 등록해두면,
    # '🏢 그룹 차입금 분석' 화면에서 기관별/과목별/총괄 분석표를 만듭니다. 시트는 위치
    # (예전엔 '맨 마지막 시트'로 고정)가 아니라 아래 '법인' 선택값과 이름이 맞는 시트를
    # 찾아서 쓴다 — 법인마다 파일 구조가 조금씩 달라 위치만으로는 엉뚱한 시트가
    # 골라질 수 있어서(실제로 이 문제로 데이터가 꼬였던 사례 확인) 이름 기준으로 바꿨다.
    # -----------------------------------------------------------------
    if page == PAGE_GROUP_LOAN_ADMIN:
        st.subheader("🏢 그룹 차입금 업로드(관리자)")
        st.write(
            "그룹차입금보고서 원본을 그대로(또는 해당 법인 시트만 남겨서) 업로드하면, "
            "아래 '법인'에서 고른 이름으로 끝나는 시트를 자동으로 찾아 씁니다(예: 법인=동양 → "
            "'8.동양' 시트, '9.동양에너지'는 이름이 달라 헷갈리지 않습니다). 이름이 안 맞으면 "
            "맨 마지막 시트로 대신 처리하고, 저장 후 어떤 시트를 썼는지 화면에 표시합니다."
        )

        try:
            correct_pw = st.secrets["admin"]["password"]
        except Exception:
            correct_pw = None

        if 'group_loan_admin_authed' not in st.session_state:
            st.session_state['group_loan_admin_authed'] = False

        if not correct_pw:
            st.error(
                "관리자 비밀번호가 설정되어 있지 않습니다. 프로그램 폴더의 .streamlit/secrets.toml 파일에서 "
                "[admin] password 값을 설정해주세요."
            )
        elif not st.session_state['group_loan_admin_authed']:
            pw_input = st.text_input("관리자 비밀번호", type="password", key="group_loan_admin_pw_input")
            if st.button("확인", type="primary", key="group_loan_admin_pw_confirm"):
                if pw_input == correct_pw:
                    st.session_state['group_loan_admin_authed'] = True
                    st.rerun()
                else:
                    st.error("비밀번호가 올바르지 않습니다.")
        else:
            st.success("✅ 관리자 인증됨")
            if st.button("로그아웃", key="group_loan_admin_logout"):
                st.session_state['group_loan_admin_authed'] = False
                st.rerun()

            st.markdown("---")
            st.markdown("#### ➕ 법인별 월간 데이터 등록 / 교체")
            current_month = datetime.now().month
            gl_col1, gl_col2, gl_col3 = st.columns(3)
            with gl_col1:
                gl_company = st.selectbox("법인", group_loan.GROUP_COMPANIES, key="gl_company")
            with gl_col2:
                gl_year = st.number_input(
                    "연도", min_value=2020, max_value=2100, value=datetime.now().year, step=1, key="gl_year"
                )
            with gl_col3:
                gl_month = st.selectbox("월", list(range(1, 13)), index=current_month - 1, key="gl_month")

            gl_existing = group_loan_db.get_report_bytes_by_company_month(gl_company, int(gl_year), int(gl_month))
            if gl_existing is not None:
                st.warning(f"⚠️ {gl_company} {gl_year}년 {gl_month}월 데이터가 이미 있습니다. 저장하면 덮어씁니다.")

            # 파일 업로더 key에 법인ㆍ연ㆍ월을 넣어, 법인을 바꿔도 이전에 골랐던 파일이
            # 그대로 남아있다가 다른 법인 이름으로 잘못 저장되는 일이 없게 한다 — key가
            # 고정이면 드롭다운만 바꾸고 파일을 다시 고르지 않아도 이전 파일이 계속
            # 선택된 상태로 남아, "저장"을 누르면 엉뚱한 법인에 그 파일이 들어갔다
            # (동양 것을 올린 뒤 법인만 금왕에프원으로 바꾸고 저장하면 금왕에프원 자리에
            # 동양 파일이 들어가는 사례로 확인됨). 법인ㆍ연ㆍ월이 하나라도 바뀌면 완전히
            # 새 위젯으로 취급돼 파일 선택이 항상 비워진 채로 시작한다.
            gl_uploader_key = f"gl_uploader_{gl_company}_{gl_year}_{gl_month}"
            gl_file = st.file_uploader(
                "그룹차입금보고서 엑셀 (.xlsx) — 해당 법인 시트 포함",
                type=["xlsx"], key=gl_uploader_key,
                help="예: 21_(주)동양 6월_그룹차입금보고서.xlsx",
            )

            # 업로드된 파일의 시트 이름만 보고 어느 법인 파일인지 미리 알아내서, 위 '법인'
            # 드롭다운 선택과 맞는지 저장 전에 대조한다 — 드롭다운을 안 바꾸고 다른 회사
            # 파일을 올려서 엉뚱한 법인 이름으로 저장되는 사고(실제 발생 사례)를 막기 위함.
            gl_detected_company = None
            if gl_file is not None:
                try:
                    gl_probe_wb = openpyxl.load_workbook(io.BytesIO(gl_file.getvalue()), data_only=True, read_only=True)
                    gl_detected_company = group_loan.detect_company_from_workbook(gl_probe_wb)
                except Exception:
                    gl_detected_company = None

            gl_mismatch = gl_detected_company is not None and gl_detected_company != gl_company
            if gl_mismatch:
                st.error(
                    f"🚫 이 파일은 시트 이름으로 볼 때 **'{gl_detected_company}'** 법인 것으로 보이는데, "
                    f"위에서 선택한 법인은 **'{gl_company}'**입니다. 위 '법인'을 '{gl_detected_company}'로 "
                    f"바꾸거나(맞다면), 올바른 법인의 파일을 다시 선택해주세요 — 서로 다른 동안에는 저장을 막아둡니다."
                )
            elif gl_file is not None and gl_detected_company is not None:
                st.caption(f"✅ 파일 확인: 시트 이름이 '{gl_detected_company}' 법인과 일치합니다.")
            elif gl_file is not None and gl_detected_company is None:
                st.caption("ℹ️ 파일에서 법인을 자동으로 알아내지 못했습니다 — 위 '법인' 선택이 맞는지 직접 확인해주세요.")

            if st.button("💾 저장", type="primary", disabled=(gl_file is None or gl_mismatch), key="gl_save"):
                try:
                    parsed_preview, matched_sheet, unit_fix_count, sheet_format = group_loan.load_company_items_from_bytes(
                        gl_file.getvalue(), company=gl_company
                    )
                except Exception as e:
                    st.error(f"파싱 실패: {e} — 시트 서식을 확인해주세요.")
                    parsed_preview, matched_sheet, unit_fix_count, sheet_format = None, None, 0, None
                if parsed_preview is not None:
                    if not parsed_preview:
                        st.error(f"'{matched_sheet}' 시트에서 차입금 항목을 하나도 찾지 못했습니다 — {gl_company} 법인의 시트가 맞는지 확인해주세요.")
                    else:
                        group_loan_db.save_report(gl_company, int(gl_year), int(gl_month), gl_file.name, gl_file.getvalue())
                        format_label = "개별법인 제출용 서식" if sheet_format == 'individual' else "그룹차입금보고서 형식"
                        st.success(
                            f"{gl_company} {gl_year}년 {gl_month}월 데이터를 저장했습니다 "
                            f"(시트 '{matched_sheet}' · {format_label}으로 인식 · {len(parsed_preview)}개 항목)."
                        )
                        if unit_fix_count:
                            st.warning(
                                f"⚠️ {unit_fix_count}개 항목에서 단위가 억원이 아니라 원(WON)으로 입력된 것으로 "
                                f"보여 자동으로 억원으로 환산했습니다 — 원본 엑셀에서 해당 줄의 단위를 한 번 확인해주세요."
                            )
                        st.rerun()

            st.markdown("---")
            # 위 업로드 폼에서 고른 연도/월을 그대로 이어받아 보여준다(별도 선택창을 두면
            # 방금 올린 달과 다른 달을 보고 있으면서 "안 올라갔다"고 착각하기 쉬워서,
            # 아예 같은 연도/월 기준으로 고정했다).
            st.markdown(f"#### 📋 {gl_year}년 {gl_month}월 등록 현황")
            st.caption("바로 위에서 고른 연도/월 기준입니다 — 다른 달을 보려면 위쪽 '연도'/'월'을 바꿔주세요.")
            gl_rows = group_loan_db.list_reports_for_month(int(gl_year), int(gl_month))
            if not gl_rows:
                st.caption("이 달에 등록된 법인이 아직 없습니다.")
            else:
                gl_list_df = pd.DataFrame([
                    {"법인": r[1], "파일명": r[2], "등록일시": r[3], "크기(KB)": round(r[4] / 1024, 1)}
                    for r in gl_rows
                ])
                st.dataframe(gl_list_df, use_container_width=True, hide_index=True)
                missing = [c for c in group_loan.GROUP_COMPANIES if c not in {r[1] for r in gl_rows}]
                if missing:
                    st.caption(f"⚠️ 이 달에 아직 등록되지 않은 법인: {', '.join(missing)}")

                gl_del_options = {f"{r[1]} · {r[2]} ({r[3]})": r[0] for r in gl_rows}
                gl_del_label = st.selectbox("삭제할 항목 선택", list(gl_del_options.keys()), key="gl_delete_pick")
                if st.button("🗑 선택한 데이터 삭제", key="gl_delete_btn"):
                    group_loan_db.delete_report(gl_del_options[gl_del_label])
                    st.success("삭제했습니다.")
                    st.rerun()

            st.markdown("---")
            st.markdown("#### 🗓 전체 등록 현황(모든 월 한눈에 보기)")
            gl_all_months = group_loan_db.list_months()
            if not gl_all_months:
                st.caption("등록된 데이터가 전혀 없습니다.")
            else:
                overview_rows = []
                for oy, om in gl_all_months:
                    orows = group_loan_db.list_reports_for_month(oy, om)
                    done = sorted({r[1] for r in orows})
                    overview_rows.append({
                        "연도": oy, "월": om,
                        "등록 법인 수": f"{len(done)}/{len(group_loan.GROUP_COMPANIES)}",
                        "등록된 법인": ', '.join(done) if done else '-',
                    })
                st.dataframe(pd.DataFrame(overview_rows), use_container_width=True, hide_index=True)
                st.caption("이 표에 없는 연도/월은 어떤 법인도 저장된 적이 없는 것입니다.")
        return

    # -----------------------------------------------------------------
    # 📑 월간보고서(표지/관리/대여/운용/차입/기타/전체 출력) — 업로드된 월간보고서 원본 엑셀의
    # 시트를 그대로 재현해 보여줍니다. 월별 차입금 스냅샷과는 별개의 데이터 소스라
    # 아래 '조회 단위 선택' 게이트보다 먼저 독립적으로 처리합니다.
    # -----------------------------------------------------------------
    REPORT_VIEW_PAGES = (
        PAGE_REPORT_ALL, PAGE_REPORT_COVER, PAGE_REPORT_MANAGE, PAGE_REPORT_LEND,
        PAGE_REPORT_INVEST, PAGE_REPORT_BORROW, PAGE_REPORT_OTHER,
    )
    if page in REPORT_VIEW_PAGES:
        st.subheader(page)

        report_rows = monthly_report_db.list_reports()
        if not report_rows:
            st.info(
                "아직 등록된 월간보고서가 없습니다. 상단 메뉴의 **⚙️ 데이터 관리 → 📑 월간보고서 업로드(관리자)** "
                "에서 월간보고서 엑셀 원본을 먼저 등록해주세요."
            )
            return

        report_options = {r[3]: r[0] for r in report_rows}
        selected_report_label = st.selectbox("🗓 조회 월 선택", list(report_options.keys()), key="report_view_select")
        report_bytes = monthly_report_db.get_report_bytes(report_options[selected_report_label])

        try:
            report_wb = openpyxl.load_workbook(io.BytesIO(report_bytes), data_only=True)
        except Exception as e:
            st.error(f"월간보고서 파일을 열 수 없습니다: {e}")
            return

        SHEET_MAP = {
            PAGE_REPORT_COVER: ['표지'],
            PAGE_REPORT_MANAGE: [
                '자금현황(통합 누적-보고용)', '자금현황(통합 누적-보고용) (2)',
                '사업본부 통합', '건재', '건설', '플랜트', '공통',
            ],
            PAGE_REPORT_LEND: ['대여금 현황'],
            PAGE_REPORT_INVEST: ['자금운용현황'],
            PAGE_REPORT_BORROW: ['차입금 및 담보현황'],
            PAGE_REPORT_OTHER: ['어음현황', '어음현황 (2)'],
        }

        if page == PAGE_REPORT_ALL:
            st.markdown(
                '<button class="no-print" onclick="window.print()" '
                'style="background:#3182f6;color:white;border:none;padding:10px 20px;border-radius:6px;'
                'font-weight:700;cursor:pointer;margin-bottom:16px;font-size:0.95rem;">🖨 인쇄하기 / PDF 저장</button>',
                unsafe_allow_html=True
            )
            sheet_sequence = []
            for group_page in [
                PAGE_REPORT_COVER, PAGE_REPORT_MANAGE, PAGE_REPORT_LEND,
                PAGE_REPORT_INVEST, PAGE_REPORT_BORROW, PAGE_REPORT_OTHER,
            ]:
                sheet_sequence.extend(SHEET_MAP[group_page])
        else:
            sheet_sequence = SHEET_MAP[page]

        html_parts = []
        for sheet_name in sheet_sequence:
            if sheet_name not in report_wb.sheetnames:
                continue
            ws = report_wb[sheet_name]
            html_parts.append('<div class="report-page">')
            html_parts.append(f'<div class="report-page-title">{sheet_name}</div>')
            html_parts.append(render_sheet_html(ws))
            html_parts.append('</div>')

        if not html_parts:
            st.warning("이 화면에 해당하는 시트를 파일에서 찾지 못했습니다.")
        else:
            st.markdown(''.join(html_parts), unsafe_allow_html=True)
        return

    # -----------------------------------------------------------------
    # 🏢 그룹 차입금 분석 — 7개 법인의 그룹차입금보고서(법인별 마지막 시트)를 모아
    # 3.금융기관별(잔액)/4.금융기관별(금리)/5.과목별(잔액)/1.그룹총괄표와 같은 구조로
    # 재현합니다. 데이터는 group_loan_db(법인·연·월별 업로드)에서 가져옵니다.
    # -----------------------------------------------------------------
    if page == PAGE_GROUP_LOAN:
        st.subheader(page)

        gl_months = group_loan_db.list_months()
        if not gl_months:
            st.info(
                "아직 등록된 그룹 차입금 데이터가 없습니다. 상단 메뉴의 "
                "**⚙️ 데이터 관리 → 🏢 그룹 차입금 업로드(관리자)** 에서 법인별로 먼저 등록해주세요."
            )
            return

        # 연도 → 월 2단계 선택 — 데이터가 여러 해에 걸쳐 쌓여도 월 목록이 한없이
        # 길어지지 않도록, 먼저 연도를 고르고 그 해에 등록된 월만 보여준다.
        gl_years = sorted({y for y, _m in gl_months}, reverse=True)
        gl_view_col1, gl_view_col2 = st.columns(2)
        with gl_view_col1:
            view_year = st.selectbox("연도", gl_years, key="gl_view_year")
        with gl_view_col2:
            months_in_year = sorted([m for y, m in gl_months if y == view_year], reverse=True)
            # 키에 연도를 포함시켜, 연도를 바꿨을 때 이전 연도의 월 값이 새 옵션 목록에
            # 없어서 에러가 나는 대신 그 연도의 최신월로 자연스럽게 초기화되게 한다.
            view_month = st.selectbox(
                "월", months_in_year, format_func=lambda m: f"{m}월", key=f"gl_view_month_{view_year}"
            )

        def _load_month_items(y, m):
            rows = group_loan_db.list_reports_for_month(y, m)
            result = {}
            for r in rows:
                company = r[1]
                fb = group_loan_db.get_report_bytes_by_company_month(company, y, m)
                try:
                    items, _matched_sheet, _unit_fix_count, _sheet_format = group_loan.load_company_items_from_bytes(fb, company=company)
                    result[company] = items
                except Exception as e:
                    st.warning(f"{company} 시트 파싱 실패: {e}")
            return result

        company_items = _load_month_items(view_year, view_month)
        registered = sorted(company_items.keys())
        missing = [c for c in group_loan.GROUP_COMPANIES if c not in company_items]
        st.caption(
            f"✅ 조회 중: {view_year}년 {view_month}월 · 등록 법인 {len(registered)}/{len(group_loan.GROUP_COMPANIES)}"
            + (f" · 미등록: {', '.join(missing)}" if missing else "")
        )

        # 전월(직전으로 등록된 월) 데이터 — 있으면 총괄표의 전월대비에 사용
        month_idx = gl_months.index((view_year, view_month))
        prev_items = None
        if month_idx + 1 < len(gl_months):
            prev_y, prev_m = gl_months[month_idx + 1]
            prev_items = _load_month_items(prev_y, prev_m)

        def _fmt_amt(v):
            return f"{v:,.2f}" if isinstance(v, (int, float)) else "-"

        def _fmt_pct(v):
            return f"{v * 100:.2f}%" if isinstance(v, (int, float)) else "-"

        st.markdown("#### 1. 총괄표")
        summary_rows, summary_total = group_loan.build_summary(company_items, prev_items)
        summary_df = pd.DataFrame([
            {
                "법인": c, "당월 총한도(억원)": _fmt_amt(summary_rows[c]['총한도']),
                "당월 잔액(억원)": _fmt_amt(summary_rows[c]['잔액']),
                "가중평균금리": _fmt_pct(summary_rows[c]['가중평균금리']),
                "전월 잔액(억원)": _fmt_amt(summary_rows[c]['전월잔액']),
                "전월대비(억원)": _fmt_amt(summary_rows[c]['전월대비']),
            }
            for c in group_loan.GROUP_COMPANIES if c in summary_rows
        ] + [{
            "법인": "계", "당월 총한도(억원)": _fmt_amt(summary_total['총한도']),
            "당월 잔액(억원)": _fmt_amt(summary_total['잔액']),
            "가중평균금리": _fmt_pct(summary_total['가중평균금리']),
            "전월 잔액(억원)": _fmt_amt(summary_total['전월잔액']),
            "전월대비(억원)": _fmt_amt(summary_total['전월대비']),
        }])
        st.dataframe(summary_df, use_container_width=True, hide_index=True)
        if prev_items is None:
            st.caption("ℹ️ 직전 등록월이 없어 전월 잔액ㆍ전월대비는 비워뒀습니다.")

        st.markdown("---")
        st.markdown("#### 2. 금융기관별 잔액 (억원)")
        bal_rows, bal_total = group_loan.build_balance_by_institution(company_items)
        bal_cols = group_loan.INSTITUTION_COLUMNS + ['합계']
        bal_df = pd.DataFrame([
            {"법인": c, **{col: _fmt_amt(bal_rows[c][col]) for col in bal_cols}}
            for c in group_loan.GROUP_COMPANIES if c in bal_rows
        ] + [{"법인": "계", **{col: _fmt_amt(bal_total[col]) for col in bal_cols}}])
        st.dataframe(bal_df, use_container_width=True, hide_index=True)

        st.markdown("---")
        st.markdown("#### 3. 금융기관별 금리")
        rate_rows, rate_total = group_loan.build_rate_by_institution(company_items)
        rate_cols = group_loan.INSTITUTION_COLUMNS + ['가중평균']
        rate_df = pd.DataFrame([
            {"법인": c, **{col: _fmt_pct(rate_rows[c][col]) for col in rate_cols}}
            for c in group_loan.GROUP_COMPANIES if c in rate_rows
        ] + [{"법인": "계", **{col: _fmt_pct(rate_total[col]) for col in rate_cols}}])
        st.dataframe(rate_df, use_container_width=True, hide_index=True)

        st.markdown("---")
        st.markdown("#### 4. 과목별 잔액 (억원)")
        subj_rows, subj_total = group_loan.build_balance_by_subject(company_items)
        subj_cols = group_loan.SUBJECT_COLUMNS + ['합계']
        subj_df = pd.DataFrame([
            {"법인": c, **{col: _fmt_amt(subj_rows[c][col]) for col in subj_cols}}
            for c in group_loan.GROUP_COMPANIES if c in subj_rows
        ] + [{"법인": "계", **{col: _fmt_amt(subj_total[col]) for col in subj_cols}}])
        st.dataframe(subj_df, use_container_width=True, hide_index=True)

        # -------------------------------------------------------------
        # 📊 도식화 — 표만으로는 한눈에 안 들어와서, 기관별/법인별/과목별 비중을
        # 원형ㆍ막대 그래프로도 보여준다(기존 '차입금현황표' 페이지의 원형도표와
        # 같은 스타일을 그대로 따랐다).
        # -------------------------------------------------------------
        st.markdown("---")
        st.markdown("### 📊 도식화")
        CHART_COLORS = [
            '#3b82f6', '#10b981', '#06b6d4', '#f59e0b', '#c084fc',
            '#f472b6', '#ec4899', '#6366f1', '#14b8a6', '#64748b', '#84cc16', '#f97316', '#a855f7',
        ]

        def _donut(labels_values, center_label):
            fig = go.Figure(data=[go.Pie(
                labels=[l for l, v in labels_values], values=[v for l, v in labels_values],
                hole=0.55,
                marker=dict(
                    colors=(CHART_COLORS * (len(labels_values) // len(CHART_COLORS) + 1))[:len(labels_values)],
                    line=dict(color='#0f172a', width=2),
                ),
                textinfo='percent', textfont=dict(color='white', size=12),
                hovertemplate='%{label}<br>%{value:,.1f}억원<br>%{percent}<extra></extra>',
            )])
            total = sum(v for _l, v in labels_values)
            fig.update_layout(
                showlegend=True,
                legend=dict(orientation="v", yanchor="middle", y=0.5, xanchor="left", x=1.02, font=dict(size=11, color='#64748b')),
                paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)',
                font=dict(color='#0f172a'), margin=dict(t=10, b=10, l=10, r=10), height=340,
                annotations=[dict(text=center_label(total), x=0.5, y=0.5, font=dict(size=15, color='#0f172a'), showarrow=False)],
            )
            return fig

        @st.dialog("금융기관 세부내역")
        def _show_institution_dialog(inst):
            st.markdown(f"**{inst}** — 법인별 세부내역")
            rows = []
            for c in group_loan.GROUP_COMPANIES:
                for it in company_items.get(c, []):
                    if group_loan.normalize_institution_column(it['institution_raw']) == inst:
                        rows.append({
                            "법인": c, "차입처": it['institution_raw'],
                            "잔액(억원)": _fmt_amt(it['balance']),
                            "금리": _fmt_pct(it['rate']) if it['rate'] is not None else "-",
                        })
            if rows:
                st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)
            else:
                st.caption("세부내역이 없습니다.")

        chart_col1, chart_col2 = st.columns(2)
        with chart_col1:
            st.markdown("#### 🥧 금융기관별 잔액 비중 (7개 법인 합산)")
            inst_lv = [(col, bal_total[col]) for col in group_loan.INSTITUTION_COLUMNS if bal_total[col] > 0]
            if inst_lv:
                st.plotly_chart(_donut(inst_lv, lambda t: f"{t:,.0f}억원"), use_container_width=True)
                # 원형(파이/도넛) 차트는 Plotly의 클릭선택(on_select)이 애초에 지원되지
                # 않는다 — 선택 기능은 x/y축이 있는 카테시안 트레이스(막대ㆍ산점도 등)
                # 전용이라 각도 기반인 파이 트레이스에는 적용되지 않는다(라이브러리
                # 자체 한계, 실제로 클릭해도 아무 반응이 없는 것으로 확인됨). 그래서
                # 여기만 기관 이름 버튼을 눌러 세부내역 팝업을 여는 방식을 쓴다.
                st.caption("👇 기관을 클릭하면 세부내역이 뜹니다")
                inst_labels = [l for l, _v in inst_lv]
                for i in range(0, len(inst_labels), 4):
                    row_cols = st.columns(4)
                    for rc, label in zip(row_cols, inst_labels[i:i + 4]):
                        with rc:
                            if st.button(label, key=f"gl_inst_btn_{label}", use_container_width=True):
                                _show_institution_dialog(label)
            else:
                st.caption("표시할 잔액이 없습니다.")
        with chart_col2:
            st.markdown("#### 🥧 과목별 잔액 비중 (7개 법인 합산)")
            subj_lv = [(col, subj_total[col]) for col in ['일반대', '한도대(잔액)', '회사채', '전단채', 'CP', '기타'] if subj_total[col] > 0]
            if subj_lv:
                st.plotly_chart(_donut(subj_lv, lambda t: f"{t:,.0f}억원"), use_container_width=True)
            else:
                st.caption("표시할 잔액이 없습니다.")

        @st.dialog("법인 세부내역")
        def _show_company_dialog(company):
            st.markdown(f"**{company}** — 금융기관별 세부내역")
            rows = [
                {
                    "구분": it['subject'] or '-', "차입처": it['institution_raw'],
                    "한도(억원)": _fmt_amt(it['limit']), "잔액(억원)": _fmt_amt(it['balance']),
                    "금리": _fmt_pct(it['rate']) if it['rate'] is not None else "-",
                }
                for it in company_items.get(company, [])
            ]
            if rows:
                st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)
            else:
                st.caption("세부내역이 없습니다.")

        st.markdown("#### 📊 법인별 잔액 비교 (억원) — 막대를 클릭하면 세부내역")
        bar_companies = [c for c in group_loan.GROUP_COMPANIES if c in summary_rows]
        if bar_companies:
            bar_values = [summary_rows[c]['잔액'] for c in bar_companies]
            bar_max = max(bar_values) if bar_values else 0
            fig_bar = go.Figure(data=[go.Bar(
                x=bar_companies, y=bar_values,
                # width를 카테고리 슬롯 대비 고정 비율로 못박아서, bargap만으로는
                # 법인 수가 적을 때(예: 2개) 슬롯 자체가 넓어져 막대가 여전히
                # 두꺼워 보이던 문제를 없앤다 — 법인이 2개든 7개든 막대 하나의
                # 두께가 항상 일정하게 얇게 유지된다.
                width=[0.28] * len(bar_companies),
                marker=dict(color=(CHART_COLORS * (len(bar_companies) // len(CHART_COLORS) + 1))[:len(bar_companies)]),
                text=[f"{v:,.0f}" for v in bar_values], textposition='outside',
                hovertemplate='%{x}<br>%{y:,.1f}억원<extra></extra>',
            )])
            fig_bar.update_layout(
                paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)',
                font=dict(color='#0f172a'), margin=dict(t=40, b=10, l=10, r=10), height=320,
                xaxis=dict(title=None),
                # 막대 위 숫자 라벨(textposition='outside')이 y축 최댓값 바로
                # 위까지 차오르면 잘려 보였던 문제 — y축 상단에 18% 여유를 둬서
                # 라벨이 항상 그래프 영역 안에 다 보이게 한다.
                yaxis=dict(title='잔액(억원)', gridcolor='rgba(15,23,42,0.06)', range=[0, bar_max * 1.18 if bar_max else 1]),
                showlegend=False,
                bargap=0.55,
            )
            bar_event = st.plotly_chart(
                fig_bar, use_container_width=True,
                on_select="rerun", selection_mode="points", key="gl_company_bar",
            )
            bar_points = (bar_event or {}).get("selection", {}).get("points", [])
            if bar_points and bar_points[0].get("x"):
                _show_company_dialog(bar_points[0]["x"])
        return

    # -----------------------------------------------------------------
    # 1) 데이터 소스: 매번 파일을 업로드하는 대신, 관리자가 월별로 미리 등록해둔
    #    스냅샷(이 PC의 로컬 DB에 저장됨)을 월/분기/연 단위로 골라 조회합니다.
    #    ('⚙️ 데이터 관리 → 🗄 데이터 관리(관리자)' 메뉴에서 등록/교체/삭제합니다.)
    # -----------------------------------------------------------------
    st.sidebar.subheader("📅 조회 단위 선택")
    snapshots = snapshot_db.list_snapshots()

    if not snapshots:
        st.sidebar.warning("⚠️ 등록된 데이터가 없습니다.")
        st.info(
            "아직 등록된 차입금 데이터가 없습니다. 상단 메뉴의 **⚙️ 데이터 관리 → 🗄 데이터 관리(관리자)** 에서 "
            "차입금 관리내역 엑셀 파일을 업로드해 월별 데이터를 먼저 등록해주세요."
        )
        return

    QUARTER_END_MONTHS = {3: 1, 6: 2, 9: 3, 12: 4}

    view_granularity = st.sidebar.radio(
        "보기 단위", ["월별", "분기별", "연도별"], horizontal=True, key="view_granularity"
    )

    if view_granularity == "분기별":
        filtered_snapshots = [s for s in snapshots if s[2] in QUARTER_END_MONTHS]
    elif view_granularity == "연도별":
        filtered_snapshots = [s for s in snapshots if s[2] == 12]
    else:
        filtered_snapshots = snapshots

    if not filtered_snapshots:
        st.sidebar.warning(f"⚠️ {view_granularity} 기준으로 볼 데이터가 없습니다.")
        st.info(
            f"'{view_granularity}'로 조회하려면 분기말(3·6·9·12월) 또는 연말(12월) 데이터가 등록되어 있어야 "
            "합니다. 월별 보기로 바꾸거나, 데이터 관리에서 해당 월 데이터를 먼저 등록해주세요."
        )
        return

    def _snapshot_display_label(s):
        year, month = s[1], s[2]
        if view_granularity == "분기별":
            return f"{year}년 {QUARTER_END_MONTHS[month]}분기 ({month}월 기준)"
        if view_granularity == "연도별":
            return f"{year}년 (12월 기준)"
        return f"{year}년 {month}월"

    snapshot_options = {
        _snapshot_display_label(s): s[0] for s in filtered_snapshots
    }
    options_list = list(snapshot_options.keys())
    state_key = f'selected_snapshot_label__{view_granularity}'

    current_label = st.session_state.get(state_key)
    if current_label not in snapshot_options:
        current_label = options_list[0]

    st.sidebar.selectbox(
        "조회할 시점", options_list, index=options_list.index(current_label),
        key=state_key, label_visibility="collapsed",
    )
    current_label = st.session_state[state_key]

    # 월/분기/연 전환 시에도(선택박스를 직접 건드리지 않아도) 항상 현재 선택된
    # 시점의 데이터로 동기화합니다.
    if st.session_state.get('excel_filename') != current_label:
        st.session_state['excel_bytes'] = snapshot_db.get_snapshot_bytes(snapshot_options[current_label])
        st.session_state['excel_filename'] = current_label

    st.sidebar.caption(f"✅ 조회 중: {st.session_state['excel_filename']}")

    excel_bytes = st.session_state['excel_bytes']

    # Load data from uploaded Excel

    excel_kpi, df_hist = load_excel_data(excel_bytes)

    # 인쇄용 요약(PAGE_PRINT) 등에서 참조하는 이름들의 기본값.
    # df_hist가 비어있는 경우에도 NameError 없이 "데이터 부족" 메시지를 보여주기 위함.
    eval_df = pd.DataFrame()
    y_a = y_b = y_c = None

    # 차입금 관리내역(현재) 시트의 '합계' 행 금리를 가져와, 당해년도 가중평균 금리로 자동 반영합니다.
    # (수동 입력값 대신 엑셀에 이미 계산되어 있는 값을 우선 사용)
    computed_curr_rate = None
    try:
        df_current_for_kpi, _ = load_current_loan_status(excel_bytes)
        if not df_current_for_kpi.empty and '금융기관' in df_current_for_kpi.columns:
            is_total_for_kpi = df_current_for_kpi['금융기관'].astype(str).str.strip().isin(['계', '합계'])
            rate_col_for_kpi = next((c for c in df_current_for_kpi.columns if c.strip() == '금리'), None)
            if rate_col_for_kpi is not None:
                total_rate_vals = pd.to_numeric(
                    df_current_for_kpi.loc[is_total_for_kpi, rate_col_for_kpi], errors='coerce'
                ).dropna()
                if not total_rate_vals.empty:
                    computed_curr_rate = float(total_rate_vals.iloc[0]) * 100
    except Exception:
        computed_curr_rate = None

    # Calculate KPIs from st.session_state inputs
    inputs = st.session_state['kpi_inputs']
    effective_avg_rate_curr = computed_curr_rate if computed_curr_rate is not None else inputs['avg_rate_curr']
    rate_diff = effective_avg_rate_curr - inputs['avg_rate_prev']
    saving_rate = (inputs['avg_rate_prev'] - effective_avg_rate_curr) - (inputs['bok_rate_prev'] - inputs['bok_rate_curr'])
    saving_amount = inputs['balance_curr'] * 1000000 * (saving_rate / 100)
    total_balance = inputs['balance_curr'] * 1000000

    kpi_calculated = {
        "saving_amount": saving_amount,
        "avg_rate_prev": inputs['avg_rate_prev'],
        "avg_rate_curr": effective_avg_rate_curr,
        "saving_rate": saving_rate,
        "total_balance": total_balance,
        "rate_is_auto": computed_curr_rate is not None,
    }

    # 요약 수치는 항상 계산해 두고(다른 화면에서도 참조), 카드 렌더링은 KPI 화면에서만 표시
    saving_amt = kpi_calculated['saving_amount']
    saving_str = f"{saving_amt/100000000:,.1f}억원" if abs(saving_amt) >= 100000000 else f"{saving_amt:,.0f}원"
    total_bal = kpi_calculated['total_balance']
    bal_str = f"{total_bal/100000000:,.1f}억원" if abs(total_bal) >= 100000000 else f"{total_bal:,.0f}원"

    diff_class = "positive" if rate_diff < 0 else "negative" if rate_diff > 0 else "neutral"
    diff_text = f"전년비 {abs(rate_diff):.2f}%p 인하" if rate_diff < 0 else f"전년비 {abs(rate_diff):.2f}%p 상승" if rate_diff > 0 else "동결"
    if kpi_calculated["rate_is_auto"]:
        diff_text += " · 차입금표 합계 자동반영"

    if page == PAGE_KPI:
        st.markdown(f"### 🏆 핵심 성과 지표 (KPI) - {CURRENT_YEAR}년 기준")
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            st.markdown(f'<div class="metric-card"><div class="metric-label">연간 이자비용 절감액</div><div class="metric-value">{saving_str}</div><div class="metric-desc positive">기준금리 대비 절감효과 반영</div></div>', unsafe_allow_html=True)
        with col2:
            st.markdown(f'<div class="metric-card"><div class="metric-label">전년대비 가중평균 금리</div><div class="metric-value">{kpi_calculated["avg_rate_curr"]:.2f}%</div><div class="metric-desc {diff_class}">{diff_text}</div></div>', unsafe_allow_html=True)
        with col3:
            st.markdown(f'<div class="metric-card"><div class="metric-label">절감 금리 (효과)</div><div class="metric-value">{abs(kpi_calculated["saving_rate"]):.2f}%p</div><div class="metric-desc">기준금리 변동 반영 성과</div></div>', unsafe_allow_html=True)
        with col4:
            st.markdown(f'<div class="metric-card"><div class="metric-label">총 차입잔액</div><div class="metric-value">{bal_str}</div><div class="metric-desc">기준 시점 기준</div></div>', unsafe_allow_html=True)
        st.markdown("---")
        st.caption("상단 메뉴에서 다른 화면(금리 추이·신용등급·차입금현황·만기도래·부수거래 연동·인쇄용 요약)으로 이동할 수 있습니다.")

        with st.expander("✏️ 수동 입력 및 관리 (KPI 성과표)"):
            st.write(f"대시보드의 전년({PREV_YEAR}년) 및 당년({CURRENT_YEAR}년) 차입금 성과 분석에 사용될 데이터 기준을 입력합니다.")

            with st.form("kpi_input_form_tab5"):
                col_y1, col_y2 = st.columns(2)

                with col_y1:
                    st.markdown(f"##### 📅 {PREV_YEAR}년말 (전년도 기준)")
                    val_bal_prev_t5 = st.number_input(
                        "차입잔액 (백만원)",
                        value=st.session_state['kpi_inputs']['balance_prev'],
                        step=100.0,
                        format="%.1f",
                        key="t5_bal_prev",
                        help=f"전년도({PREV_YEAR}년) 차입잔액"
                    )
                    st.caption(f"💡 약 {val_bal_prev_t5/100:,.1f} 억원")
                    val_rate_prev_t5 = st.number_input(
                        "가중평균 차입금리 (%)",
                        value=st.session_state['kpi_inputs']['avg_rate_prev'],
                        min_value=0.0,
                        max_value=20.0,
                        step=0.01,
                        format="%.2f",
                        key="t5_rate_prev"
                    )
                    val_bok_prev_t5 = st.number_input(
                        "기준금리 평균 (%)",
                        value=st.session_state['kpi_inputs']['bok_rate_prev'],
                        min_value=0.0,
                        max_value=20.0,
                        step=0.01,
                        format="%.2f",
                        key="t5_bok_prev"
                    )

                with col_y2:
                    st.markdown(f"##### 📅 {CURRENT_YEAR}년말 (당해년도 기준)")
                    val_bal_curr_t5 = st.number_input(
                        "차입잔액 (백만원)",
                        value=st.session_state['kpi_inputs']['balance_curr'],
                        step=100.0,
                        format="%.1f",
                        key="t5_bal_curr",
                        help=f"당해년도({CURRENT_YEAR}년) 차입잔액"
                    )
                    st.caption(f"💡 약 {val_bal_curr_t5/100:,.1f} 억원")
                    val_rate_curr_t5 = st.number_input(
                        "가중평균 차입금리 (%)",
                        value=st.session_state['kpi_inputs']['avg_rate_curr'],
                        min_value=0.0,
                        max_value=20.0,
                        step=0.01,
                        format="%.2f",
                        key="t5_rate_curr"
                    )
                    val_bok_curr_t5 = st.number_input(
                        "기준금리 평균 (%)",
                        value=st.session_state['kpi_inputs']['bok_rate_curr'],
                        min_value=0.0,
                        max_value=20.0,
                        step=0.01,
                        format="%.2f",
                        key="t5_bok_curr"
                    )

                submit_btn_t5 = st.form_submit_button("입력값 저장 및 대시보드 반영")
                if submit_btn_t5:
                    st.session_state['kpi_inputs'] = {
                        'balance_prev': val_bal_prev_t5,
                        'avg_rate_prev': val_rate_prev_t5,
                        'bok_rate_prev': val_bok_prev_t5,
                        'balance_curr': val_bal_curr_t5,
                        'avg_rate_curr': val_rate_curr_t5,
                        'bok_rate_curr': val_bok_curr_t5
                    }
                    try:
                        save_kpi_values(st.session_state['kpi_inputs'])
                        st.success("성과표 입력값이 성공적으로 저장되었습니다! 대시보드를 업데이트합니다.")
                        st.rerun()
                    except Exception as e:
                        st.error(f"저장 중 오류 발생: {e}")

    if df_hist is not None and not df_hist.empty:
        # 연도 라벨: 실제 당해년도(CURRENT_YEAR)에 해당하는 컬럼에만 "(현재)" 표시
        years = []
        for yc in YEAR_COLS:
            yr_num = int(yc.replace('년', ''))
            years.append(f"{yr_num}년(현재)" if yr_num == CURRENT_YEAR else yc)

        avg_bank = []
        for y in YEAR_COLS:
            if y in df_hist.columns:
                valid_rates = df_hist[df_hist[y] > 0][y]
                avg_bank.append(valid_rates.mean() if not valid_rates.empty else None)
            else:
                avg_bank.append(None)

        bok_data = [bok_rates.get(str(int(y.replace('년', ''))), None) for y in YEAR_COLS]
        # 값이 없으면 합리적 기본값으로 대체 (그래프가 완전히 비지 않도록)
        bok_data = [v if v is not None else 3.0 for v in bok_data]

        if page == PAGE_RATE_TREND:
            st.markdown("### 📊 기준금리 vs 가중평균금리 추이 (종합)")
            st.write("한국은행 기준금리 대비 시중은행들의 연도별 가중평균 금리 추이입니다.")

            rate_change = fetch_base_rate_change_this_year()
            if rate_change:
                chg_date_str, old_rate, new_rate = rate_change
                chg_date_fmt = f"{chg_date_str[:4]}-{chg_date_str[4:6]}-{chg_date_str[6:]}"
                bp = round((new_rate - old_rate) * 100)
                direction = "인상" if bp > 0 else "인하"
                st.info(
                    f"📌 한국은행 기준금리가 **{chg_date_fmt}부터 {old_rate:.2f}% → {new_rate:.2f}%로 "
                    f"{abs(bp)}bp {direction}**되었습니다. 위 그래프의 '{HIST_YEAR_END}년' 막대는 이 최신 금리를 "
                    "반영합니다. 아래 대출별 스프레드 평가(1차 지표=기준금리 대비)는 대출금리 변동분에서 이 기준금리 "
                    "변동분을 이미 자동으로 차감하고 있어, 대출금리가 기준금리와 같은 폭(예: 이번처럼 25bp)만큼만 "
                    "움직였다면 스프레드는 그대로 유지된 것으로 — 즉 시장 전체가 움직인 결과일 뿐 협상 성과와는 "
                    "무관한 것으로 — 평가됩니다."
                )

            if all(v is None for v in avg_bank):
                st.info("업로드한 엑셀의 '과거 이자율변동내역' 시트에서 유효한 금리 데이터를 찾지 못했습니다. 시트 구성을 확인해주세요.")
            else:
                fig1 = go.Figure()
                fig1.add_trace(go.Bar(
                    x=years,
                    y=bok_data,
                    name='한국은행 기준금리',
                    marker_color='#f43f5e',  # soft premium red
                    text=[f"{v:.2f}%" for v in bok_data],
                    textposition='inside',
                ))
                fig1.add_trace(go.Bar(
                    x=years,
                    y=avg_bank,
                    name='시중은행 가중평균',
                    marker_color='#3b82f6',  # soft premium blue
                    text=[f"{v:.2f}%" if v is not None else "" for v in avg_bank],
                    textposition='inside',
                ))
                fig1.update_layout(
                    barmode='group',
                    plot_bgcolor='rgba(0,0,0,0)',
                    paper_bgcolor='rgba(0,0,0,0)',
                    font=dict(color='#0f172a'),
                    yaxis=dict(title='금리 (%)', gridcolor='rgba(15,23,42,0.06)'),
                    xaxis=dict(gridcolor='rgba(15,23,42,0.06)'),
                    legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1)
                )
                st.plotly_chart(fig1, use_container_width=True)
 
        elif page == PAGE_RATE_BANK:
            st.markdown("### 🏢 금융기관별 금리 비교 (연도별)")
            st.write("선택한 연도의 금융기관별 금리를 낮은 순으로 정렬하여 비교합니다. 빨간 점선은 해당 연도의 한국은행 기준금리입니다.")
            st.caption(
                f"※ {YEAR_COLS[-1]}은 '차입금현황(현재)' 표를 은행별로 잔액 가중평균해 실시간으로 계산합니다"
                "(차입금현황이 바뀌면 이 그래프도 같이 바뀝니다). 그 이전 연도는 '과거 이자율변동내역' 시트 값을 그대로 씁니다."
            )
            default_idx = len(YEAR_COLS) - 1  # 기본값: 당해년도
            selected_year = st.selectbox("조회 연도 선택", YEAR_COLS, index=default_idx)

            # Filter and sort — 당해년도(가장 최근 연도)는 '과거 이자율변동내역' 시트에
            # 따로 타이핑해둔 값 대신 '★ 차입금 관리내역(현재)' 표를 은행별로 직접 집계한
            # 값을 쓴다. 그래야 차입금현황(현재) 데이터를 고치면 이 화면도 같이 바뀐다 —
            # 예전에는 두 표가 따로 관리돼서 하나를 고쳐도 다른 하나가 그대로였다.
            if selected_year == YEAR_COLS[-1]:
                df_cur_for_rate, _ = load_current_loan_status(excel_bytes)
                live_summary = compute_current_bank_rate_summary(df_cur_for_rate)
                if live_summary:
                    df_year = pd.DataFrame(
                        [{"금융기관": bank, selected_year: rate} for bank, rate in live_summary.items()]
                    )
                    df_year = df_year[df_year[selected_year] > 0].sort_values(by=selected_year, ascending=True)
                else:
                    df_year = pd.DataFrame()
            elif selected_year in df_hist.columns:
                df_year = df_hist[['금융기관', selected_year]].copy()
                df_year = df_year[df_year[selected_year] > 0]
                df_year = df_year.sort_values(by=selected_year, ascending=True)
            else:
                df_year = pd.DataFrame()

            if not df_year.empty:
                fig2 = go.Figure()
 
                # Get BOK rate for selected year — 값이 없으면(드묾) 3.0% 같은 임의의
                # 숫자 대신, 조회 가능한 연도 중 가장 최근 값으로 대신한다.
                year_key = selected_year.replace('년', '')
                bok_rate_val = bok_rates.get(year_key)
                if bok_rate_val is None and bok_rates:
                    bok_rate_val = bok_rates[max(bok_rates.keys())]
                elif bok_rate_val is None:
                    bok_rate_val = 3.0
 
                fig2.add_trace(go.Bar(
                    x=df_year['금융기관'],
                    y=df_year[selected_year],
                    marker_color='#10b981',  # premium green
                    text=[f"{v:.2f}%" for v in df_year[selected_year]],
                    textposition='outside',
                    name='금융기관 금리'
                ))
 
                fig2.add_hline(
                    y=bok_rate_val,
                    line_dash="dash",
                    line_color="#f43f5e",
                    line_width=2,
                    annotation_text=f"한국은행 기준금리 ({bok_rate_val:.2f}%)",
                    annotation_position="bottom right"
                )
 
                fig2.update_layout(
                    plot_bgcolor='rgba(0,0,0,0)',
                    paper_bgcolor='rgba(0,0,0,0)',
                    font=dict(color='#0f172a'),
                    yaxis=dict(title='금리 (%)', gridcolor='rgba(15,23,42,0.06)', range=[0, max(df_year[selected_year].max() * 1.15, 6.0)]),
                    xaxis=dict(gridcolor='rgba(15,23,42,0.06)'),
                )
                st.plotly_chart(fig2, use_container_width=True)
            else:
                st.info("선택한 연도의 금리 데이터가 없습니다.")

            st.markdown("---")
            st.markdown("### 📈 개별 금융기관 금리 추이 (연도별 필터)")
            st.write("비교하고 싶은 금융기관을 선택하여 개별 추이를 분석할 수 있습니다.")
            all_banks = df_hist['금융기관'].unique().tolist()
            selected_banks = st.multiselect("비교할 금융기관 선택 (기본값: 상위 3개 기관)", all_banks, default=all_banks[:3])
 
            fig3 = go.Figure()
 
            # Weighted average and BOK as reference
            fig3.add_trace(go.Scatter(x=years, y=avg_bank, mode='lines+markers+text', name='시중은행 가중평균',
                                     line=dict(color='#3b82f6', width=4), marker=dict(size=8, symbol='diamond'),
                                     text=[f"{v:.2f}%" if v else "" for v in avg_bank], textposition="top right"))
 
            fig3.add_trace(go.Scatter(x=years, y=bok_data, mode='lines+markers+text', name='한국은행 기준금리(ECOS)',
                                     line=dict(color='#f43f5e', width=4, dash='dash'), marker=dict(size=8, symbol='star'),
                                     text=[f"{v:.2f}%" for v in bok_data], textposition="bottom right"))
 
            # Selected banks
            colors = ['#10b981', '#f59e0b', '#8b5cf6', '#ec4899', '#14b8a6', '#06b6d4', '#eab308']
            for i, bank in enumerate(selected_banks):
                row_matches = df_hist[df_hist['금융기관'] == bank]
                if not row_matches.empty:
                    row = row_matches.iloc[0]
                    rates = [row.get(y, None) for y in YEAR_COLS]
                    fig3.add_trace(go.Scatter(x=years, y=rates, mode='lines+markers', name=bank,
                                             opacity=0.7, line=dict(color=colors[i % len(colors)], width=2),
                                             marker=dict(size=6)))
 
            fig3.update_layout(
                plot_bgcolor='rgba(0,0,0,0)',
                paper_bgcolor='rgba(0,0,0,0)',
                font=dict(color='#0f172a'),
                yaxis=dict(title='이자율 (%)', gridcolor='rgba(15,23,42,0.06)'),
                xaxis=dict(gridcolor='rgba(15,23,42,0.06)'),
                legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1)
            )
            st.plotly_chart(fig3, use_container_width=True)
 
        elif page == PAGE_CREDIT:
            st.subheader("🔍 금융기관별 신용등급 변동 추이 (5개년)")
 
            try:
                xl_credit = pd.ExcelFile(io.BytesIO(excel_bytes))
                credit_sheets = [s for s in xl_credit.sheet_names if '신용등급' in s]
                if credit_sheets:
                    df_c_raw = xl_credit.parse(credit_sheets[0], header=None)
 
                    # 그래프 x축(연도) 라벨은 시스템 날짜(CURRENT_YEAR)가 아니라
                    # 엑셀 시트에 실제로 적힌 연도 헤더(6행)를 그대로 사용한다.
                    # 시스템 날짜 기준으로 라벨을 계산하면, 시트의 마지막 기록 연도와
                    # 실제 당해년도가 어긋날 때(예: 시트는 2025년까지만 있는데 시스템 날짜는 2026년)
                    # 라벨과 실제 데이터가 한 해씩 밀려서 등급이 잘못 표시되는 문제가 생긴다.
                    all_years_header = [str(df_c_raw.iloc[5, col]).strip() for col in range(2, 11)]
                    years_5 = all_years_header[-5:]
                    st.write(f"{years_5[0]}년부터 {years_5[-1]}년까지의 주요 금융기관 신용등급 변동 추이입니다. 선 그래프를 통해 각 은행의 신용도 추세를 한눈에 파악할 수 있으며, 하단에서 전체 로우데이터를 확인할 수 있습니다.")
 
                    chart_data = []
                    r_list_all = []
 
                    for r in range(6, 14):
                        row_c = df_c_raw.iloc[r]
                        bank_c = str(row_c.iloc[1]).strip()
 
                        # Full history for raw data table
                        h_all = {}
                        all_years = [str(df_c_raw.iloc[5, col]).strip() for col in range(2, 11)]
                        for idx, col in enumerate(range(2, 11)):
                            yr_c = all_years[idx]
                            v_c = str(row_c.iloc[col]).strip() if pd.notna(row_c.iloc[col]) else "-"
                            v_c = v_c.replace('\n', ' ')
                            h_all[yr_c] = v_c
                        r_list_all.append({'금융기관': bank_c, **h_all})
 
                        # 5-year numeric mapping for chart (columns 6~10 = last 5 years in sheet)
                        nums_5 = []
                        last_val = None
                        # 5개년 창(window) 시작 이전 컬럼(2~5)을 미리 훑어서, 만약 창의 첫 해(col=6)가
                        # '전년등급 활용'처럼 실제 등급 없이 이전 등급을 그대로 쓰라는 표시일 경우에도
                        # 그 이전에 실제로 기록된 등급을 그대로 이어받을 수 있도록 한다.
                        for col in range(2, 6):
                            prev_str = str(row_c.iloc[col]).strip() if pd.notna(row_c.iloc[col]) else ""
                            prev_val = rating_to_numeric(prev_str)
                            if prev_val is not None:
                                last_val = prev_val

                        for col in range(6, 11):
                            val_str = str(row_c.iloc[col]).strip() if pd.notna(row_c.iloc[col]) else ""
                            num_val = rating_to_numeric(val_str)

                            if num_val is None:
                                if ('전년등급' in val_str or val_str == '-' or not val_str) and last_val is not None:
                                    num_val = last_val
                            if num_val is not None:
                                last_val = num_val

                            nums_5.append(num_val)
 
                        chart_data.append({
                            '금융기관': bank_c,
                            'ratings': nums_5
                        })
 
                    df_c_table = pd.DataFrame(r_list_all)
 
                    # Plotly chart
                    fig_credit = go.Figure()
                    colors_credit = ['#10b981', '#f59e0b', '#8b5cf6', '#ec4899', '#14b8a6', '#06b6d4', '#eab308', '#3b82f6']
 
                    for i, item in enumerate(chart_data):
                        matched = df_c_table[df_c_table['금융기관'] == item['금융기관']]
                        text_vals = [matched.iloc[0][yr] if (not matched.empty and yr in matched.columns) else "-" for yr in years_5]
                        fig_credit.add_trace(go.Scatter(
                            x=years_5,
                            y=item['ratings'],
                            mode='lines+markers',
                            name=item['금융기관'],
                            line=dict(color=colors_credit[i % len(colors_credit)], width=3),
                            marker=dict(size=8),
                            text=text_vals,
                            hovertemplate="<b>" + item['금융기관'] + "</b><br>연도: %{x}<br>등급: %{text}<extra></extra>"
                        ))
 
                    fig_credit.update_layout(
                        plot_bgcolor='rgba(0,0,0,0)',
                        paper_bgcolor='rgba(0,0,0,0)',
                        font=dict(color='#0f172a'),
                        yaxis=dict(
                            title='신용등급',
                            gridcolor='rgba(15,23,42,0.06)',
                            tickvals=[1, 2, 3, 4, 5, 6, 7, 8, 9, 10],
                            ticktext=['BBB-', 'BBB', 'BBB+', 'A-', 'A', 'A+', 'AA-', 'AA', 'AA+', 'AAA'],
                            range=[0.5, 10.5]
                        ),
                        xaxis=dict(
                            title='연도',
                            gridcolor='rgba(15,23,42,0.06)',
                            type='category',
                            categoryorder='array',
                            categoryarray=years_5
                        ),
                        legend=dict(orientation="h", yanchor="bottom", y=1.05, xanchor="right", x=1)
                    )
 
                    st.plotly_chart(fig_credit, use_container_width=True)
 
                    st.markdown("#### 📋 로우데이터 (전체 신용등급 현황)")
                    st.dataframe(df_c_table, use_container_width=True)
                else:
                    st.caption("신용등급현황 시트가 존재하지 않습니다.")
            except Exception as e:
                st.error(f"신용등급 로드 오류: {e}")

        elif page == PAGE_COLLATERAL:
            st.subheader("🏗 담보현황")
            st.markdown(
                '<button class="no-print" onclick="window.print()" '
                'style="background:#3182f6;color:white;border:none;padding:9px 16px;border-radius:6px;'
                'font-weight:700;cursor:pointer;margin-bottom:12px;font-size:0.9rem;">🖨 인쇄하기 / PDF 저장</button>',
                unsafe_allow_html=True,
            )
            st.write("차입금현황(현재) 표에서 '담보내역'이 기재된 대출만 모아, 담보로 제공한 물건지를 정리합니다.")
            coll_summary_df = pd.DataFrame()
            try:
                df_current_coll, sheet_used_coll = load_current_loan_status(excel_bytes)
                st.caption(f"📄 참조 시트: '{sheet_used_coll}'")

                if df_current_coll.empty:
                    st.info("표시할 차입금 현황 데이터가 없습니다.")
                else:
                    coll_detail_df, coll_summary_df = build_collateral_overview(df_current_coll)
                    if coll_detail_df.empty:
                        st.info("담보내역이 기재된 대출을 찾지 못했습니다. 엑셀의 '담보내역' 컬럼을 확인해주세요.")
                    else:
                        n_props = len(coll_summary_df)
                        n_loans = len(coll_detail_df)
                        total_secured = coll_detail_df['잔액(원)'].sum()
                        shared_props = coll_summary_df[coll_summary_df['담보대출건수'] > 1]

                        c1, c2, c3 = st.columns(3)
                        with c1:
                            st.markdown(f'<div class="metric-card"><div class="metric-label">담보 물건지 수</div><div class="metric-value">{n_props}개</div><div class="metric-desc">서로 다른 담보 기준</div></div>', unsafe_allow_html=True)
                        with c2:
                            st.markdown(f'<div class="metric-card"><div class="metric-label">담보 제공 대출 건수</div><div class="metric-value">{n_loans}건</div><div class="metric-desc">잔액 합계 {total_secured/100000000:,.0f}억원</div></div>', unsafe_allow_html=True)
                        with c3:
                            st.markdown(f'<div class="metric-card"><div class="metric-label">복수 대출 담보 물건</div><div class="metric-value">{len(shared_props)}개</div><div class="metric-desc">2건 이상 대출에 공동 담보</div></div>', unsafe_allow_html=True)

                        st.markdown("---")
                        st.markdown("#### 📍 물건지별 합계")
                        st.caption("같은 물건지가 여러 대출의 담보로 잡혀 있으면(공동담보) '담보대출건수'가 2건 이상으로 표시됩니다.")
                        summary_display = coll_summary_df.copy()
                        summary_display['잔액합계'] = summary_display['잔액합계'].map(lambda v: f"{v:,.0f}")
                        summary_column_config = {
                            "담보물건지": st.column_config.Column(width="large"),
                            "잔액합계": st.column_config.Column(width="small"),
                            "담보대출건수": st.column_config.Column(width="small"),
                            "담보제공기관": st.column_config.Column(width="medium"),
                        }
                        st.dataframe(summary_display, use_container_width=True, hide_index=True, column_config=summary_column_config)

                        st.markdown("#### 📋 대출별 상세")
                        st.caption(
                            "위 '물건지별 합계' 표에서 궁금한 물건지를 아래 드롭다운으로 골라도 됩니다 "
                            "(Streamlit은 표 더블클릭 이동을 지원하지 않아, 대신 필터로 바로 연결해뒀습니다)."
                        )
                        coll_site_options = ["(전체 보기)"] + list(coll_summary_df['담보물건지'])
                        picked_site = st.selectbox(
                            "🔎 담보물건지로 찾기", coll_site_options, key="collateral_site_filter",
                        )
                        detail_display = coll_detail_df.copy()
                        if picked_site != "(전체 보기)":
                            detail_display = detail_display[detail_display['담보물건지'] == picked_site]
                        detail_display['잔액(원)'] = detail_display['잔액(원)'].map(lambda v: f"{v:,.0f}")
                        detail_column_config = {
                            "담보물건지": st.column_config.Column(width="large"),
                            "금융기관": st.column_config.Column(width="small"),
                            "대출과목": st.column_config.Column(width="small"),
                            "구분": st.column_config.Column(width="small"),
                            "잔액(원)": st.column_config.Column(width="small"),
                            "만기일": st.column_config.Column(width="small"),
                        }
                        st.dataframe(detail_display, use_container_width=True, hide_index=True, column_config=detail_column_config)
            except Exception as e:
                st.error(f"담보현황 로드 오류: {e}")
                st.caption("엑셀의 '★ 차입금 관리내역(현재)' 시트명/구조를 확인해주세요.")

            st.markdown("---")
            st.markdown("### 🏢 부동산 자산 현황")
            st.caption(
                "팀에서 관리하는 부동산(담보 가능 자산) 현황입니다 — 위쪽 '차입금현황에서 뽑은 담보내역'과는 "
                "별도의 데이터로, 사업장별 장부가ㆍ감정평가액ㆍ담보제공 여부를 직접 관리합니다."
            )

            re_input_mode = st.radio(
                "입력 방식", ["📤 엑셀 업로드(전체 교체)", "✏️ 직접입력(표에서 편집)"],
                horizontal=True, key="real_estate_input_mode", label_visibility="collapsed",
            )

            if re_input_mode == "📤 엑셀 업로드(전체 교체)":
                re_file = st.file_uploader(
                    "부동산 현황 엑셀 파일(.xlsx) — '담보현황_사업장별' 시트 포함", type=["xlsx"],
                    key="real_estate_uploader", help="예: ★ (주)동양 부동산 현황.xlsx",
                )
                if re_file is not None:
                    try:
                        re_wb = openpyxl.load_workbook(io.BytesIO(re_file.getvalue()), data_only=True)
                        re_preview_rows = real_estate_db.parse_real_estate_workbook(re_wb)
                        re_preview_summary = real_estate_db.parse_real_estate_summary(re_wb)
                    except Exception as e:
                        re_preview_rows = None
                        re_preview_summary = []
                        st.error(f"파싱 실패: {e}")
                    if re_preview_rows is not None:
                        if not re_preview_rows:
                            st.error("'담보현황_사업장별' 시트에서 사업장을 하나도 찾지 못했습니다 — 시트 구조를 확인해주세요.")
                        else:
                            st.caption(f"미리보기: 총 {len(re_preview_rows)}개 사업장을 찾았습니다(업무용/비업무용 포함).")
                            if st.button("💾 저장 (기존 데이터 전체 교체)", type="primary", key="real_estate_save_upload"):
                                real_estate_db.replace_all_sites(re_preview_rows, source=re_file.name)
                                real_estate_db.replace_summary(re_preview_summary, source=re_file.name)
                                st.session_state['real_estate_edit_df_loaded'] = False
                                st.success(f"부동산 현황 {len(re_preview_rows)}건을 저장했습니다.")
                                st.rerun()
            else:
                st.caption(
                    "표에서 직접 추가ㆍ수정ㆍ삭제할 수 있습니다. 저장을 누르면 기존 데이터 전체가 지금 표 내용으로 교체됩니다."
                )
                re_edit_columns = [
                    "구분", "사업장", "토지(원)", "건물(원)", "소계(원)", "감정평가액(원)",
                    "평가년도", "담보은행", "담보내역", "비고", "사업부",
                ]
                if not st.session_state.get('real_estate_edit_df_loaded'):
                    current_sites = real_estate_db.list_sites()
                    st.session_state['real_estate_edit_df'] = pd.DataFrame([
                        {
                            "구분": r["category"], "사업장": r["site_name"],
                            "토지(원)": r["land_value"], "건물(원)": r["building_value"], "소계(원)": r["subtotal_value"],
                            "감정평가액(원)": r["appraisal_value"], "평가년도": r["appraisal_year"],
                            "담보은행": r["bank"], "담보내역": r["collateral_detail"], "비고": r["note"], "사업부": r["biz_unit"],
                        }
                        for r in current_sites
                    ]) if current_sites else pd.DataFrame(columns=re_edit_columns)
                    st.session_state['real_estate_edit_df_loaded'] = True

                re_edited_df = st.data_editor(
                    st.session_state['real_estate_edit_df'],
                    num_rows="dynamic", use_container_width=True, key="real_estate_editor",
                    column_config={
                        "구분": st.column_config.SelectboxColumn(options=["업무용", "비업무용"]),
                        "소계(원)": st.column_config.NumberColumn(
                            disabled=True, help="토지+건물을 저장할 때 자동으로 계산합니다 — 직접 입력할 수 없습니다.",
                        ),
                    },
                )
                if st.button("💾 저장", type="primary", key="real_estate_save_manual"):
                    clean_re_df = re_edited_df.dropna(how='all')
                    rows_to_save = []
                    for _, r in clean_re_df.iterrows():
                        if not str(r.get('사업장') or '').strip():
                            continue
                        land = pd.to_numeric(r.get('토지(원)'), errors='coerce')
                        building = pd.to_numeric(r.get('건물(원)'), errors='coerce')
                        # 소계는 사람이 직접 입력한 값을 믿지 않고 토지+건물로 항상 다시 계산한다
                        # (건물 가격만 고쳐도 소계가 저절로 맞게 저장되도록 — 사용자 요청).
                        subtotal = None if (pd.isna(land) and pd.isna(building)) else (
                            (0 if pd.isna(land) else land) + (0 if pd.isna(building) else building)
                        )
                        rows_to_save.append({
                            "category": r.get('구분') or '업무용',
                            "site_name": str(r.get('사업장')).strip(),
                            "land_value": None if pd.isna(land) else land,
                            "building_value": None if pd.isna(building) else building,
                            "subtotal_value": subtotal,
                            "appraisal_value": pd.to_numeric(r.get('감정평가액(원)'), errors='coerce'),
                            "appraisal_year": r.get('평가년도'),
                            "bank": r.get('담보은행'),
                            "collateral_detail": r.get('담보내역'),
                            "note": r.get('비고'),
                            "biz_unit": r.get('사업부'),
                        })
                    if not rows_to_save:
                        st.error("최소 1개 사업장 이상 입력해주세요.")
                    else:
                        real_estate_db.replace_all_sites(rows_to_save)
                        st.session_state['real_estate_edit_df_loaded'] = False
                        st.success(f"부동산 현황 {len(rows_to_save)}건을 저장했습니다.")
                        st.rerun()

            st.markdown("---")
            re_sites = real_estate_db.list_sites()
            if not re_sites:
                st.info("아직 등록된 부동산 현황이 없습니다. 위에서 엑셀을 업로드하거나 직접 입력해주세요.")
            else:
                re_updated_at = real_estate_db.get_meta('updated_at')
                if re_updated_at:
                    st.caption(f"마지막 업데이트: {re_updated_at}")

                re_diffs = find_real_estate_loan_diffs(re_sites, coll_summary_df)
                if re_diffs:
                    st.warning(
                        f"⚠️ 차입금현황과 대조했을 때 다시 확인이 필요한 항목 {len(re_diffs)}건이 있습니다(노란색 행) "
                        "— 사업장 이름 표기 차이로 인한 오탐일 수 있으니 실제로 한 번 확인해주세요."
                    )

                def _re_fmt(v):
                    return f"{v:,.0f}" if isinstance(v, (int, float)) else (v or '-')

                def _re_total_row(label, rows):
                    land_sum = sum((r['land_value'] or 0) for r in rows)
                    bldg_sum = sum((r['building_value'] or 0) for r in rows)
                    sub_sum = sum((r['subtotal_value'] or 0) for r in rows)
                    appr_sum = sum((r['appraisal_value'] or 0) for r in rows)
                    return (
                        '<tr style="background:rgba(15,23,42,0.06);font-weight:700;">'
                        f'<td>{label}</td>'
                        f'<td style="text-align:right;">{_re_fmt(land_sum)}</td>'
                        f'<td style="text-align:right;">{_re_fmt(bldg_sum)}</td>'
                        f'<td style="text-align:right;">{_re_fmt(sub_sum)}</td>'
                        f'<td style="text-align:right;">{_re_fmt(appr_sum)}</td>'
                        '<td></td><td></td><td></td><td></td><td></td>'
                        '</tr>'
                    )

                for cat in ["업무용", "비업무용"]:
                    cat_rows = [r for r in re_sites if r['category'] == cat]
                    if not cat_rows:
                        continue
                    cat_total = sum((r['subtotal_value'] or 0) for r in cat_rows)
                    st.markdown(f"#### 📍 {cat} ({len(cat_rows)}건, 장부가 합계 {cat_total/1_000_000:,.0f}백만원)")

                    table_rows_html = []
                    for r in cat_rows:
                        diff_msg = re_diffs.get(r['id'])
                        row_style = 'background:rgba(217,119,6,0.15);' if diff_msg else ''
                        table_rows_html.append(
                            f'<tr style="{row_style}">'
                            f'<td>{r["site_name"]}</td>'
                            f'<td style="text-align:right;">{_re_fmt(r["land_value"])}</td>'
                            f'<td style="text-align:right;">{_re_fmt(r["building_value"])}</td>'
                            f'<td style="text-align:right;">{_re_fmt(r["subtotal_value"])}</td>'
                            f'<td style="text-align:right;">{_re_fmt(r["appraisal_value"])}</td>'
                            f'<td style="text-align:center;">{r["appraisal_year"] or "-"}</td>'
                            f'<td>{r["bank"] or "-"}</td>'
                            f'<td>{r["collateral_detail"] or "-"}</td>'
                            f'<td>{r["biz_unit"] or "-"}</td>'
                            f'<td style="color:var(--danger);font-size:0.78rem;">{diff_msg or ""}</td>'
                            f'</tr>'
                        )
                    table_rows_html.append(_re_total_row(f"{cat}계", cat_rows))
                    table_html = (
                        '<table class="real-estate-table" style="border-collapse:collapse;width:100%;font-size:0.78rem;">'
                        '<thead><tr style="background:rgba(15,23,42,0.04);">'
                        '<th>사업장</th><th>토지(원)</th><th>건물(원)</th><th>소계(원)</th>'
                        '<th>감정평가액(원)</th><th>평가년도</th><th>담보은행</th><th>담보내역</th>'
                        '<th>사업부</th><th>차입금현황 대조</th>'
                        '</tr></thead><tbody>' + ''.join(table_rows_html) + '</tbody></table>'
                    )
                    st.markdown(table_html, unsafe_allow_html=True)

                st.markdown("---")
                st.markdown("#### 📎 담보 서류 및 품의서 링크")
                st.caption(
                    "사업장별로 품의서 링크(그룹웨어 완료함에서 찾은 URL)와 근저당권설정계약서ㆍ등기필증ㆍ"
                    "채권최고액 확인서류 등을 등록해두면, 팀원 누구나 다시 찾아볼 수 있습니다."
                )

                # 담보로 실제 잡혀있는(담보은행이 채워진) 사업장만 골라 보여준다 —
                # 담보 가능(미제공) 사업장까지 다 나오면 목록이 너무 길고, 애초에
                # 연결할 품의서 자체가 없는 곳들이라 의미도 없다.
                pledged_sites = [r for r in re_sites if (r.get('bank') or '').strip()]
                if not pledged_sites:
                    st.info("담보은행이 등록된 사업장이 없습니다 — 담보내역 표에 담보은행을 먼저 입력해주세요.")
                    doc_site_row = None
                else:
                    doc_site_options = {f"{r['site_name']} ({r['category']})": r['id'] for r in pledged_sites}
                    with st.container(border=True):
                        st.markdown("**1️⃣ 사업장을 먼저 선택하세요** (담보은행이 등록된 사업장만 표시)")
                        doc_site_label = st.selectbox(
                            "서류를 등록/확인할 사업장", list(doc_site_options.keys()),
                            key="re_doc_site_select", label_visibility="collapsed",
                        )
                        doc_site_id = doc_site_options[doc_site_label]
                        st.info(f"👉 현재 선택된 사업장: **{doc_site_label}**", icon="📍")
                    doc_site_row = next(r for r in pledged_sites if r['id'] == doc_site_id)

                if doc_site_row is not None:

                    with st.container(border=True):
                        st.markdown(f"**2️⃣ 품의서 링크 — {doc_site_label}**")
                        new_approval_url = st.text_input(
                            "🔗 품의서 링크(그룹웨어 URL)", value=doc_site_row.get('approval_doc_url') or '',
                            key=f"re_approval_url__{doc_site_id}",
                            help="그룹웨어 완료함에서 해당 대출 품의서를 열어 주소창 URL을 그대로 붙여넣으세요.",
                        )
                        if st.button("🔗 이 사업장에 링크 저장", key=f"re_approval_url_save__{doc_site_id}", type="primary"):
                            real_estate_db.update_site(doc_site_id, approval_doc_url=new_approval_url.strip() or None)
                            st.success(f"'{doc_site_label}'의 품의서 링크를 저장했습니다.")
                            st.rerun()
                        if doc_site_row.get('approval_doc_url'):
                            st.markdown(f"📄 저장된 링크: [{doc_site_row['approval_doc_url']}]({doc_site_row['approval_doc_url']})")
                        else:
                            st.caption("아직 저장된 링크가 없습니다.")

                        st.markdown("---")
                        st.caption(
                            f"또는 그룹웨어 완료함에서 담보은행({doc_site_row.get('bank') or '정보없음'}) 기준으로 "
                            "자동 검색해서 링크와 첨부 서류를 한 번에 가져올 수 있습니다."
                        )
                        if not _collateral_doc_search_ready():
                            st.caption("⚙️ 자동 검색은 최초 1회 설정이 필요합니다 (아래 안내 참고).")
                        if st.button("🔍 그룹웨어에서 자동 검색", key=f"re_gw_search__{doc_site_id}"):
                            if not doc_site_row.get('bank'):
                                st.error("이 사업장엔 담보은행 정보가 없어 자동 검색을 할 수 없습니다. 수동으로 링크를 저장해주세요.")
                            else:
                                with st.spinner(f"그룹웨어 완료함에서 '{doc_site_row['bank']}' 관련 품의서 검색 중... (최대 3분)"):
                                    ok, msg = run_collateral_doc_search(doc_site_row['site_name'])
                                if ok:
                                    st.success("검색 완료! 결과:")
                                    st.code(msg, language=None)
                                    st.rerun()
                                else:
                                    st.error(msg)

                    with st.container(border=True):
                        st.markdown(f"**3️⃣ 서류 업로드 — {doc_site_label}**")
                        st.caption("업로드 후 반드시 아래 '이 서류 저장' 버튼을 눌러야 저장됩니다.")
                        re_doc_upload = st.file_uploader(
                            "📎 파일 선택 (근저당권설정계약서ㆍ등기필증ㆍ채권최고액 확인서류 등)",
                            key=f"re_doc_upload__{doc_site_id}",
                        )
                        re_up_col1, re_up_col2 = st.columns([1, 1])
                        with re_up_col1:
                            re_doc_type = st.selectbox(
                                "서류 종류", real_estate_db.DOC_TYPES, key=f"re_doc_type__{doc_site_id}",
                            )
                        with re_up_col2:
                            re_doc_uploader_name = st.text_input(
                                "업로드자(선택)", key=f"re_doc_uploader__{doc_site_id}",
                            )
                        if re_doc_upload is None:
                            st.button("💾 이 서류 저장", key=f"re_doc_save__{doc_site_id}", disabled=True,
                                      help="먼저 위에서 파일을 선택하세요.")
                        elif st.button("💾 이 서류 저장", key=f"re_doc_save__{doc_site_id}", type="primary"):
                            real_estate_db.add_document(
                                doc_site_id, re_doc_type, re_doc_upload.name, re_doc_upload.getvalue(),
                                uploaded_by=re_doc_uploader_name.strip() or None,
                            )
                            st.success(f"'{re_doc_upload.name}' 저장했습니다. 아래 목록에서 다운로드할 수 있습니다.")
                            st.rerun()

                    st.markdown(f"**📂 '{doc_site_label}'에 등록된 서류**")
                    re_docs = real_estate_db.list_documents(doc_site_id)
                    if re_docs:
                        for d in re_docs:
                            ddl_col1, ddl_col2, ddl_col3 = st.columns([3, 1, 1])
                            with ddl_col1:
                                uploader_suffix = f" ({d['uploaded_by']})" if d['uploaded_by'] else ""
                                st.write(f"**[{d['doc_type']}]** {d['original_filename']} — {d['uploaded_at']}{uploader_suffix}")
                            with ddl_col2:
                                doc_file = real_estate_db.get_document_file(d['id'])
                                if doc_file:
                                    st.download_button(
                                        "⬇ 다운로드", data=doc_file[1], file_name=doc_file[0],
                                        key=f"re_doc_dl__{d['id']}",
                                    )
                            with ddl_col3:
                                if st.button("🗑 삭제", key=f"re_doc_del__{d['id']}"):
                                    real_estate_db.delete_document(d['id'])
                                    st.rerun()
                    else:
                        st.caption("등록된 서류가 없습니다.")

                st.markdown("#### 📊 전체 합계 현황")
                summary_rows = real_estate_db.list_summary()
                if summary_rows:
                    st.caption("부동산 현황 파일 원본의 '담보 제공중 / 담보 불가 / 담보 가능' 요약표를 그대로 반영했습니다.")

                    def _re_pct(v):
                        return f"{v * 100:,.1f}%" if isinstance(v, (int, float)) else '-'

                    summary_rows_html = []
                    for row in summary_rows:
                        is_total = row['label'] == '합계'
                        row_style = 'font-weight:700;background:rgba(15,23,42,0.06);' if is_total else ''
                        summary_rows_html.append(
                            f'<tr style="{row_style}">'
                            f'<td>{row["label"]}</td>'
                            f'<td style="text-align:right;">{_re_fmt(row["land_value"])}</td>'
                            f'<td style="text-align:right;">{_re_fmt(row["building_value"])}</td>'
                            f'<td style="text-align:right;">{_re_fmt(row["subtotal_value"])}</td>'
                            f'<td style="text-align:right;">{_re_fmt(row["appraisal_value"])}</td>'
                            f'<td style="text-align:right;">{_re_pct(row["land_ratio"])}</td>'
                            f'<td style="text-align:right;">{_re_pct(row["building_ratio"])}</td>'
                            f'<td style="text-align:right;">{_re_pct(row["subtotal_ratio"])}</td>'
                            f'<td style="font-size:0.75rem;color:var(--text-secondary);">{row["note"] or ""}</td>'
                            f'</tr>'
                        )
                    summary_html = (
                        '<table class="real-estate-table" style="border-collapse:collapse;width:100%;font-size:0.8rem;">'
                        '<thead><tr style="background:rgba(15,23,42,0.04);">'
                        '<th>구분</th><th>토지(원)</th><th>건물(원)</th><th>소계(원)</th><th>감정평가액(원)</th>'
                        '<th>토지비율</th><th>건물비율</th><th>소계비율</th><th>비고</th>'
                        '</tr></thead><tbody>' + ''.join(summary_rows_html) + '</tbody></table>'
                    )
                    st.markdown(summary_html, unsafe_allow_html=True)
                else:
                    st.caption("업로드한 파일에 요약표가 없어 현재 등록된 사업장을 그대로 합산한 값입니다.")
                    grand_total_html = (
                        '<table class="real-estate-table" style="border-collapse:collapse;width:100%;font-size:0.78rem;">'
                        '<thead><tr style="background:rgba(15,23,42,0.04);">'
                        '<th>구분</th><th>토지(원)</th><th>건물(원)</th><th>소계(원)</th>'
                        '<th>감정평가액(원)</th><th></th><th></th><th></th><th></th><th></th>'
                        '</tr></thead><tbody>' + _re_total_row("합계", re_sites) + '</tbody></table>'
                    )
                    st.markdown(grand_total_html, unsafe_allow_html=True)

        elif page == PAGE_LOAN:
            st.subheader("📋 차입금현황표 (현재)")
            st.write("업로드한 엑셀의 '★ 차입금 관리내역(현재)' 시트를 기준으로 한 차입금 현황입니다.")
            try:
                snapshot_sheets = list_loan_snapshot_sheets(excel_bytes)
                chosen_sheet = None
                if len(snapshot_sheets) > 1:
                    snapshot_labels = [
                        (f"{d.strftime('%Y-%m-%d')} 기준" if d is not None else name)
                        for name, d in snapshot_sheets
                    ]
                    picked_label = st.selectbox(
                        "조회 시점(스냅샷) 선택",
                        options=snapshot_labels,
                        index=0,
                        # 사이드바에서 다른 월/스냅샷 파일로 전환하면 이 안에 들어있는 '현재' 시트
                        # 목록 자체가 통째로 바뀌는데, key가 고정되어 있으면 이전 파일에서 골랐던
                        # 값이 세션에 남아있다가 새 파일의 옵션 목록에 없어 선택이 안 먹히거나
                        # 에러로 화면이 안 바뀌는 문제가 있었다 — 파일명을 key에 포함해 파일이
                        # 바뀌면 이 선택도 항상 첫 번째 옵션으로 새로 시작하게 한다.
                        key=f"loan_status_snapshot__{st.session_state.get('excel_filename', '')}",
                        help="엑셀에 저장된 여러 시점의 '현재' 시트 중 하나를 골라 그 시점 기준 차입금 현황을 볼 수 있습니다 "
                             "(분기말 스냅샷을 남겨두면 분기별 조회로 활용할 수 있습니다).",
                    )
                    chosen_sheet = snapshot_sheets[snapshot_labels.index(picked_label)][0]

                df_current, sheet_used = load_current_loan_status(excel_bytes, sheet_name_override=chosen_sheet)
                st.caption(f"📄 참조 시트: '{sheet_used}'")

                if df_current.empty:
                    st.info("표시할 차입금 현황 데이터가 없습니다. 시트 구성을 확인해주세요.")
                else:
                    balance_col = next((c for c in df_current.columns if '잔액' in c), None)
                    # '금리' 컬럼을 정확히 찾는다.
                    # (버그) '금리구분', '금리변경주기(개월)' 컬럼도 '금리'로 시작하기 때문에
                    # startswith만 쓰면 이 컬럼들이 먼저 잡혀버려서(값이 '변동금리'/'600.00%' 같은
                    # 텍스트라 숫자 변환 실패) 가중평균 금리가 항상 '-'로 표시되는 문제가 있었음.
                    # → 정확히 '금리'인 컬럼을 우선 찾고, 없을 때만 '구분/변경/조건'이 없는
                    #   금리로 시작하는 컬럼을 대체 후보로 사용한다.
                    rate_col = next((c for c in df_current.columns if c.strip() == '금리'), None)
                    if rate_col is None:
                        rate_col = next(
                            (c for c in df_current.columns
                             if ('이자율' in c) or (c.strip().startswith('금리') and not any(x in c for x in ('구분', '변경', '조건')))),
                            None
                        )
                    is_total_row = df_current['금융기관'].astype(str).str.strip().isin(['계', '합계'])

                    if balance_col is not None:
                        bal_numeric = pd.to_numeric(df_current[balance_col], errors='coerce')
                        total_now = bal_numeric[~is_total_row].sum()

                        col_a, col_b, col_c = st.columns(3)
                        with col_a:
                            st.markdown(f'<div class="metric-card"><div class="metric-label">총 차입잔액 (표 합계)</div><div class="metric-value">{total_now:,.1f}</div><div class="metric-desc">단위: 시트 원본 그대로</div></div>', unsafe_allow_html=True)
                        with col_b:
                            n_banks = df_current[~is_total_row]['금융기관'].nunique()
                            st.markdown(f'<div class="metric-card"><div class="metric-label">거래 금융기관 수</div><div class="metric-value">{n_banks}개</div><div class="metric-desc">현재 기준</div></div>', unsafe_allow_html=True)
                        with col_c:
                            weighted_avg = "-"
                            if rate_col is not None:
                                # 엑셀의 '합계' 행 금리(=이자합계/잔액합계로 계산되어 있음)를 그대로 사용
                                total_rate_series = pd.to_numeric(df_current.loc[is_total_row, rate_col], errors='coerce')
                                if not total_rate_series.dropna().empty:
                                    weighted_avg = f"{total_rate_series.dropna().iloc[0] * 100:.2f}%"
                            st.markdown(f'<div class="metric-card"><div class="metric-label">가중평균 금리 (표 기준)</div><div class="metric-value">{weighted_avg}</div><div class="metric-desc">엑셀 합계행 기준</div></div>', unsafe_allow_html=True)
                        st.markdown("---")

                    # ------------------------------------------------------------
                    # 한눈에 보기: 좌측 상세내역 표 / 우측 금융기관별 잔액 비중 원형도표
                    # ------------------------------------------------------------
                    detail_col, pie_col = st.columns([3, 2])

                    with detail_col:
                        with st.container(border=True):
                            st.markdown("#### 📋 상세내역")

                            # 필터 UI: ① 표시할 컬럼 선택, ② 잔액 유무로 행 필터링
                            all_columns = df_current.columns.tolist()
                            # 컬럼이 많으면(전형적으로 15개 이상) 표가 화면 폭을 넘어가 잘려 보이므로,
                            # 기본값은 핵심 항목만 우선 선택하고 나머지는 필요할 때 사용자가 직접 추가하도록 한다.
                            priority_columns = [
                                '금융기관', '대출과목', '금리구분', '만기일', '차입한도(원)', '잔액(원)', '금리'
                            ]
                            default_columns = [c for c in priority_columns if c in all_columns] or all_columns
                            selected_columns = st.multiselect(
                                "표시할 항목(컬럼) 선택",
                                options=all_columns,
                                default=default_columns,
                                key="loan_status_columns",
                                help="기본으로는 핵심 항목만 표시됩니다. 필요한 열을 추가/삭제해서 볼 수 있어요.",
                            )
                            balance_filter = st.radio(
                                "잔액 표시",
                                options=["전체", "잔액 있는 건만", "잔액 없는 건만"],
                                index=0,
                                key="loan_status_balance_filter",
                                horizontal=True,
                                help="잔액(원)이 있는 대출만 보거나, 없는 대출만 볼 수 있어요.",
                            )

                            view_df = df_current.copy()

                            if balance_col is not None and balance_filter != "전체":
                                bal_numeric_all = pd.to_numeric(view_df[balance_col], errors='coerce')
                                has_balance = bal_numeric_all.notna() & (bal_numeric_all != 0)
                                if balance_filter == "잔액 있는 건만":
                                    view_df = view_df[has_balance | is_total_row.reindex(view_df.index, fill_value=False)]
                                else:  # 잔액 없는 건만
                                    view_df = view_df[(~has_balance) | is_total_row.reindex(view_df.index, fill_value=False)]

                            if not selected_columns:
                                st.warning("표시할 컬럼을 1개 이상 선택해주세요.")
                            else:
                                fmt_df = view_df[selected_columns].copy()
                                for col in fmt_df.columns:
                                    numeric_col = pd.to_numeric(fmt_df[col], errors='coerce')
                                    if numeric_col.notna().mean() > 0.5:
                                        if '금리' in col or '이자율' in col:
                                            fmt_df[col] = numeric_col.map(lambda v: f"{v*100:.2f}%" if pd.notna(v) else "-")
                                        elif any(k in col for k in ('잔액', '한도', '금액', '이자')):
                                            fmt_df[col] = numeric_col.map(lambda v: f"{v:,.0f}" if pd.notna(v) else "-")

                                st.dataframe(fmt_df, use_container_width=True, hide_index=True)
                                st.caption(f"총 {len(view_df):,}행 표시 중 (전체 {len(df_current):,}행)")

                    with pie_col:
                        with st.container(border=True):
                            st.markdown("#### 🥧 금융기관별 잔액 비중")
                            if balance_col is None:
                                st.caption("잔액 컬럼을 찾을 수 없어 원형도표를 표시할 수 없습니다.")
                            else:
                                # 상세내역의 '잔액 표시' 필터와 동일한 대상(view_df)을 기준으로,
                                # 합계행을 제외하고 금융기관별 잔액을 합산해 비중을 계산한다.
                                pie_src = view_df[~view_df['금융기관'].astype(str).str.strip().isin(['계', '합계'])].copy()
                                pie_src['_잔액숫자'] = pd.to_numeric(pie_src[balance_col], errors='coerce').fillna(0)
                                pie_group = pie_src.groupby('금융기관')['_잔액숫자'].sum()
                                pie_group = pie_group[pie_group > 0].sort_values(ascending=False)

                                if pie_group.empty:
                                    st.caption("표시할 잔액 데이터가 없습니다.")
                                else:
                                    inst_colors = [
                                        '#3b82f6', '#10b981', '#06b6d4', '#f59e0b', '#c084fc',
                                        '#f472b6', '#ec4899', '#6366f1', '#14b8a6', '#64748b'
                                    ]
                                    fig_loan_pie = go.Figure(data=[go.Pie(
                                        labels=pie_group.index,
                                        values=pie_group.values,
                                        hole=0.55,
                                        marker=dict(
                                            colors=(inst_colors * (len(pie_group) // len(inst_colors) + 1))[:len(pie_group)],
                                            line=dict(color='#0f172a', width=2)
                                        ),
                                        textinfo='percent',
                                        textfont=dict(color='white', size=12),
                                        hovertemplate='%{label}<br>%{value:,.0f}<br>%{percent}<extra></extra>'
                                    )])
                                    fig_loan_pie.update_layout(
                                        showlegend=True,
                                        legend=dict(orientation="v", yanchor="middle", y=0.5, xanchor="left", x=1.02, font=dict(size=11, color='#64748b')),
                                        paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)',
                                        font=dict(color='#0f172a'),
                                        margin=dict(t=10, b=10, l=10, r=10),
                                        height=340,
                                        annotations=[dict(
                                            text=f"{pie_group.sum()/100000000:,.0f}억원",
                                            x=0.5, y=0.5, font=dict(size=15, color='#0f172a'), showarrow=False
                                        )]
                                    )
                                    st.plotly_chart(fig_loan_pie, use_container_width=True)
                                    st.caption(f"총 {len(pie_group)}개 기관 · 잔액 합계 {pie_group.sum():,.0f}")
            except Exception as e:
                st.error(f"차입금현황표 로드 오류: {e}")
                st.caption("엑셀의 '★ 차입금 관리내역(현재)' 시트명/구조를 확인해주세요.")

        elif page == PAGE_INTEREST:
            st.subheader("💵 원리금관리")
            st.write(
                "차입금현황(현재) 표의 '금리변경주기', '최근금리변경일', '이자지급일' 컬럼을 기준으로 "
                "① 변동금리 대출의 다음 금리변경 예정일과 ② 다음 이자지급일에 예상되는 이자를 계산합니다."
            )
            try:
                df_current, sheet_used = load_current_loan_status(excel_bytes)
                st.caption(f"📄 참조 시트: '{sheet_used}'")

                if df_current.empty:
                    st.info("표시할 차입금 현황 데이터가 없습니다. 시트 구성을 확인해주세요.")
                else:
                    as_of_date = st.date_input(
                        "기준일",
                        value=datetime.now().date(),
                        key="interest_as_of_date",
                        help="이 날짜를 기준으로 '다음 금리변경 예정일'과 '다음 이자지급일'을 계산합니다.",
                    )

                    # ------------------------------------------------------------
                    # ⓪ 이자 지급 전망 (월별 · 연별 · 기관별 한눈에 보기)
                    # ------------------------------------------------------------
                    with st.container(border=True):
                        st.markdown("#### 📆 이자 지급 전망 (월별·연별·기관별)")
                        st.caption(
                            "단순히 '잔액 × 현재금리'로 1년치를 어림잡지 않고, 아래 ②의 실제 지급주기·일수·영업일 조정 "
                            "로직을 그대로 재사용해 월별/연별/기관별 합계를 계산합니다 — 그래서 별도 입력값 없이도 이 "
                            "수치가 더 정확합니다. 다만 전망 기간 안에 금리변경이 예정된 구간(대출별 상세의 비고 참고)은 "
                            "변경 후 확정금리를 알 수 없어 현재금리로 잠정 산정한 값이라, 그 이후로 갈수록 실제와는 차이가 "
                            "날 수 있습니다. 더 정밀하게 보려면 예상 금리 변동분을 직접 입력하는 기능이 필요한데, 필요하시면 말씀해주세요."
                        )
                        outlook_period_options = {"6개월": 6, "12개월": 12, "24개월": 24}
                        outlook_period_label = st.radio(
                            "전망 기간", list(outlook_period_options.keys()), index=1,
                            key="interest_outlook_period", horizontal=True,
                        )
                        months_ahead = outlook_period_options[outlook_period_label]
                        monthly_df, yearly_df, by_bank_df = build_interest_outlook(df_current, as_of_date, months_ahead)

                        if monthly_df.empty:
                            st.caption("전망할 예상이자 데이터가 있는 대출이 없습니다.")
                        else:
                            outlook_total = monthly_df['예상이자합계'].sum()
                            outlook_avg_monthly = monthly_df['예상이자합계'].mean()
                            oc1, oc2 = st.columns(2)
                            with oc1:
                                st.markdown(f'<div class="metric-card"><div class="metric-label">{outlook_period_label} 예상이자 합계</div><div class="metric-value">{outlook_total/100000000:,.1f}억원</div><div class="metric-desc">전체 대출 · 실제 지급일 기준</div></div>', unsafe_allow_html=True)
                            with oc2:
                                st.markdown(f'<div class="metric-card"><div class="metric-label">월평균 예상이자</div><div class="metric-value">{outlook_avg_monthly/100000000:,.1f}억원</div><div class="metric-desc">{len(monthly_df)}개월 평균</div></div>', unsafe_allow_html=True)

                            outlook_view = st.radio(
                                "보기 방식", ["📅 월별", "🗓 연별", "🏦 기관별"], index=0,
                                key="interest_outlook_view", horizontal=True,
                            )

                            if outlook_view == "📅 월별":
                                view_df, x_col, x_title = monthly_df, '연월', '연월'
                            elif outlook_view == "🗓 연별":
                                view_df, x_col, x_title = yearly_df, '연도', '연도'
                            else:
                                view_df, x_col, x_title = by_bank_df, '금융기관', '금융기관'

                            fig_outlook = go.Figure()
                            fig_outlook.add_trace(go.Bar(
                                x=view_df[x_col],
                                y=view_df['예상이자합계'] / 100000000,
                                marker_color='#3b82f6',
                                text=[f"{v/100000000:,.1f}억원" for v in view_df['예상이자합계']],
                                textposition='outside',
                            ))
                            fig_outlook.update_layout(
                                plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
                                font=dict(color='#0f172a'),
                                yaxis=dict(gridcolor='rgba(15,23,42,0.06)'), xaxis=dict(title=x_title, type='category'),
                                margin=dict(t=70),
                                annotations=[dict(
                                    text='예상이자(억원)', xref='paper', yref='paper', x=0, y=1.1,
                                    xanchor='left', yanchor='bottom', showarrow=False, font=dict(size=12, color='#64748b')
                                )]
                            )
                            st.plotly_chart(fig_outlook, use_container_width=True)

                            view_display = view_df.copy()
                            view_display['예상이자합계'] = view_display['예상이자합계'].map(lambda v: f"{v:,.0f}")
                            st.dataframe(view_display, use_container_width=True, hide_index=True)

                    # ------------------------------------------------------------
                    # ① 금리 변경 예정일 체크
                    # ------------------------------------------------------------
                    with st.container(border=True):
                        st.markdown("#### 🔔 금리 변경 예정일 체크")
                        st.caption(
                            "'최근금리변경일'에 미래 날짜가 기재돼 있으면 그 값을 그대로 쓰고, 없으면 기표일(또는 차입일) + "
                            "변경주기로 다음 예정일을 추정합니다. 12개월 등 장기 주기는 대개 고정금리를 매년 갱신하는 방식이라 "
                            "'🔒 고정금리 갱신'으로 별도 표기하고, 변동금리 위험도(🔴/🟡) 집계에서는 제외했습니다."
                        )
                        rate_sched_df = build_rate_change_schedule(df_current, as_of_date)
                        if rate_sched_df.empty:
                            st.caption("금리변경/갱신 일정 정보가 있는 대출이 없습니다.")
                        else:
                            variable_df = rate_sched_df[rate_sched_df['금리유형'] == '변동금리']
                            n_urgent = int((variable_df['D-day'] <= 30).sum())
                            n_watch = int(((variable_df['D-day'] > 30) & (variable_df['D-day'] <= 90)).sum())
                            c1, c2, c3 = st.columns(3)
                            with c1:
                                st.markdown(f'<div class="metric-card"><div class="metric-label">🔴 30일 이내 임박</div><div class="metric-value">{n_urgent}건</div><div class="metric-desc">변동금리 변경 예정</div></div>', unsafe_allow_html=True)
                            with c2:
                                st.markdown(f'<div class="metric-card"><div class="metric-label">🟡 31~90일 관심</div><div class="metric-value">{n_watch}건</div><div class="metric-desc">사전 검토 권장</div></div>', unsafe_allow_html=True)
                            with c3:
                                if not variable_df.empty:
                                    nearest = variable_df.iloc[0]
                                    st.markdown(f'<div class="metric-card"><div class="metric-label">가장 임박한 변동금리 건</div><div class="metric-value">{nearest["금융기관"]}</div><div class="metric-desc">{nearest["다음 금리변경 예정일"]} (D{nearest["D-day"]:+d})</div></div>', unsafe_allow_html=True)
                                else:
                                    st.markdown('<div class="metric-card"><div class="metric-label">가장 임박한 변동금리 건</div><div class="metric-value">-</div><div class="metric-desc">해당 없음</div></div>', unsafe_allow_html=True)

                            rate_sched_column_config = {
                                "금융기관": st.column_config.Column(width="small"),
                                "대출과목": st.column_config.Column(width="small"),
                                "금리유형": st.column_config.Column(width="small"),
                                "현재금리": st.column_config.Column(width="small"),
                                "변경주기": st.column_config.Column(width="small"),
                                "다음 금리변경 예정일": st.column_config.Column(width="small"),
                                "D-day": st.column_config.Column(width="small"),
                                "상태": st.column_config.Column(width="small"),
                                "산출근거": st.column_config.Column(width="small"),
                            }
                            st.dataframe(rate_sched_df, use_container_width=True, hide_index=True, column_config=rate_sched_column_config)

                    # ------------------------------------------------------------
                    # ② 이자지급일 기준 예상이자 (신용카드 명세서처럼 지급일별 캘린더)
                    # ------------------------------------------------------------
                    with st.container(border=True):
                        st.markdown("#### 💰 예상 이자지급 캘린더")
                        cycle_options = {"다음 3회": 3, "다음 6회": 6, "다음 12회": 12}
                        cycle_label = st.radio(
                            "표시 범위 (대출별 다음 N회 지급주기)", options=list(cycle_options.keys()), index=1,
                            key="interest_n_cycles", horizontal=True,
                            help="대부분 매월이지만 일부(예: 3개월단위 회사채)는 분기 주기라, 대출마다 지급 간격이 달라 "
                                 "'N회'가 곧 N개월을 뜻하지는 않습니다.",
                        )
                        n_cycles = cycle_options[cycle_label]
                        st.caption(
                            "각 대출의 '이자지급일'(미기재 시 28일·매월 가정)을 기준으로 예상이자 = 잔액 × 금리 × 일수 / 365 를 "
                            "계산합니다. 이자지급일이 토·일요일이나 공휴일과 겹치면 신용카드 결제일처럼 다음 영업일에 실제로 "
                            "지급되므로, 이 화면은 그 '실제 지급일' 기준으로 날짜별 합계를 보여줍니다. 주기 중간에 금리변경 예정일이 "
                            "있으면 구간을 나눠 계산하고, 변경 이후 회차는 확정금리를 알 수 없어 현재금리로 잠정 산정합니다(비고에 표시)."
                        )
                        detail_df, summary_df = build_expected_interest(df_current, as_of_date, n_cycles=n_cycles)
                        if detail_df.empty:
                            st.caption("예상이자를 계산할 잔액·금리 데이터가 있는 대출이 없습니다.")
                        else:
                            total_expected = int(detail_df['예상이자(원)'].sum())
                            nearest_pay = summary_df.iloc[0]
                            c1, c2 = st.columns(2)
                            with c1:
                                st.markdown(f'<div class="metric-card"><div class="metric-label">{cycle_label} 예상이자 합계</div><div class="metric-value">{total_expected:,.0f}원</div><div class="metric-desc">전체 대출 · 실제 지급일 기준</div></div>', unsafe_allow_html=True)
                            with c2:
                                st.markdown(f'<div class="metric-card"><div class="metric-label">가장 가까운 실제 지급일</div><div class="metric-value">{nearest_pay["실제 지급일(영업일)"]}</div><div class="metric-desc">{nearest_pay["예상이자합계"]:,.0f}원 · {nearest_pay["대출건수"]}건</div></div>', unsafe_allow_html=True)

                            fig_interest = go.Figure()
                            fig_interest.add_trace(go.Bar(
                                x=summary_df['실제 지급일(영업일)'],
                                y=summary_df['예상이자합계'] / 100000000,
                                marker_color='#3b82f6',
                            ))
                            fig_interest.update_layout(
                                plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
                                font=dict(color='#0f172a'),
                                yaxis=dict(gridcolor='rgba(15,23,42,0.06)'), xaxis=dict(title='실제 지급일(영업일)'),
                                margin=dict(t=70),
                                annotations=[dict(
                                    text='예상이자(억원)', xref='paper', yref='paper', x=0, y=1.1,
                                    xanchor='left', yanchor='bottom', showarrow=False, font=dict(size=12, color='#64748b')
                                )]
                            )
                            st.plotly_chart(fig_interest, use_container_width=True)

                            st.markdown("##### 지급일별 합계 (실제 영업일 기준)")
                            summary_display = summary_df.copy()
                            summary_display['예상이자합계'] = summary_display['예상이자합계'].map(lambda v: f"{v:,.0f}")
                            st.dataframe(summary_display, use_container_width=True, hide_index=True)

                            st.markdown("##### 대출별 상세 내역")
                            detail_display = detail_df.copy()
                            detail_display['잔액(원)'] = detail_display['잔액(원)'].map(lambda v: f"{v:,.0f}")
                            detail_display['예상이자(원)'] = detail_display['예상이자(원)'].map(lambda v: f"{v:,.0f}")
                            detail_column_config = {
                                "금융기관": st.column_config.Column(width="small"),
                                "대출과목": st.column_config.Column(width="small"),
                                "회차": st.column_config.Column(width="small"),
                                "잔액(원)": st.column_config.Column(width="small"),
                                "적용금리": st.column_config.Column(width="small"),
                                "적용기간": st.column_config.Column(width="medium"),
                                "일수": st.column_config.Column(width="small"),
                                "이론상 지급일": st.column_config.Column(width="small"),
                                "실제 지급일(영업일)": st.column_config.Column(width="small"),
                                "예상이자(원)": st.column_config.Column(width="small"),
                                "비고": st.column_config.Column(width="large"),
                            }
                            st.dataframe(detail_display, use_container_width=True, hide_index=True, column_config=detail_column_config)

                    # ------------------------------------------------------------
                    # 🗓 과거 지급이자 실적 (재계산 없이 그대로) — "이자 이력 관리"에
                    # 업로드해둔 실제 지급 내역을 연도별로 보여줍니다. 위 ⓪~②는 전부
                    # 미래를 잔액×금리로 어림한 값인 반면, 이 표는 실제로 지급된 금액입니다.
                    # ------------------------------------------------------------
                    ih_years_view = interest_history_db.list_actual_interest_years()
                    if ih_years_view:
                        with st.container(border=True):
                            st.markdown("#### 🗓 과거 지급이자 실적")
                            st.caption(
                                "'💵 이자 이력 관리(관리자)'에 등록된 실제 지급 내역입니다 — 위의 예상치와 달리 "
                                "잔액×금리로 어림하지 않고, 업로드된 원본 그대로 보여줍니다. 이 중 확정 스케줄로 "
                                "등록된 구간은 위 ①~② 예상이자 계산에도 자동 반영됩니다."
                            )
                            ih_pick_year = st.selectbox("연도", ih_years_view, key="interest_history_view_year")
                            ih_rows_view = interest_history_db.list_actual_interest(ih_pick_year)
                            if ih_rows_view:
                                ih_view_df = pd.DataFrame(
                                    ih_rows_view,
                                    columns=["연도", "은행", "금액", "이율", "기간시작", "기간종료", "이자금액"],
                                )
                                ih_total = ih_view_df['이자금액'].sum()
                                st.markdown(
                                    f'<div class="metric-card"><div class="metric-label">{ih_pick_year}년 실제 이자 합계</div>'
                                    f'<div class="metric-value">{ih_total/100000000:,.1f}억원</div>'
                                    f'<div class="metric-desc">{len(ih_view_df)}건</div></div>',
                                    unsafe_allow_html=True,
                                )
                                ih_view_display = ih_view_df.copy()
                                ih_view_display['금액'] = ih_view_display['금액'].map(lambda v: f"{v:,.0f}" if pd.notna(v) else '-')
                                ih_view_display['이율'] = ih_view_display['이율'].map(lambda v: f"{v*100:.2f}%" if pd.notna(v) else '-')
                                ih_view_display['이자금액'] = ih_view_display['이자금액'].map(lambda v: f"{v:,.0f}")
                                st.dataframe(ih_view_display, use_container_width=True, hide_index=True)
            except Exception as e:
                st.error(f"원리금관리 계산 오류: {e}")
                st.caption("엑셀의 '★ 차입금 관리내역(현재)' 시트명/구조를 확인해주세요.")

        elif page == PAGE_MATURITY:
            st.subheader("📅 1년이내 만기도래 및 상환계획표")
            today = datetime.now().date()
            as_of_date = st.date_input(
                "조회일자(기준일) 선택",
                value=today,
                key="maturity_as_of_date",
                help="이 날짜를 기준으로 1년 이내 만기가 도래하는 차입금을 조회합니다.",
            )
            cutoff = (pd.Timestamp(as_of_date) + pd.DateOffset(years=1)).date()
            st.write(
                f"조회일자 **{as_of_date.strftime('%Y-%m-%d')}** 기준으로 "
                f"**{cutoff.strftime('%Y-%m-%d')}**까지 만기가 도래하는 차입금 내역입니다. "
                "(단위: 백만원)"
            )
            try:
                df_current, sheet_used = load_current_loan_status(excel_bytes)
                if df_current.empty:
                    st.info("차입금현황 데이터가 없어 만기도래표를 만들 수 없습니다.")
                else:
                    schedule_df, schedule_keys = build_maturity_schedule(df_current, as_of_date=as_of_date)

                    if schedule_df.empty:
                        st.success("✅ 조회일자 기준 1년 이내 만기가 도래하는 차입금이 없습니다.")
                    else:
                        saved_status = load_repayment_status()

                        # 저장되어 있던 상환여부/비고 값을 반영
                        for i, key in enumerate(schedule_keys):
                            if key in saved_status:
                                schedule_df.at[i, '상환여부'] = saved_status[key].get('상환여부', '')
                                schedule_df.at[i, '비고'] = saved_status[key].get('비고', '')

                        total_amount_mm = schedule_df.loc[schedule_df['차입기관명'] == '계', '금액(백만원)'].iloc[0]
                        n_items = len(schedule_df) - 1  # 합계 행 제외
                        col_a, col_b = st.columns(2)
                        with col_a:
                            st.markdown(f'<div class="metric-card"><div class="metric-label">1년 이내 만기도래 건수</div><div class="metric-value">{n_items}건</div><div class="metric-desc">기준일: {as_of_date.strftime("%Y-%m-%d")}</div></div>', unsafe_allow_html=True)
                        with col_b:
                            st.markdown(f'<div class="metric-card"><div class="metric-label">만기도래 예정 총액</div><div class="metric-value">{total_amount_mm:,.0f} 백만원</div><div class="metric-desc">잔액 기준(잔액이 없으면 한도액)</div></div>', unsafe_allow_html=True)
                        st.markdown("---")

                        # 표시용 포맷 (금액에 콤마)
                        display_df = schedule_df.copy()
                        display_df['금액(백만원)'] = display_df['금액(백만원)'].map(
                            lambda v: f"{v:,.0f}" if pd.notna(v) else "-"
                        )

                        edited_df = st.data_editor(
                            display_df,
                            use_container_width=True,
                            hide_index=True,
                            key="maturity_schedule_editor",
                            disabled=['차입기관명', '과목', '금액(백만원)', '만기일', '담보종류'],
                            column_config={
                                "상환여부": st.column_config.TextColumn(
                                    "상환여부", help="상환 또는 연장으로 기재 (일부상환의 경우 상환예정금액 기재)"
                                ),
                                "비고": st.column_config.TextColumn("비고"),
                            },
                        )

                        # 편집된 상환여부/비고를 로컬에 저장 (합계 행 제외)
                        updated = False
                        for i, key in enumerate(schedule_keys):
                            if key == '__TOTAL__':
                                continue
                            new_status = edited_df.at[i, '상환여부']
                            new_note = edited_df.at[i, '비고']
                            prev = saved_status.get(key, {})
                            if prev.get('상환여부', '') != new_status or prev.get('비고', '') != new_note:
                                saved_status[key] = {'상환여부': new_status, '비고': new_note}
                                updated = True
                        if updated:
                            save_repayment_status(saved_status)

                        st.caption(
                            "가) 상환여부: 상환 또는 연장으로 기재(일부상환의 경우 상환예정금액 기재)  \n"
                            "나) 유동성장기부채의 상환계획도 표시  \n"
                            "다) 동일 차입기관에 2개 이상 대출이 있는 경우 만기일 순으로 구분 기재"
                        )
            except Exception as e:
                st.error(f"만기도래표 생성 오류: {e}")
                st.caption("엑셀의 '★ 차입금 관리내역(현재)' 시트에서 만기일/잔액 컬럼을 확인해주세요.")

        # Load credit ratings downgrades
        credit_downgrades = check_credit_downgrade(excel_bytes)
 
        # 최근 3개년 평가 로직: 두 가지 연도 기준을 선택할 수 있게 제공
        #  - 당해년도 포함 기준 (예: 24~26년): 26년 최근 금리를 반영하고, 25년 대비 26년 스프레드/등급
        #    변동을 보기 위한 기본값
        #  - 완전 연도 기준 (예: 23~25년): 당해년도 데이터가 아직 전혀 없을 때 참고용 대체 옵션
        window_incl_curr = [f"{y}년" for y in range(CURRENT_YEAR - 2, CURRENT_YEAR + 1)]
        window_incl_curr = [y for y in window_incl_curr if y in YEAR_COLS]
        window_complete = [f"{y}년" for y in range(CURRENT_YEAR - 3, CURRENT_YEAR)]
        window_complete = [y for y in window_complete if y in YEAR_COLS]

        year_window_options = {}
        if len(window_incl_curr) == 3:
            year_window_options[f"{window_incl_curr[0]}~{window_incl_curr[-1]} (당해년도 포함, 기본)"] = window_incl_curr
        if len(window_complete) == 3 and window_complete != window_incl_curr:
            year_window_options[f"{window_complete[0]}~{window_complete[-1]} (완전 연도만)"] = window_complete
        if not year_window_options:
            year_window_options["최근 3개년"] = YEAR_COLS[-3:]

        default_window_label = list(year_window_options.keys())[0]
        if page == PAGE_ANCILLARY:
            selected_window_label = st.radio(
                "평가 기준 연도 선택",
                options=list(year_window_options.keys()),
                index=0,
                horizontal=True,
                key="negotiation_eval_year_window"
            )
        else:
            selected_window_label = st.session_state.get("negotiation_eval_year_window", default_window_label)
            if selected_window_label not in year_window_options:
                selected_window_label = default_window_label
        eval_years = year_window_options[selected_window_label]
        y_a, y_b, y_c = eval_years[-3], eval_years[-2], eval_years[-1]  # 과거 -> 전년 -> 최근(당해)

        if page == PAGE_ANCILLARY:
            st.markdown(f"### ✅ 금융기관 신용스프레드 기반 금리 협상력 평가 (3개년 기준: {y_a}~{y_c})")
            st.caption(
                "💡 **1차 지표(기준금리 대비)**: 스프레드 = 대출금리 − 한국은행 기준금리. "
                "신용평가사가 등급별 회사채 스프레드를 공시하는 방식과 동일한 원리입니다.  \n"
                "**2차 지표(실제 벤치마크 대비)**: 당사가 실제로 선택한 벤치마크(코리보·CD·금융채 등, "
                "1년 고정물/3개월 변동물 등 만기·종류가 저마다 다름)의 그 기간 자체 변동분을 ECOS에서 가져와 비교합니다. "
                "한은 기준금리는 안 움직였는데 벤치마크가 움직였다면, 스프레드 확대가 협상 실패가 아니라 "
                "'그 시점에 그 벤치마크가 더 비싸게 움직인 결과'일 수 있어 이 구분이 중요합니다.  \n"
                "신용등급 변동 방향도 함께 반영합니다. (평가 대상 기간은 위에서 선택한 3개년만 사용합니다.)"
            )

        bok_vals = [bok_rates.get(str(int(y.replace('년', ''))), None) for y in eval_years]
        bok_diff_curr = None
        bok_diff_prev = None
        bok_a = bok_b = bok_c = None
        if len(bok_vals) >= 3 and all(v is not None for v in bok_vals[-3:]):
            bok_a, bok_b, bok_c = bok_vals[-3], bok_vals[-2], bok_vals[-1]
            bok_diff_curr = bok_c - bok_b   # 최근년 vs 전년
            bok_diff_prev = bok_b - bok_a   # 전년 vs 그 이전년

        # 실제 사용된 벤치마크별 ECOS 시계열은 처음 필요해질 때만 조회해서 캐시(중복 API 호출 방지)
        bench_series_cache = {}
        bench_name_cache = {}

        def get_bench_value(bench_key, year_label):
            """bench_key(코리보/CD/금융채 등)의 특정 연도(예: '2024년') ECOS 값을 반환. 없으면 None."""
            if bench_key not in bench_series_cache:
                series, item_name = fetch_ecos_benchmark_series(bench_key)
                bench_series_cache[bench_key] = series
                bench_name_cache[bench_key] = item_name
            series = bench_series_cache[bench_key]
            return series.get(str(int(year_label.replace('년', ''))), None)

        eval_data = []
        pending_data = []       # 최근(당해)년도 금리 미반영 = 연말 연장 대기 중인 건
        insufficient_count = 0  # 과거 이력 자체가 부족해 평가 불가한 건

        if bok_a is not None:
            for _, row in df_hist.iterrows():
                if any(y not in df_hist.columns for y in eval_years):
                    continue

                val_a, val_b, val_c = row[y_a], row[y_b], row[y_c]

                if val_a == 0 or val_b == 0:
                    # 과거 이력 자체가 없어 스프레드 추세를 계산할 수 없는 건
                    insufficient_count += 1
                    continue

                if val_c == 0:
                    # 최근년도({y_c}) 금리가 아직 미반영 = 연말 재협상/연장 전 상태.
                    # 평가에서 제외하지 않고, 별도의 '연장 대기' 목록으로 보여준다.
                    spread_b_pending = val_b - bok_b
                    pending_data.append({
                        "금융기관": row['금융기관'],
                        f"{y_b} 금리": f"{val_b:.2f}%",
                        f"스프레드({y_b})": f"{spread_b_pending:+.2f}%p",
                        "상태": "🕒 연장 대기",
                        "비고": f"{y_c} 금리 갱신(연장) 완료 후 자동으로 평가 대상에 반영됩니다."
                    })
                    continue

                rate_a, rate_b, rate_c = val_a, val_b, val_c

                # --- 1차 지표: 스프레드(가산금리) = 대출금리 - 한국은행 기준금리 ---
                spread_a = rate_a - bok_a
                spread_b = rate_b - bok_b
                spread_c = rate_c - bok_c
                d_spread_prev = spread_b - spread_a   # 전년도 스프레드 변동
                d_spread_curr = spread_c - spread_b   # 최근년도 스프레드 변동 (기준금리 기준)

                # --- 2차 지표: 실제 선택 벤치마크(코리보/CD/금융채 등) 대비 실질 스프레드 ---
                cond_b = row.get(f"{y_b}_조건", "")
                cond_c = row.get(f"{y_c}_조건", "")
                bench_key = parse_benchmark_key(cond_c) or parse_benchmark_key(cond_b)

                bench_b_val = get_bench_value(bench_key, y_b) if bench_key else None
                bench_c_val = get_bench_value(bench_key, y_c) if bench_key else None

                d_spread_bench_curr = None
                bench_display_name = None
                if bench_key and bench_b_val is not None and bench_c_val is not None:
                    bench_display_name = bench_name_cache.get(bench_key) or bench_key
                    spread_bench_b = rate_b - bench_b_val
                    spread_bench_c = rate_c - bench_c_val
                    d_spread_bench_curr = spread_bench_c - spread_bench_b

                # 평가에 쓸 '핵심 스프레드 변동' — 벤치마크 매칭이 되면 그쪽을 우선 사용(더 정밀),
                # 안 되면 기존처럼 기준금리 기준으로 평가(하위 호환)
                d_spread_key = d_spread_bench_curr if d_spread_bench_curr is not None else d_spread_curr

                # Map bank name to normalized name
                norm_name = get_normalized_bank_name(row['금융기관'])

                # Determine credit rating and change text/direction
                credit_change = str(row['신용등급변동'])
                rating_24 = "-"
                rating_25 = "-"
                is_downgraded = False
                is_upgraded = any(tok in credit_change for tok in ("상향", "상승"))

                if norm_name and norm_name in credit_downgrades:
                    info = credit_downgrades[norm_name]
                    rating_24 = info["r_24"]
                    rating_25 = info["r_25"]
                    is_downgraded = info["is_downgrade"]

                    if is_downgraded:
                        if "하락" not in credit_change:
                            credit_change = f"등급 하락 ({rating_24} → {rating_25})"
                        is_upgraded = False
                    elif credit_change == "-" and rating_25 != "-":
                        credit_change = f"안정 ({rating_25})"

                rating_dir = "downgrade" if is_downgraded else ("upgrade" if is_upgraded else "stable")

                # --- 등급 변동 방향 x 스프레드 변동(d_spread_key)으로 4분면 평가 ---
                basis_txt = f"{bench_display_name} 대비" if d_spread_bench_curr is not None else "기준금리 대비(벤치마크 미매칭)"
                if rating_dir == "downgrade":
                    if d_spread_key <= 0.05:
                        status, eval_txt = "🟢 우수", f"신용등급 하락 악재에도 스프레드를 방어/축소 (협상력 우수, {basis_txt})"
                    elif d_spread_key <= 0.30:
                        status, eval_txt = "🟡 주의", f"등급 하락을 감안하면 스프레드 확대가 시장 관행상 불가피한 수준 ({basis_txt})"
                    else:
                        status, eval_txt = "🔴 경고", f"등급 하락폭 대비 스프레드가 과도하게 확대됨 (재협상 필요, {basis_txt})"
                elif rating_dir == "upgrade":
                    if d_spread_key <= -0.05:
                        status, eval_txt = "🟢 우수", f"등급 상승분이 스프레드 축소로 충분히 반영됨 ({basis_txt})"
                    elif d_spread_key <= 0.05:
                        status, eval_txt = "🟡 주의", f"등급이 개선됐음에도 스프레드 개선폭이 미미함 (재협상 여지, {basis_txt})"
                    else:
                        status, eval_txt = "🔴 경고", f"등급 상승에도 스프레드가 오히려 확대됨 (협상 미흡, {basis_txt})"
                else:  # stable
                    if d_spread_key <= -0.05:
                        status, eval_txt = "🟢 우수", f"등급 변동 없이 스프레드 축소에 성공 ({basis_txt})"
                    elif d_spread_key <= 0.05:
                        status, eval_txt = "🟢 양호", f"스프레드는 유지, 금리 변동은 {basis_txt.replace(' 대비','')} 변동분만 반영됨"
                    else:
                        status, eval_txt = "🔴 경고", f"등급 변동이 없는데도 스프레드가 확대됨 (협상 필요, {basis_txt})"

                # 과거(전년도) 스프레드가 이미 과도하게 벌어졌던 경우, 올해 개선이 회복 수준일 수 있음을 참고 표기
                if status == "🟢 우수" and d_spread_prev > 0.30:
                    eval_txt += " · 단, 전년도 확대분 회복 수준이라 지속 모니터링 필요"

                # 한도대(미사용 한도 포함)는 실제로 인출해서 쓰기 전까지는 그 금리가 당장
                # 이자비용으로 나가지 않는다 — 그래서 스프레드가 넓어졌어도 일반대ㆍ회사채처럼
                # "협상 미흡"으로 단정하기는 애매하고, 사용량이 늘어날 때 비로소 문제가 되는
                # 성격이다(사용자 피드백). 반대로 스프레드를 줄여놓은 건 인출 대비 안전판을
                # 미리 확보한 것이니 그대로 잘한 것으로 인정한다 — 그래서 축소(🟢)는 건드리지
                # 않고, 확대(🟡/🔴)만 "⚪ 보통"으로 한 단계 완화한다.
                is_undrawn_limit = '한도대' in str(row['금융기관'])
                if is_undrawn_limit and status in ("🟡 주의", "🔴 경고"):
                    status = "⚪ 보통"
                    eval_txt += " · 한도대(미사용 한도 포함)라 당장 인출해 쓰기 전까지는 이자비용에 영향이 없어 보통으로 완화 평가"

                # '잘한 것' / '보통' / '못한 것' 세 그룹으로 나누기 위한 판정 (🟢 계열 = 잘한 것)
                if "🟢" in status:
                    group = "잘한 것"
                elif status == "⚪ 보통":
                    group = "보통"
                else:
                    group = "못한 것"

                eval_data.append({
                    "금융기관": row['금융기관'],
                    f"{y_b} 금리": f"{rate_b:.2f}%",
                    f"{y_c} 금리(최근)": f"{rate_c:.2f}%",
                    f"스프레드({y_b},기준금리대비)": f"{spread_b:+.2f}%p",
                    f"스프레드({y_c},기준금리대비)": f"{spread_c:+.2f}%p",
                    "Δ스프레드(기준금리대비,%p)": f"{d_spread_curr:+.2f}",
                    # bench_key는 "금리조건 텍스트에서 이 종류로 읽었다"는 뜻일 뿐, ECOS에
                    # 실제로 그 항목이 없거나(예: 이 통계표엔 '금융채' 자체가 없음) 데이터를
                    # 못 가져오면 bench_display_name이 끝까지 None으로 남는다. 예전엔 이 경우도
                    # bench_key 텍스트를 그대로 보여줘서 "매칭된 것처럼" 보였는데, 실제로는
                    # Δ스프레드(벤치마크대비)가 항상 '-'인 미매칭 건이라 화면과 실제 계산이
                    # 어긋나 보였다(실제로 겪은 문제). 이제는 진짜 매칭(=값까지 확보)된 경우만
                    # 벤치마크명을 보여주고, 나머지는 어떤 이유든 전부 "미인식(기준금리로 대체)"로
                    # 통일해서 Δ스프레드 컬럼의 '-'와 항상 일치하게 한다.
                    "선택 벤치마크": bench_display_name if bench_display_name else "미인식(기준금리로 대체)",
                    "Δ스프레드(벤치마크대비,%p)": f"{d_spread_bench_curr:+.2f}" if d_spread_bench_curr is not None else "-",
                    "_핵심Δ스프레드": d_spread_key,  # 요약 지표 계산용(숫자, 표시 안 함)
                    "신용등급 변동": credit_change,
                    "평가 사유": eval_txt,
                    "종합 상태": status,
                    "구분": group
                })

        eval_df = pd.DataFrame()
        if eval_data:
            eval_df = pd.DataFrame(eval_data)

        if page == PAGE_ANCILLARY:
            tab_ancillary_spread, tab_ancillary_link = st.tabs(["📊 금리협상력 평가", "🔗 부수거래(예금·카드·퇴직연금) 연계분석"])
            with tab_ancillary_spread:
                if not eval_df.empty:
                    # 요약 지표: 스프레드 방어(축소·유지) 비율
                    n_total = len(eval_df)
                    n_good = int((eval_df["구분"] == "잘한 것").sum())
                    n_neutral = int((eval_df["구분"] == "보통").sum())
                    n_bad = n_total - n_good - n_neutral
                    avg_d_spread = eval_df["_핵심Δ스프레드"].astype(float).mean() if n_total else 0.0
                    n_bench_matched = int((eval_df["선택 벤치마크"] != "미인식(기준금리로 대체)").sum())

                    s1, s2, s3 = st.columns(3)
                    with s1:
                        neutral_desc = f" · 보통(한도대) {n_neutral}개 별도" if n_neutral else ""
                        st.markdown(f'<div class="metric-card"><div class="metric-label">스프레드 방어/축소 기관</div><div class="metric-value">{n_good}/{n_total}개</div><div class="metric-desc positive">전체 평가 대상 중 비중{neutral_desc}</div></div>', unsafe_allow_html=True)
                    with s2:
                        cls = "positive" if avg_d_spread <= 0 else "negative"
                        st.markdown(f'<div class="metric-card"><div class="metric-label">평균 Δ스프레드(핵심지표)</div><div class="metric-value">{avg_d_spread:+.2f}%p</div><div class="metric-desc {cls}">가능하면 벤치마크 대비, 없으면 기준금리 대비</div></div>', unsafe_allow_html=True)
                    with s3:
                        st.markdown(f'<div class="metric-card"><div class="metric-label">벤치마크 매칭 건수</div><div class="metric-value">{n_bench_matched}/{n_total}개</div><div class="metric-desc">ECOS 벤치마크로 정밀 평가된 비중</div></div>', unsafe_allow_html=True)

                    with st.expander("ℹ️ '벤치마크 매칭 건수'가 무슨 뜻인가요? — 대출별 매칭 상세 보기"):
                        st.markdown(
                            "대출마다 '세부 금리조건' 텍스트(예: \"KORIBOR + 1.60%\", \"3개월 CD변동 + 1.69%\")에서 "
                            "벤치마크 종류를 자동으로 인식해, 한국은행 ECOS 시장금리 통계표(817Y002)에서 그 벤치마크의 "
                            "실제 연도별 시계열을 찾아 매칭합니다. 매칭에 성공하면 스프레드 변동을 "
                            "**\"그 대출이 실제로 걸려있는 지표금리\" 대비**로 정밀 평가하고(2차 지표), 매칭에 실패하면 "
                            "(신규/생소한 벤치마크, ECOS 미제공 종목 등) 대신 **한국은행 기준금리 대비**로 대체 평가합니다(1차 지표) — "
                            "그래서 매칭 건수가 많을수록 협상 성과 평가가 더 정밀하다는 뜻입니다."
                        )
                        match_detail_df = eval_df[["금융기관", "선택 벤치마크", "Δ스프레드(벤치마크대비,%p)"]].copy()
                        match_detail_df["매칭 여부"] = match_detail_df["선택 벤치마크"].apply(lambda v: "✅ 매칭" if v != "미인식(기준금리로 대체)" else "⚠️ 미인식(기준금리로 대체)")
                        st.dataframe(match_detail_df, use_container_width=True, hide_index=True)

                    st.markdown("---")

                    good_df = eval_df[eval_df["구분"] == "잘한 것"].drop(columns=["구분", "_핵심Δ스프레드"]).reset_index(drop=True)
                    neutral_df = eval_df[eval_df["구분"] == "보통"].drop(columns=["구분", "_핵심Δ스프레드"]).reset_index(drop=True)
                    bad_df = eval_df[eval_df["구분"] == "못한 것"].drop(columns=["구분", "_핵심Δ스프레드"]).reset_index(drop=True)

                    # 컬럼 수가 많아 화면 폭을 넘기기 쉬우므로, 짧은 항목은 좁게·서술형 항목은
                    # 넓게 지정해 표가 잘리지 않고 한 화면에 정렬되도록 한다.
                    eval_column_config = {
                        "금융기관": st.column_config.Column(width="small"),
                        f"{y_b} 금리": st.column_config.Column(width="small"),
                        f"{y_c} 금리(최근)": st.column_config.Column(width="small"),
                        "선택 벤치마크": st.column_config.Column(width="small"),
                        "신용등급 변동": st.column_config.Column(width="small"),
                        "종합 상태": st.column_config.Column(width="small"),
                        "평가 사유": st.column_config.Column(width="large"),
                    }

                    st.markdown(f"#### ✅ 잘한 것 (스프레드 방어·축소) — {n_good}개 기관")
                    if not good_df.empty:
                        st.dataframe(good_df, use_container_width=True, hide_index=True, column_config=eval_column_config)
                    else:
                        st.caption("해당 기관이 없습니다.")

                    if n_neutral:
                        st.markdown(f"#### ⚪ 보통 (한도대 — 미사용 한도라 스프레드 확대해도 당장 비용 영향 적음) — {n_neutral}개 기관")
                        st.caption("한도대(미사용 한도 포함)는 실제로 인출하기 전까지 이자비용이 발생하지 않아, 스프레드가 넓어졌어도 일반대·회사채처럼 '협상 미흡'으로 단정하지 않고 별도로 분리했습니다. 스프레드를 줄여놓은 한도대는 그대로 '잘한 것'에 포함됩니다.")
                        st.dataframe(neutral_df, use_container_width=True, hide_index=True, column_config=eval_column_config)

                    st.markdown(f"#### ⚠️ 못한 것 (스프레드 확대·협상 미흡) — {n_bad}개 기관")
                    if not bad_df.empty:
                        st.dataframe(bad_df, use_container_width=True, hide_index=True, column_config=eval_column_config)
                    else:
                        st.caption("해당 기관이 없습니다.")
                else:
                    st.info("평가 대상 연도의 금리 데이터가 충분하지 않습니다 (선택한 3개년 모두 0 초과 데이터가 필요).")

                # 연말 연장 대기 중인 건 (최근/당해년도 금리 미반영) — 사라지지 않고 별도로 표시
                if pending_data:
                    st.markdown("---")
                    st.markdown(f"#### 🕒 연장 대기 중 (당해년도 미반영) — {len(pending_data)}건")
                    st.caption(f"아직 {y_c} 갱신(연장)이 완료되지 않아 이번 평가에서는 제외했습니다. 연말에 연장이 완료되면 자동으로 위 평가표에 반영됩니다.")
                    st.dataframe(pd.DataFrame(pending_data), use_container_width=True, hide_index=True)

                if insufficient_count:
                    st.caption(f"ℹ️ 과거 이력 데이터 부족으로 평가 대상에서 제외된 건: {insufficient_count}건")

                # 평가 기준표
                st.markdown(f"""
                #### 📋 스프레드 기반 평가 기준표 (신용등급 변동 + 실제 벤치마크 반영)
                **핵심 스프레드 = 대출금리 − 비교금리.** 비교금리는 가능하면 당사가 실제로 계약한 벤치마크(코리보·CD·금융채 등, ECOS에서 자동 조회)를 쓰고,
                벤치마크를 텍스트에서 인식하지 못했거나 ECOS 값이 없을 때만 한국은행 기준금리로 대체합니다.
                이 스프레드의 **당해년도({y_c}) 변동폭**을 핵심 지표로 삼고, 신용등급 변동 방향(하락/유지/상승)에 따라 같은 스프레드 확대라도 다르게 평가합니다.

                | 상태 | 등급 하락 시 | 등급 유지 시 | 등급 상승 시 |
                |---|---|---|---|
                | **🟢 우수** | 스프레드 방어/축소 (+0.05%p 이하) | 스프레드 축소 (-0.05%p 이하) | 스프레드 축소 (-0.05%p 이하) |
                | **🟢 양호** | - | 스프레드 유지 (±0.05%p 이내) | - |
                | **🟡 주의** | 스프레드 소폭 확대 (~+0.30%p, 등급하락 감안 시 불가피) | - | 스프레드 개선폭 미미 (±0.05%p 이내) |
                | **🔴 경고** | 스프레드 과도 확대 (+0.30%p 초과) | 스프레드 확대 (+0.05%p 초과) | 스프레드 오히려 확대 |

                > 참고: 국내 신용평가사들은 등급별 회사채 스프레드 통계를 정기 공시하며, 신용등급이 하락한 채권/여신은 시장 스프레드가 확대되는 것이 일반적입니다.
                > 따라서 등급 하락이 있었던 해의 스프레드 확대를 무조건 '협상 실패'로 보지 않고, 확대폭이 시장 관행 대비 과도한지를 기준으로 평가합니다.
                > **한도대 예외**: 위 표에서 🟡/🔴에 해당하더라도 한도대(미사용 한도 포함)는 실제 인출 전까지 이자비용이 없어 "⚪ 보통"으로 한 단계 완화합니다 — 단, 스프레드를 줄여놓은 경우(🟢)는 한도대여도 그대로 잘한 것으로 인정합니다.
                > **벤치마크 자동 매칭에 대해**: 대출조건 텍스트(예: "KORIBOR + 1.60%", "금융채 1년 + 1.90%")에서 벤치마크 종류를 인식해 ECOS 시장금리(817Y002) 통계표에서
                > 항목명을 검색·매칭합니다. 텍스트에서 벤치마크 종류를 읽었더라도 ECOS에 그 항목 자체가 없거나 값을 못 가져오면(신규/생소한 벤치마크, ECOS 미제공 종목 등)
                > 실제로 매칭된 것으로 치지 않고 "선택 벤치마크"란에 "미인식(기준금리로 대체)"로 표시하고 기준금리 기준으로 자동 대체됩니다 — 값을 실제로 확보한 경우만 매칭으로 인정합니다.
                """)

            with tab_ancillary_link:
                # ---------------------------------------------------------------
                # 예금 기여도 · 법인카드 사용액 ↔ 대출금리 스프레드 성과 연계 분석
                # ---------------------------------------------------------------
                st.markdown("---")
                st.markdown("### 💰 예금ㆍ법인카드ㆍ퇴직연금 거래 기여도 ↔ 대출금리 스프레드 연계 분석")
                st.caption(
                    "당사가 각 금융기관에 예치한 예금 잔액ㆍ법인카드 이용 실적ㆍ퇴직연금 적립금(별도 프로그램인 "
                    "퇴직연금통합관리시스템 데이터를 그대로 연동)을 '거래 관계 규모'로 환산하여, "
                    "위에서 평가한 대출금리 스프레드 성과와 같은 기관 기준으로 나란히 비교합니다.  \n"
                    "예금·카드·퇴직연금 등 우량 거래 관계를 유지해온 기관일수록 대출금리 협상에서도 유리한 결과"
                    "(스프레드 방어·축소)로 이어졌는지를 확인하기 위한 자료입니다."
                )

                # 세션에 아직 없으면(=새로 접속했거나 서버가 재시작됐으면) DB에 저장해둔
                # 마지막 업로드본을 먼저 불러온다 — 예전엔 st.session_state에만 올려둬서
                # 재시작하거나 다른 사람이 접속하면 업로드한 파일이 사라졌었다.
                if 'deposit_bytes' not in st.session_state:
                    saved = ancillary_db.get_file('deposit')
                    if saved:
                        st.session_state['deposit_bytes'] = saved[2]
                        st.session_state['deposit_filename'] = saved[0]
                        st.session_state['deposit_uploaded_at'] = saved[1]
                if 'card_bytes' not in st.session_state:
                    saved = ancillary_db.get_file('card')
                    if saved:
                        st.session_state['card_bytes'] = saved[2]
                        st.session_state['card_filename'] = saved[0]
                        st.session_state['card_uploaded_at'] = saved[1]

                dep_col, card_col = st.columns(2)
                with dep_col:
                    deposit_file = st.file_uploader(
                        "예금 기여도 분석 엑셀 (.xlsx)",
                        type=["xlsx"], key="deposit_uploader",
                        help="예: 금융기관별_예금_기여도_및_수익률_분석_양식.xlsx"
                    )
                    if deposit_file is not None:
                        st.session_state['deposit_bytes'] = deposit_file.getvalue()
                        st.session_state['deposit_filename'] = deposit_file.name
                        st.session_state['deposit_uploaded_at'] = datetime.now().strftime("%Y-%m-%d %H:%M")
                        ancillary_db.save_file('deposit', deposit_file.name, deposit_file.getvalue())
                    if 'deposit_filename' in st.session_state:
                        st.caption(f"✅ 등록됨: {st.session_state['deposit_filename']} ({st.session_state['deposit_uploaded_at']})")
                with card_col:
                    card_file = st.file_uploader(
                        "법인카드 사용금액 엑셀 (.xlsx)",
                        type=["xlsx"], key="card_uploader",
                        help="예: 법인카드_사용_금액.xlsx"
                    )
                    if card_file is not None:
                        st.session_state['card_bytes'] = card_file.getvalue()
                        st.session_state['card_filename'] = card_file.name
                        st.session_state['card_uploaded_at'] = datetime.now().strftime("%Y-%m-%d %H:%M")
                        ancillary_db.save_file('card', card_file.name, card_file.getvalue())
                    if 'card_filename' in st.session_state:
                        st.caption(f"✅ 등록됨: {st.session_state['card_filename']} ({st.session_state['card_uploaded_at']})")

                pension_base_date, pension_amounts_raw = fetch_pension_institution_amounts()

                if 'deposit_bytes' not in st.session_state and 'card_bytes' not in st.session_state and not pension_amounts_raw:
                    st.info("👆 예금 기여도 분석 파일과 법인카드 사용금액 파일을 업로드하면, 거래 관계 규모와 대출금리 개선 효과를 연계한 분석이 표시됩니다. 한 번 등록해두면 서버를 재시작하거나 다른 사람이 접속해도 계속 남아있습니다. (퇴직연금 적립금은 퇴직연금통합관리시스템에서 자동으로 연동되어 별도 업로드가 필요 없습니다.)")
                else:
                    deposit_by_year = load_deposit_contribution(st.session_state['deposit_bytes']) if 'deposit_bytes' in st.session_state else {}
                    card_df = load_card_usage(st.session_state['card_bytes']) if 'card_bytes' in st.session_state else pd.DataFrame()

                    # ---- 차입금 현황이 있는 기관만 남기고 나머지는 제외 ----
                    # (예금·카드 거래는 있지만 당사와 차입 관계가 없는 기관은 대출금리 협상과
                    #  무관하므로, 이 화면의 모든 표·차트에서 애초에 제외한다.)
                    active_loan_banks = get_active_loan_bank_names(excel_bytes)
                    excluded_no_loan_names = set()

                    if active_loan_banks:
                        if deposit_by_year:
                            filtered_by_year = {}
                            for yr, ydf in deposit_by_year.items():
                                keep_mask = ydf['금융기관'].map(lambda v: normalize_relation_name(v) in active_loan_banks)
                                excluded_no_loan_names.update(ydf.loc[~keep_mask, '금융기관'].astype(str).unique())
                                filtered_by_year[yr] = ydf[keep_mask].reset_index(drop=True)
                            deposit_by_year = filtered_by_year
                        if not card_df.empty:
                            keep_mask = card_df['은행'].map(lambda v: normalize_relation_name(v) in active_loan_banks)
                            excluded_no_loan_names.update(card_df.loc[~keep_mask, '은행'].astype(str).unique())
                            card_df = card_df[keep_mask].reset_index(drop=True)
                    else:
                        st.warning(
                            "⚠️ '★ 차입금 관리내역(현재)' 시트에서 차입금 현황을 확인할 수 없어, 예금·카드 목록에 "
                            "차입금 거래 기관 필터를 적용하지 못했습니다. 시트 구성을 확인해주세요."
                        )

                    # ---- 예금 기여도 (차입금 거래 기관만) ----
                    dep_latest = pd.DataFrame()
                    latest_dep_year = None
                    if deposit_by_year:
                        latest_dep_year = max(deposit_by_year.keys())
                        dep_years_sorted = sorted(deposit_by_year.keys())
                        dep_latest = deposit_by_year[latest_dep_year].copy()

                        st.markdown(f"#### 🏦 예금 기여도 현황 ({latest_dep_year}년 기준, 차입금 거래 기관만)")
                        if dep_latest.empty:
                            st.caption("차입금 거래가 있는 기관 중에서는 예금 기여도 데이터와 매칭되는 항목이 없습니다.")
                        else:
                            dep_latest_display = dep_latest.copy()
                            dep_latest_display['잔액'] = dep_latest_display['잔액'].map(lambda v: f"{v/100000000:,.1f}억원")
                            dep_latest_display['비중'] = dep_latest_display['비중'].map(lambda v: f"{v*100:,.1f}%")
                            dep_latest_display['금리'] = dep_latest_display['금리'].map(lambda v: f"{v*100:,.2f}%")
                            st.dataframe(dep_latest_display, use_container_width=True, hide_index=True)

                            trend_rows = []
                            for yr in dep_years_sorted:
                                for _, r in deposit_by_year[yr].iterrows():
                                    trend_rows.append({'연도': str(yr), '금융기관': r['금융기관'], '잔액': r['잔액']})
                            if trend_rows:
                                trend_df = pd.DataFrame(trend_rows)
                                fig_dep = go.Figure()
                                for bank in trend_df['금융기관'].unique():
                                    sub = trend_df[trend_df['금융기관'] == bank]
                                    fig_dep.add_trace(go.Bar(x=sub['연도'], y=sub['잔액'] / 100000000, name=bank))
                                fig_dep.update_layout(
                                    barmode='stack', plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
                                    font=dict(color='#0f172a'), yaxis=dict(gridcolor='rgba(15,23,42,0.06)'),
                                    xaxis=dict(title='연도'), legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
                                    margin=dict(t=70),
                                    annotations=[dict(
                                        text='잔액(억원)', xref='paper', yref='paper', x=0, y=1.1,
                                        xanchor='left', yanchor='bottom', showarrow=False, font=dict(size=12, color='#64748b')
                                    )]
                                )
                                st.plotly_chart(fig_dep, use_container_width=True)
                    else:
                        st.caption("예금 기여도 파일이 업로드되지 않았습니다.")

                    # ---- 법인카드 사용액 (최근 3개년, 차입금 거래 기관만) ----
                    card_latest = pd.DataFrame()
                    if not card_df.empty:
                        card_df = card_df.copy()
                        card_df['_연도숫자'] = card_df['연도'].astype(str).str.extract(r'(\d{4})').astype(float)
                        recent_years = sorted(card_df['_연도숫자'].dropna().unique())[-3:]
                        card_df_recent = card_df[card_df['_연도숫자'].isin(recent_years)]

                        if card_df_recent.empty:
                            st.caption("법인카드 사용금액 데이터가 없습니다.")
                        else:
                            st.markdown("#### 💳 법인카드 사용 실적 추이 (최근 3개년, 차입금 거래 기관만)")
                            fig_card = go.Figure()
                            for bank in card_df_recent['은행'].unique():
                                sub = card_df_recent[card_df_recent['은행'] == bank].sort_values('_연도숫자')
                                fig_card.add_trace(go.Bar(x=sub['연도'], y=sub['금액'] / 100000000, name=bank))
                            fig_card.update_layout(
                                barmode='group',
                                plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)', font=dict(color='#0f172a'),
                                yaxis=dict(gridcolor='rgba(15,23,42,0.06)'), xaxis=dict(title='연도'),
                                legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
                                margin=dict(t=70),
                                annotations=[dict(
                                    text='사용금액(억원)', xref='paper', yref='paper', x=0, y=1.1,
                                    xanchor='left', yanchor='bottom', showarrow=False, font=dict(size=12, color='#64748b')
                                )]
                            )
                            st.plotly_chart(fig_card, use_container_width=True)

                            latest_card_year = recent_years[-1]
                            card_latest = card_df_recent[card_df_recent['_연도숫자'] == latest_card_year]
                    else:
                        st.caption("법인카드 사용금액 파일이 업로드되지 않았습니다.")

                    # ---- 퇴직연금 적립금 현황 (차입금 거래 기관만) ----
                    # 별도 프로그램(퇴직연금통합관리시스템)이 이미 관리 중인 기관별 적립금
                    # 데이터를 그대로 끌어온다 — 예금·카드처럼 별도 업로드가 필요 없다.
                    # (pension_amounts_raw는 위쪽 게이트 조건에서 이미 조회해둔 값을 재사용)
                    pension_latest = {}
                    if pension_amounts_raw:
                        if active_loan_banks:
                            pension_latest = {k: v for k, v in pension_amounts_raw.items() if k in active_loan_banks}
                        else:
                            pension_latest = dict(pension_amounts_raw)

                        st.markdown(f"#### 🏦 퇴직연금 적립금 현황 ({pension_base_date} 기준, 차입금 거래 기관만)")
                        if not pension_latest:
                            st.caption("차입금 거래가 있는 기관 중에서는 퇴직연금 적립금 데이터와 매칭되는 항목이 없습니다.")
                        else:
                            # "거래관계 규모(원화 금액)"만으로는 그게 우리 전체 포트폴리오에서
                            # 큰 비중인지 작은 비중인지 알 수 없다 — 차입금 잔액ㆍ퇴직연금 적립금
                            # 각각의 "전체 대비 비중(%)"을 나란히 보여줘서, 이 기관과의 관계가
                            # 실제로 협상 레버리지가 될 만큼 큰지를 바로 판단할 수 있게 한다.
                            loan_balance_by_bank = get_loan_balance_by_bank(excel_bytes)
                            total_loan_balance = sum(loan_balance_by_bank.values())
                            total_pension = sum(pension_amounts_raw.values())

                            pension_rows = []
                            for k, v in pension_latest.items():
                                loan_bal = loan_balance_by_bank.get(k, 0.0)
                                pension_rows.append({
                                    "금융기관": k,
                                    "차입금 잔액(억원)": round(loan_bal / 100_000_000, 1),
                                    "차입금 비중(%)": round(loan_bal / total_loan_balance * 100, 1) if total_loan_balance else 0.0,
                                    "퇴직연금 적립금(억원)": round(v / 100_000_000, 1),
                                    "적립금 비중(%)": round(v / total_pension * 100, 1) if total_pension else 0.0,
                                })
                            pension_df_display = pd.DataFrame(pension_rows).sort_values(
                                "퇴직연금 적립금(억원)", ascending=False
                            ).reset_index(drop=True)

                            # 원형 도표 두 개를 나란히 놓으면 "같은 기관"이라는 연결이 눈에 안 들어온다는
                            # 피드백에 따라, 슬로프 차트(연결선 그래프)로 교체 — 왼쪽 축(차입금 비중)과
                            # 오른쪽 축(적립금 비중)에 같은 기관을 선으로 직접 이어서, 차입금 규모에 비해
                            # 퇴직연금이 얼마나 함께(비례해서) 움직이고 있는지를 한눈에 보여준다.
                            inst_colors = [
                                '#3b82f6', '#10b981', '#06b6d4', '#f59e0b', '#c084fc',
                                '#f472b6', '#ec4899', '#6366f1', '#14b8a6', '#64748b'
                            ]
                            slope_df = pension_df_display.sort_values("차입금 비중(%)", ascending=False).reset_index(drop=True)
                            inst_order = slope_df["금융기관"].tolist()
                            color_map = {name: inst_colors[i % len(inst_colors)] for i, name in enumerate(inst_order)}

                            fig_slope = go.Figure()
                            for _, row in slope_df.iterrows():
                                name = row["금융기관"]
                                loan_pct = row["차입금 비중(%)"]
                                pension_pct = row["적립금 비중(%)"]
                                gap = pension_pct - loan_pct
                                fig_slope.add_trace(go.Scatter(
                                    x=[0, 1], y=[loan_pct, pension_pct],
                                    mode='lines+markers+text',
                                    line=dict(color=color_map[name], width=2 + min(abs(gap) / 3, 4)),
                                    marker=dict(color=color_map[name], size=10, line=dict(color='#0f172a', width=1)),
                                    text=[f"{name}  {loan_pct:.1f}%", f"{name}  {pension_pct:.1f}%"],
                                    textposition=['middle left', 'middle right'],
                                    textfont=dict(size=12, color=color_map[name]),
                                    hovertemplate=f"{name}<br>%{{x}}<br>비중 %{{y:.1f}}%<extra></extra>",
                                    showlegend=False,
                                ))
                            fig_slope.update_layout(
                                xaxis=dict(
                                    tickmode='array', tickvals=[0, 1],
                                    ticktext=['💰 차입금 비중', '🏦 퇴직연금 적립금 비중'],
                                    range=[-0.55, 1.55], tickfont=dict(size=13, color='#0f172a'),
                                    showgrid=False, zeroline=False,
                                ),
                                yaxis=dict(title='전체 대비 비중(%)', gridcolor='rgba(15,23,42,0.06)', ticksuffix='%'),
                                plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
                                font=dict(color='#0f172a'), margin=dict(t=20, b=20, l=10, r=10),
                                height=120 + 55 * len(inst_order),
                            )
                            st.plotly_chart(fig_slope, use_container_width=True)

                            st.caption(
                                "선이 오른쪽으로 갈수록 내려가는 기관(차입금 비중 대비 적립금 비중이 작은 기관)일수록 "
                                "아직 다 쓰지 않은 협상 레버리지가 있다고 볼 수 있습니다.  \n"
                                "반대로 올라가는 기관은 이미 차입금 규모 이상으로 퇴직연금을 함께 고려해 운용하고 있다는 뜻입니다.  \n"
                                "선의 굵기는 두 비중의 격차(연계 정도)가 클수록 굵어집니다.  \n"
                                "비중은 각각 '당사 전체 차입금 잔액'ㆍ'당사 전체 퇴직연금 적립금'을 분모로 계산합니다"
                                "(위 도표는 차입금 거래가 있는 기관만 추렸지만, 비중 계산의 분모는 전체 기준입니다)."
                            )
                    else:
                        st.caption(
                            "🏦 퇴직연금 적립금 현황을 불러오지 못했습니다 — 퇴직연금통합관리시스템 서버(포트 8000)가 "
                            "켜져 있는지 확인해주세요."
                        )

                    if excluded_no_loan_names:
                        st.caption(
                            f"ℹ️ 차입금 현황이 없어 예금·카드 목록에서 제외된 기관: {', '.join(sorted(excluded_no_loan_names))}"
                        )

                    # ---- 관계 규모 x 대출금리 스프레드 성과 연계: 부수거래가 협상에 영향을 미쳤는지 분석 ----
                    if not dep_latest.empty or not card_latest.empty or pension_latest:
                        st.markdown("---")
                        st.markdown("#### 🔗 거래 관계 규모 vs 대출금리 협상 성과 — 부수거래가 협상에 영향을 미쳤는가?")
                        st.caption(
                            "⚠️ 당사와 **차입금(대출) 거래가 있는 금융기관**만 표시합니다(위 예금·카드·퇴직연금 표와 동일 기준).  \n"
                            "예금 잔액·법인카드 이용실적·퇴직연금 적립금을 합산한 '거래 관계 규모'가 클수록 대출금리 협상 성과"
                            "(🟢 방어·축소)로 이어졌는지를 확인해, 부수거래가 실제로 금리 협상에 영향을 미쳤는지를 판단합니다."
                        )

                        relation = {}
                        for _, r in dep_latest.iterrows():
                            norm = normalize_relation_name(r['금융기관'])
                            relation.setdefault(norm, {'예금잔액': 0.0, '카드사용액': 0.0, '퇴직연금적립금': 0.0})
                            relation[norm]['예금잔액'] += float(r['잔액'])
                        for _, r in card_latest.iterrows():
                            norm = normalize_relation_name(r['은행'])
                            relation.setdefault(norm, {'예금잔액': 0.0, '카드사용액': 0.0, '퇴직연금적립금': 0.0})
                            relation[norm]['카드사용액'] += float(r['금액'])
                        for norm, amt in pension_latest.items():
                            relation.setdefault(norm, {'예금잔액': 0.0, '카드사용액': 0.0, '퇴직연금적립금': 0.0})
                            relation[norm]['퇴직연금적립금'] += float(amt)

                        spread_map = {}
                        if not eval_df.empty:
                            for _, r in eval_df.iterrows():
                                norm = get_normalized_bank_name(r['금융기관'])
                                if norm:
                                    spread_map[norm] = {
                                        '핵심Δ스프레드': r['_핵심Δ스프레드'],
                                        '종합 상태': r['종합 상태']
                                    }

                        combo_rows = []
                        excluded_no_eval = []
                        for norm, vals in relation.items():
                            rel_size = vals['예금잔액'] + vals['카드사용액'] + vals['퇴직연금적립금']
                            if rel_size <= 0:
                                continue
                            sp = spread_map.get(norm)
                            if sp is None:
                                # 차입금 잔액은 있으나(위에서 이미 필터링됨) 3개년 금리이력이 부족해
                                # 스프레드 협상성과 평가 자체가 불가능한 기관 (연장 대기 등)
                                excluded_no_eval.append(norm)
                                continue
                            combo_rows.append({
                                '금융기관': norm,
                                '예금잔액(억원)': round(vals['예금잔액'] / 100000000, 1),
                                '카드사용액(억원)': round(vals['카드사용액'] / 100000000, 1),
                                '퇴직연금적립금(억원)': round(vals['퇴직연금적립금'] / 100000000, 1),
                                '거래관계 규모(억원)': round(rel_size / 100000000, 1),
                                'Δ스프레드(%p)': f"{sp['핵심Δ스프레드']:+.2f}",
                                '대출금리 협상 결과': sp['종합 상태']
                            })

                        if combo_rows:
                            combo_df = pd.DataFrame(combo_rows).sort_values('거래관계 규모(억원)', ascending=False).reset_index(drop=True)

                            # 거래 관계 규모(중앙값 기준 상/하위) x 협상 성과(양호 여부)를 교차 판정하여
                            # 부수거래가 실제 협상에 영향을 미쳤는지를 기관별로 명시적으로 표시한다.
                            median_size = combo_df['거래관계 규모(억원)'].median()

                            def _judge_influence(row):
                                is_large = row['거래관계 규모(억원)'] >= median_size
                                is_good = "🟢" in row['대출금리 협상 결과']
                                if is_large and is_good:
                                    return "🔵 영향 있음 (관계 상위·협상 우호적)"
                                if is_large and not is_good:
                                    return "🟡 영향 제한적 (관계는 크나 협상 미흡)"
                                if (not is_large) and is_good:
                                    return "⚪ 부수거래 외 요인 (관계는 작지만 협상 양호)"
                                return "🔴 관계·협상 모두 약세"

                            combo_df['부수거래 영향 판단'] = combo_df.apply(_judge_influence, axis=1)
                            combo_column_config = {
                                "금융기관": st.column_config.Column(width="small"),
                                "예금잔액(억원)": st.column_config.Column(width="small"),
                                "카드사용액(억원)": st.column_config.Column(width="small"),
                                "퇴직연금적립금(억원)": st.column_config.Column(width="small"),
                                "거래관계 규모(억원)": st.column_config.Column(width="small"),
                                "Δ스프레드(%p)": st.column_config.Column(width="small"),
                                "대출금리 협상 결과": st.column_config.Column(width="small"),
                                "부수거래 영향 판단": st.column_config.Column(width="large"),
                            }
                            st.dataframe(combo_df, use_container_width=True, hide_index=True, column_config=combo_column_config)

                            good_matched = [r for r in combo_rows if "🟢" in r['대출금리 협상 결과']]
                            c1, c2 = st.columns(2)
                            with c1:
                                st.markdown(f'<div class="metric-card"><div class="metric-label">차입 거래 기관 중 대출금리 협상 성공</div><div class="metric-value">{len(good_matched)}/{len(combo_rows)}개</div><div class="metric-desc positive">스프레드 방어·축소(🟢)로 이어진 비중</div></div>', unsafe_allow_html=True)
                            with c2:
                                top_bank = combo_df.iloc[0]
                                st.markdown(f'<div class="metric-card"><div class="metric-label">최대 거래 관계 기관</div><div class="metric-value">{top_bank["금융기관"]}</div><div class="metric-desc">{top_bank["거래관계 규모(억원)"]:,.0f}억원 · {top_bank["대출금리 협상 결과"]}</div></div>', unsafe_allow_html=True)

                            # 관계 규모 상위 절반 vs 하위 절반의 협상 성공률을 비교해,
                            # 부수거래 규모와 협상 성과 사이에 실제 상관관계가 있는지를 판단한다.
                            n = len(combo_df)
                            half = max(1, n // 2)
                            top_half = combo_df.iloc[:half]
                            bottom_half = combo_df.iloc[half:]
                            top_good_rate = top_half['대출금리 협상 결과'].str.contains("🟢").mean()
                            bottom_good_rate = (
                                bottom_half['대출금리 협상 결과'].str.contains("🟢").mean() if not bottom_half.empty else None
                            )

                            if bottom_good_rate is not None and top_good_rate > bottom_good_rate:
                                correlation_txt = (
                                    f"거래 관계 규모 상위 {half}개 기관의 협상 성공률(**{top_good_rate*100:.0f}%**)이 "
                                    f"하위 {n - half}개 기관(**{bottom_good_rate*100:.0f}%**)보다 높게 나타나, "
                                    "**예금·법인카드 등 부수거래를 확대한 것이 대출금리 협상에 긍정적인 영향을 미쳤다고 판단됩니다.**"
                                )
                            elif bottom_good_rate is not None and top_good_rate < bottom_good_rate:
                                correlation_txt = (
                                    f"거래 관계 규모 상위 {half}개 기관의 협상 성공률(**{top_good_rate*100:.0f}%**)이 오히려 "
                                    f"하위 {n - half}개 기관(**{bottom_good_rate*100:.0f}%**)보다 낮아, "
                                    "**이 시기에는 부수거래 규모보다 신용등급·시장금리 변동 등 다른 요인이 협상 결과에 더 크게 작용한 것으로 보입니다.**"
                                )
                            else:
                                correlation_txt = (
                                    "관계 규모 상·하위 그룹의 협상 성공률에 뚜렷한 차이가 없어, 부수거래 규모와 협상 성과 간 "
                                    "**직접적인 상관관계는 이번 분석 대상 기간에서는 확인되지 않았습니다.**"
                                )

                            n_influenced = int((combo_df['부수거래 영향 판단'].str.contains("🔵")).sum())
                            st.markdown(
                                "**📝 해석**  \n"
                                f"차입금 거래가 있는 {n}개 기관 중 **{len(good_matched)}개 기관**에서 "
                                "예금·법인카드 등 부수거래를 유지·확대한 것이 대출금리 스프레드 방어·축소로 이어졌습니다.  \n"
                                f"이 중 관계 규모도 크고 협상 성과도 좋은(🔵 영향 있음) 기관은 **{n_influenced}개**입니다. {correlation_txt}  \n"
                                "다만 스프레드가 확대(🔴)된 기관은 거래 관계 규모와 무관하게 재협상이 필요한 대상이므로, "
                                "위 표의 개별 기관별 '부수거래 영향 판단'을 함께 참고해주세요."
                            )
                            if excluded_no_eval:
                                st.caption(
                                    f"ℹ️ 차입금 현황은 있으나 3개년 금리이력 부족 등으로 협상성과 평가에서 제외된 기관: "
                                    f"{', '.join(sorted(set(excluded_no_eval)))}"
                                )
                        else:
                            st.caption("차입금 거래가 있는 기관 중에서는 예금·카드 부수거래 데이터와 매칭되는 항목이 없습니다.")


                    # ---- 이자비용 절감 KPI와 연결 ----
                    if latest_dep_year is not None and not dep_latest.empty:
                        total_dep_balance = float(dep_latest['잔액'].sum())
                        st.markdown("---")
                        st.markdown(
                            f"##### 📌 참고: 상단에서 계산된 연간 이자비용 절감액(**{saving_str}**)은 총 차입잔액 "
                            f"**{bal_str}** 대비 산출된 수치이며, 같은 시점 당사의 총 예치금 규모는 "
                            f"**{total_dep_balance/100000000:,.0f}억원({latest_dep_year}년 기준)** 입니다. "
                            "차입은행과 예치은행이 겹치는 구간에서는, 예금·카드 등 관계 거래를 꾸준히 유지한 것이 금리 협상 "
                            "여력 확보에 기여했다고 볼 수 있습니다."
                        )

        # -------------------------------------------------------------------
        # 🖨 인쇄용 요약(PDF) 화면 — KPI·스프레드 평가·부수거래 연동을 한 페이지로 압축
        # -------------------------------------------------------------------
        if page == PAGE_PRINT:
            st.info(
                "💡 아래 **'인쇄하기 / PDF 저장'** 버튼을 누르거나 브라우저 단축키(Windows: Ctrl+P, Mac: Cmd+P)를 사용하면 "
                "메뉴 없이 이 화면만 A4 용지 기준으로 인쇄하거나 PDF로 저장할 수 있습니다."
            )
            st.markdown(
                '<button class="no-print" onclick="window.print()" '
                'style="background:#3182f6;color:white;border:none;padding:10px 20px;border-radius:6px;'
                'font-weight:700;cursor:pointer;margin-bottom:16px;font-size:0.95rem;">🖨 인쇄하기 / PDF 저장</button>',
                unsafe_allow_html=True
            )

            gen_dt = datetime.now().strftime('%Y-%m-%d %H:%M')

            html_parts = ['<div class="print-page">']
            html_parts.append('<div class="print-title">📊 금융기관 금리협상 및 성과분석 — 인쇄용 요약</div>')
            html_parts.append(
                f'<div class="print-sub">생성일시: {gen_dt} · 기준연도: {CURRENT_YEAR}년(당해) / {PREV_YEAR}년(전년)</div>'
            )

            # --- KPI ---
            html_parts.append('<div class="print-section-title">🏆 핵심 성과 지표 (KPI)</div>')
            html_parts.append('<div class="print-kpi-grid">')
            html_parts.append(f'<div class="print-kpi-box"><div class="lbl">연간 이자비용 절감액</div><div class="val">{saving_str}</div></div>')
            html_parts.append(f'<div class="print-kpi-box"><div class="lbl">당해년도 가중평균 금리</div><div class="val">{kpi_calculated["avg_rate_curr"]:.2f}%</div></div>')
            html_parts.append(f'<div class="print-kpi-box"><div class="lbl">절감 금리(효과)</div><div class="val">{abs(kpi_calculated["saving_rate"]):.2f}%p</div></div>')
            html_parts.append(f'<div class="print-kpi-box"><div class="lbl">총 차입잔액</div><div class="val">{bal_str}</div></div>')
            html_parts.append('</div>')

            # --- 스프레드 평가 요약 ---
            html_parts.append('<div class="print-section-title">✅ 금융기관 신용스프레드 기반 금리 협상력 평가</div>')
            if df_hist is not None and not df_hist.empty and not eval_df.empty:
                n_total = len(eval_df)
                n_good = int((eval_df["구분"] == "잘한 것").sum())
                n_bad = n_total - n_good
                html_parts.append(
                    f'<div class="print-sub">평가 대상 {n_total}개 기관 중 <b>{n_good}개 기관</b> 스프레드 방어·축소(🟢), '
                    f'{n_bad}개 기관 확대(🟡/🔴) — 평가기간: {y_a}~{y_c}</div>'
                )
                small_eval = eval_df.drop(columns=["구분", "_핵심Δ스프레드"])
                html_parts.append(small_eval.to_html(index=False, classes="print-table", border=0, escape=False))
            else:
                html_parts.append('<div class="print-sub">평가 가능한 데이터가 부족합니다.</div>')

            # --- 부수거래 연동 (차입 거래 기관만) ---
            html_parts.append('<div class="print-section-title">💰 부수거래(예금·카드) ↔ 대출금리 연동 (차입 거래 기관만)</div>')

            if 'deposit_bytes' in st.session_state or 'card_bytes' in st.session_state:
                dep_by_year_p = load_deposit_contribution(st.session_state['deposit_bytes']) if 'deposit_bytes' in st.session_state else {}
                card_df_p = load_card_usage(st.session_state['card_bytes']) if 'card_bytes' in st.session_state else pd.DataFrame()

                dep_latest_p = pd.DataFrame()
                if dep_by_year_p:
                    dep_latest_p = dep_by_year_p[max(dep_by_year_p.keys())]
                card_latest_p = pd.DataFrame()
                if not card_df_p.empty:
                    card_latest_p = card_df_p[card_df_p['연도'] == sorted(card_df_p['연도'].unique())[-1]]

                relation_p = {}
                for _, r in dep_latest_p.iterrows():
                    norm = normalize_relation_name(r['금융기관'])
                    relation_p.setdefault(norm, {'예금잔액': 0.0, '카드사용액': 0.0})
                    relation_p[norm]['예금잔액'] += float(r['잔액'])
                for _, r in card_latest_p.iterrows():
                    norm = normalize_relation_name(r['은행'])
                    relation_p.setdefault(norm, {'예금잔액': 0.0, '카드사용액': 0.0})
                    relation_p[norm]['카드사용액'] += float(r['금액'])

                spread_map_p = {}
                if not eval_df.empty:
                    for _, r in eval_df.iterrows():
                        norm = get_normalized_bank_name(r['금융기관'])
                        if norm:
                            spread_map_p[norm] = {'핵심Δ스프레드': r['_핵심Δ스프레드'], '종합 상태': r['종합 상태']}

                combo_rows_p = []
                for norm, vals in relation_p.items():
                    rel_size = vals['예금잔액'] + vals['카드사용액']
                    if rel_size <= 0:
                        continue
                    sp = spread_map_p.get(norm)
                    if sp is None:
                        continue  # 차입 거래가 없는 기관은 인쇄용 요약에서도 동일하게 제외
                    combo_rows_p.append({
                        '금융기관': norm,
                        '예금잔액(억원)': round(vals['예금잔액'] / 100000000, 1),
                        '카드사용액(억원)': round(vals['카드사용액'] / 100000000, 1),
                        '거래관계 규모(억원)': round(rel_size / 100000000, 1),
                        'Δ스프레드(%p)': f"{sp['핵심Δ스프레드']:+.2f}",
                        '대출금리 협상 결과': sp['종합 상태']
                    })

                if combo_rows_p:
                    combo_df_p = pd.DataFrame(combo_rows_p).sort_values('거래관계 규모(억원)', ascending=False).reset_index(drop=True)
                    html_parts.append(combo_df_p.to_html(index=False, classes="print-table", border=0, escape=False))
                else:
                    html_parts.append('<div class="print-sub">차입 거래가 있는 기관 중 예금·카드 데이터와 매칭되는 항목이 없습니다.</div>')
            else:
                html_parts.append(
                    '<div class="print-sub">예금·카드 파일이 업로드되지 않아 부수거래 연동 내역을 표시할 수 없습니다. '
                    '("📈 차입 분석 → 💰 부수거래 연동 대출금리 영향도" 화면에서 먼저 업로드해주세요)</div>'
                )

            html_parts.append('</div>')  # /.print-page
            st.markdown("".join(html_parts), unsafe_allow_html=True)

if __name__ == "__main__":
    render_dashboard()
