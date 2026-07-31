# -*- coding: utf-8 -*-
"""월간보고서 시트를 엑셀과 최대한 비슷한 HTML로 그려주는 공용 모듈.

Streamlit에 의존하지 않아서(openpyxl만 사용) 차입금관리(Streamlit)와 통합포털(Flask)
양쪽에서 그대로 import해 재사용합니다. 두 군데서 서로 다르게 구현해 결과가 미묘하게
어긋나는 걸 막기 위해 이 파일 하나로 합쳤습니다.
"""
import re
from datetime import datetime

import openpyxl
from openpyxl.utils import column_index_from_string


def _format_report_cell_value(cell):
    """셀 값을 엑셀 표시 형식에 최대한 가깝게 문자열로 변환합니다."""
    val = cell.value
    if val is None:
        return ''
    if isinstance(val, datetime):
        return val.strftime('%Y-%m-%d')
    fmt = (cell.number_format or '').lower()
    if isinstance(val, (int, float)) and not isinstance(val, bool):
        if '%' in fmt:
            return f"{val * 100:,.1f}%"
        if val == int(val):
            return f"{int(val):,}"
        return f"{val:,.2f}"
    return str(val)


def _merge_span_map(ws):
    """{(top_left_row, top_left_col): (rowspan, colspan)}와 병합에 가려진 (r,c) 집합을 반환."""
    merged_span = {}
    covered = set()
    for merged_range in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds
        merged_span[(min_row, min_col)] = (max_row - min_row + 1, max_col - min_col + 1)
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                if (r, c) != (min_row, min_col):
                    covered.add((r, c))
    return merged_span, covered


def _print_area_bounds(ws):
    """시트에 인쇄영역(print_area)이 지정돼 있으면 그 범위의 (max_row, max_col)을
    반환하고, 없으면 None을 반환합니다. 어음현황처럼 실제 사용범위(max_row/max_column)가
    인쇄영역보다 넓은 시트가 있어(예: 인쇄영역은 A1:T31인데 사용범위는 AA열까지 있음),
    그 차이만큼의 부속 표/메모가 같은 HTML 표에 섞여 화면이 뒤죽박죽 보이는 문제가
    있었습니다. 엑셀에서 실제로 인쇄되는 범위만 그대로 그리면 이 문제가 해결됩니다."""
    area = ws.print_area
    if not area:
        return None
    first_area = area.split(',')[0].strip()
    m = re.search(r'\$?([A-Za-z]+)\$?(\d+):\$?([A-Za-z]+)\$?(\d+)', first_area)
    if not m:
        return None
    try:
        col1 = column_index_from_string(m.group(1).upper())
        row1 = int(m.group(2))
        col2 = column_index_from_string(m.group(3).upper())
        row2 = int(m.group(4))
    except ValueError:
        return None
    return max(row1, row2), max(col1, col2)


def get_print_scale(ws):
    """엑셀에 이미 지정돼 있는 인쇄배율(페이지 설정 > 확대/축소, %)을 0~1 사이 소수로
    반환합니다. 지정돼 있지 않으면 None — 이 경우 화면단에서 A4 폭에 맞춰 자동으로
    배율을 계산하는 방식으로 대체합니다. 각 시트마다 이미 사람이 확인하고 맞춰둔
    배율이 있으므로, 있으면 그 값을 그대로 쓰는 편이 항상 더 정확합니다."""
    scale = ws.page_setup.scale
    if not scale:
        return None
    try:
        return round(float(scale) / 100.0, 4)
    except (TypeError, ValueError):
        return None


_BORDER_COLOR = '#8a8f98'


def _box_border_style(ws, min_row, min_col, max_row, max_col):
    """병합(또는 단일) 셀 박스의 실제 테두리만 반환합니다 — 엑셀 원본에 그어진 테두리가
    있는 변(top/left는 좌상단 셀 기준, bottom/right는 우하단 셀 기준)만 CSS로 그리고,
    빈 채움 셀처럼 테두리가 없는 칸은 선 없이(투명하게) 둡니다. 그래야 화면/인쇄 모두
    엑셀에 실제로 그어진 표 모양만 보이고, 사용범위 전체에 깔리는 격자선이 사라집니다."""
    top_left = ws.cell(row=min_row, column=min_col)
    bottom_right = ws.cell(row=max_row, column=max_col)
    styles = []
    if top_left.border and top_left.border.top and top_left.border.top.style:
        styles.append(f'border-top:1px solid {_BORDER_COLOR}')
    if top_left.border and top_left.border.left and top_left.border.left.style:
        styles.append(f'border-left:1px solid {_BORDER_COLOR}')
    if bottom_right.border and bottom_right.border.bottom and bottom_right.border.bottom.style:
        styles.append(f'border-bottom:1px solid {_BORDER_COLOR}')
    if bottom_right.border and bottom_right.border.right and bottom_right.border.right.style:
        styles.append(f'border-right:1px solid {_BORDER_COLOR}')
    return styles


def render_sheet_html(ws, css_class="report-sheet"):
    """openpyxl 워크시트를 병합 셀(rowspan/colspan)·굵은글씨·정렬을 최대한 보존한
    읽기 전용 HTML 표로 변환합니다(데이터를 재해석하지 않고 눈에 보이는 그대로 재현).
    엑셀에 인쇄영역이 지정돼 있으면 그 범위만 그리고(사용범위 전체가 아님), 엑셀에
    지정된 인쇄배율이 있으면 표에 data-excel-scale로 실어 화면단이 그대로 쓸 수 있게 합니다."""
    merged_span, covered = _merge_span_map(ws)
    bounds = _print_area_bounds(ws)
    max_row = bounds[0] if bounds else ws.max_row
    max_col = bounds[1] if bounds else ws.max_column

    rows_html = []
    for r in range(1, max_row + 1):
        cells_html = []
        for c in range(1, max_col + 1):
            if (r, c) in covered:
                continue
            cell = ws.cell(row=r, column=c)
            rowspan, colspan = merged_span.get((r, c), (1, 1))
            rowspan = min(rowspan, max_row - r + 1)
            colspan = min(colspan, max_col - c + 1)
            text = _format_report_cell_value(cell)
            style = _box_border_style(ws, r, c, r + rowspan - 1, c + colspan - 1)
            if cell.font and cell.font.bold:
                style.append('font-weight:700')
            align = cell.alignment.horizontal if cell.alignment else None
            if align:
                style.append(f'text-align:{align}')
            elif isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
                style.append('text-align:right')
            attrs = ''
            if rowspan > 1:
                attrs += f' rowspan="{rowspan}"'
            if colspan > 1:
                attrs += f' colspan="{colspan}"'
            style_attr = f' style="{";".join(style)}"' if style else ''
            escaped = text.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
            cells_html.append(f'<td{attrs}{style_attr}>{escaped}</td>')
        if cells_html:
            rows_html.append('<tr>' + ''.join(cells_html) + '</tr>')

    scale = get_print_scale(ws)
    scale_attr = f' data-excel-scale="{scale}"' if scale else ''
    return f'<table class="{css_class}"{scale_attr}>' + ''.join(rows_html) + '</table>'


def render_sheet_editable_html(ws, css_class="report-sheet report-sheet-edit"):
    """render_sheet_html과 같은 병합/서식 구조를 보존하되, 각 셀을 편집 가능(contenteditable)
    하게 만들고 data-row/data-col을 달아둡니다. 화면에서 Excel 값을 그대로 복사해 붙여넣고
    수정한 뒤 저장할 수 있게 하기 위한 용도입니다 — 병합된 셀은 좌상단 위치 하나만
    수정 대상이 됩니다(엑셀에서도 병합 셀의 실제 값은 좌상단에만 있는 것과 동일)."""
    merged_span, covered = _merge_span_map(ws)

    rows_html = []
    for r in range(1, ws.max_row + 1):
        cells_html = []
        for c in range(1, ws.max_column + 1):
            if (r, c) in covered:
                continue
            cell = ws.cell(row=r, column=c)
            rowspan, colspan = merged_span.get((r, c), (1, 1))
            text = _format_report_cell_value(cell)
            style = _box_border_style(ws, r, c, r + rowspan - 1, c + colspan - 1)
            if cell.font and cell.font.bold:
                style.append('font-weight:700')
            align = cell.alignment.horizontal if cell.alignment else None
            if align:
                style.append(f'text-align:{align}')
            attrs = f' data-row="{r}" data-col="{c}" contenteditable="true"'
            if rowspan > 1:
                attrs += f' rowspan="{rowspan}"'
            if colspan > 1:
                attrs += f' colspan="{colspan}"'
            style_attr = f' style="{";".join(style)}"' if style else ''
            escaped = text.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
            cells_html.append(f'<td{attrs}{style_attr}>{escaped}</td>')
        if cells_html:
            rows_html.append('<tr>' + ''.join(cells_html) + '</tr>')
    return f'<table class="{css_class}" data-max-row="{ws.max_row}" data-max-col="{ws.max_column}">' + ''.join(rows_html) + '</table>'


def _coerce_cell_value(text):
    """편집 화면에서 넘어온 문자열을 숫자로 바꿀 수 있으면 바꾸고, 아니면 문자열 그대로 둡니다."""
    text = (text or '').strip()
    if text == '':
        return None
    cleaned = text.replace(',', '')
    try:
        if cleaned.lstrip('-').isdigit():
            return int(cleaned)
        return float(cleaned)
    except ValueError:
        return text


def apply_grid_edits_to_workbook(wb, sheet_name, max_row, max_col, values):
    """편집 화면에서 받은 {(row, col): text} 값을 워크북의 시트에 그대로 씁니다.
    같은 이름의 시트가 있으면 지우고 새로 만듭니다 — 병합은 다시 만들지 않고
    평범한 셀로 저장합니다(데이터는 정확히 보존되고, 다시 열람할 때도
    render_sheet_html이 그대로 보여줄 수 있습니다 — 다만 셀 병합 모양 자체는
    사라집니다. 인접 셀에 값이 반복돼 보일 수 있다는 점만 다름)."""
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            val = values.get(f"{r}_{c}")
            if val is None:
                continue
            ws.cell(row=r, column=c, value=_coerce_cell_value(val))
    return wb


def check_ftc_disclosure(transaction_type, amount, capital_base, is_goods_services_target=True):
    """대규모내부거래 이사회 의결ㆍ공시 대상 여부를 판단한다(공정거래법 제26조,
    시행령 제33조, 공정위 고시 기준). 주가현황/server.py의 동일 함수와 로직을
    그대로 맞춰뒀습니다(자금팀이 공시 탭에서 이미 쓰고 있는 판단 기준과 동일).

    - 공통 거래금액 기준: 거래금액이 100억원 이상이거나, 회사의 자본총계ㆍ자본금 중
      큰 금액(capital_base)의 5% 이상(단, 그 5% 값이 5억원 미만이면 5억원을 기준으로
      적용 — 즉 최소 5억원은 넘어야 공시대상).
    - transaction_type: 'fund_securities_asset'(자금ㆍ유가증권ㆍ자산 거래) 또는
      'goods_services'(상품ㆍ용역 거래).
    - goods_services 거래는 위 금액기준을 충족하더라도, 거래상대방이 "동일인 및
      동일인 친족이 발행주식총수의 20% 이상을 소유한 계열회사(또는 그 계열회사의
      상법상 50% 초과 자회사)"인 경우에만 대상이다.
    """
    if amount is None or capital_base is None or capital_base <= 0:
        return None
    threshold = max(capital_base * 0.05, 500_000_000)
    amount_ge_100eok = amount >= 10_000_000_000
    amount_ge_capital_pct = amount >= threshold
    amount_triggered = amount_ge_100eok or amount_ge_capital_pct
    is_goods = transaction_type == 'goods_services'

    base = {
        "transaction_type": transaction_type,
        "amount_ge_100eok": amount_ge_100eok,
        "amount_ge_capital_pct": amount_ge_capital_pct,
        "amount_triggered": amount_triggered,
        "threshold_amount": int(threshold),
    }

    if is_goods and not is_goods_services_target:
        reason = "거래상대방이 상품ㆍ용역거래 특례 대상 계열회사(20% 계열사)에 해당하지 않아, 금액과 무관하게 공시대상이 아닙니다."
        return {**base, "is_disclosure_required": False, "reason": reason}

    if amount_ge_100eok and amount_ge_capital_pct:
        amount_reason = "거래금액이 100억원 이상이고 자본총계ㆍ자본금 중 큰 금액의 5%(최소 5억원) 기준도 충족해 공시대상입니다."
    elif amount_ge_100eok:
        amount_reason = "거래금액이 100억원 이상이므로 공시대상입니다."
    elif amount_ge_capital_pct:
        amount_reason = "거래금액은 100억원 미만이지만, 자본총계ㆍ자본금 중 큰 금액의 5%(최소 5억원) 기준을 충족해 공시대상입니다."
    else:
        amount_reason = "거래금액이 100억원 미만이고 자본총계ㆍ자본금 중 큰 금액의 5%(최소 5억원) 기준도 충족하지 않아 공시대상이 아닙니다."

    reason = f"거래상대방이 상품ㆍ용역거래 특례 대상 계열회사(20% 계열사)에 해당하며, {amount_reason}" if is_goods else amount_reason

    return {**base, "is_disclosure_required": bool(amount_triggered), "reason": reason}


def _norm(s):
    return str(s).replace(' ', '').replace('\n', '') if s is not None else ''


def _find_row_containing(ws, predicate, start_row, end_row, max_scan_col=6):
    """조건을 만족하는 첫 번째 행."""
    for r in range(start_row, min(end_row, ws.max_row) + 1):
        for c in range(1, min(ws.max_column, max_scan_col) + 1):
            v = ws.cell(row=r, column=c).value
            if v is not None and predicate(_norm(v)):
                return r
    return None


def _find_last_row_containing(ws, predicate, start_row, end_row, max_scan_col=6):
    """조건을 만족하는 마지막 행 — '합계/총계'는 소구간(1년미만 등)마다 반복되고
    구간 전체의 진짜 총합은 맨 마지막에 나오는 경우가 많아, 총계를 찾을 땐 이 함수를 쓴다."""
    found = None
    for r in range(start_row, min(end_row, ws.max_row) + 1):
        for c in range(1, min(ws.max_column, max_scan_col) + 1):
            v = ws.cell(row=r, column=c).value
            if v is not None and predicate(_norm(v)):
                found = r
                break
    return found


def _find_col_by_header(ws, header_row, label):
    target = _norm(label)
    for c in range(1, ws.max_column + 1):
        if _norm(ws.cell(row=header_row, column=c).value) == target:
            return c
    return None


def _extract_rate_after_label(ws, row, label_substr='평균금리'):
    seen = False
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=row, column=c).value
        if isinstance(v, str) and label_substr in _norm(v):
            seen = True
            continue
        if seen and isinstance(v, (int, float)) and not isinstance(v, bool):
            return v
    return None


def compute_invest_summary(ws):
    """자금운용현황 시트에서 이미 계산돼 있는 '가용자금 총계'/'비가용자금 합계' 행을
    그대로 읽어옵니다(직접 재계산하지 않음 — 원본 엑셀의 수치와 항상 일치하도록).
    반환: {"available": {"amount":..,"rate":..} or None, "unavailable": {...} or None}
    금액은 원 단위, 금리는 0.0463 같은 소수(퍼센트 아님) 그대로 반환합니다."""
    avail_title_row = _find_row_containing(ws, lambda s: '가용자금' in s and '비가용' not in s, 1, ws.max_row)
    nonavail_title_row = _find_row_containing(ws, lambda s: '비가용' in s, 1, ws.max_row)
    # 비가용자금 구간 뒤에 '외화(USD) 운용 현황' 같은 별도 섹션이 더 있을 수 있어(그
    # 섹션도 자체 '합계' 행을 갖고 있음), 비가용자금 구간의 끝을 거기서 잘라야 한다.
    next_section_row = _find_row_containing(
        ws, lambda s: '외화' in s or 'usd' in s.lower(), (nonavail_title_row or 1) + 1, ws.max_row
    ) if nonavail_title_row else None

    result = {"available": None, "unavailable": None}

    if avail_title_row:
        # 다음 섹션(비가용자금 또는 외화 운용현황) 시작 전까지가 가용자금 구간.
        section_end = (nonavail_title_row - 1) if nonavail_title_row else ws.max_row
        header_row = _find_row_containing(ws, lambda s: s == '금액', avail_title_row, section_end)
        # '합계'는 소구간(수시입출금/1년미만/1년이상)마다 반복되고 구간 전체 총합("총계")은
        # 맨 마지막에 나오므로, 마지막으로 매칭되는 행을 총합으로 본다. 정확히 '합계'/'총계'만
        # 매칭해 "운용 총계(가용+비가용)"처럼 다른 말이 섞인 행은 제외한다.
        total_row = _find_last_row_containing(
            ws, lambda s: s in ('총계', '합계'), avail_title_row, section_end
        )
        if header_row and total_row:
            amount_col = _find_col_by_header(ws, header_row, '금액')
            amount = ws.cell(row=total_row, column=amount_col).value if amount_col else None
            rate = _extract_rate_after_label(ws, total_row)
            if isinstance(amount, (int, float)):
                result["available"] = {"amount": amount, "rate": rate}

    if nonavail_title_row:
        section_end = (next_section_row - 1) if next_section_row else ws.max_row
        header_row = _find_row_containing(ws, lambda s: s == '금액', nonavail_title_row, section_end)
        total_row = _find_last_row_containing(
            ws, lambda s: s in ('총계', '합계'), nonavail_title_row, section_end
        )
        if header_row and total_row:
            amount_col = _find_col_by_header(ws, header_row, '금액')
            amount = ws.cell(row=total_row, column=amount_col).value if amount_col else None
            rate = _extract_rate_after_label(ws, total_row)
            if isinstance(amount, (int, float)):
                result["unavailable"] = {"amount": amount, "rate": rate}

    return result


def load_workbook_from_bytes(file_bytes, data_only=False):
    import io
    return openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=data_only)


def workbook_to_bytes(wb):
    import io
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
