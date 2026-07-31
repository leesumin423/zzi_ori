# -*- coding: utf-8 -*-
"""그룹(계열사) 차입금 보고서 분석 — streamlit 의존성 없음(report_shared.py와 같은 이유로
분리: 화면 쪽 app.py와 통합포털 양쪽에서 재사용 가능하게).

유진그룹에 매월 제출하는 '그룹차입금보고서.xlsx'의 맨 마지막 시트(예: '8.동양')는
회사마다 같은 서식이다 — 금융기관/지점/과목/한도/잔액/대출일자/만기일/이율 등을
기관별로 나열하고 기관마다 '소계' 행이 있으며, 맨 끝에 전체 '계' 행이 있고 그 뒤에
(집계 범위 밖인) '관계사차입' 섹션이 따로 붙는다. 이 모듈은 그 시트를 파싱해서,
그룹차입금보고서의 '3.금융기관별(잔액)'/'4.금융기관별(금리)'/'5.과목별(잔액)'/
'1.그룹총괄표' 시트와 같은 구조의 분석표를 7개 법인(GROUP_COMPANIES) 기준으로
직접 만든다 — 원래는 유진그룹 전체(23개 법인)를 담당자가 취합해야 나오는 표지만,
우리가 관리하는 7개 법인만 모아서 동일한 서식으로 재현한다."""
import openpyxl

GROUP_COMPANIES = [
    '동양', '동양에너지', '금왕에프원', '디씨아이티와이부천', '인천피에프브이', '유진마포130',
    '유진한일합섬',
]

# 3.금융기관별(잔액)/4.금융기관별(금리) 시트와 동일한 열 순서(은행권 10개 + 비은행 3개)
BANK_COLUMNS = ['농협', '산업', '우리', '국민', '하나', '신한', '대구', '기업', '수협', '기타']
NONBANK_COLUMNS = ['증권금융', '우리종금', '증권사 및 기타']
INSTITUTION_COLUMNS = BANK_COLUMNS + NONBANK_COLUMNS

_BANK_ALIASES = {
    '농협': '농협', 'NH농협': '농협', 'NH농협은행': '농협', '농협은행': '농협',
    '산업': '산업', '산업은행': '산업', 'KDB산업은행': '산업',
    '우리': '우리', '우리은행': '우리',
    '국민': '국민', 'KB국민': '국민', '국민은행': '국민', 'KB': '국민',
    '하나': '하나', '하나은행': '하나',
    '신한': '신한', '신한은행': '신한',
    '대구': '대구', '대구은행': '대구', 'im뱅크': '대구', 'iM뱅크': '대구',
    '아이엠뱅크': '대구', 'DGB대구은행': '대구',
    '기업': '기업', '기업은행': '기업', 'IBK': '기업', 'IBK기업은행': '기업',
    '수협': '수협', '수협은행': '수협',
    '증권금융': '증권금융', '한국증권금융': '증권금융',
    '우리종금': '우리종금', '우리종합금융': '우리종금',
}
_NONBANK_LITERALS = {'회사채', '전단채', 'CP', '증권사및기타', '증권사및\n기타'}


def normalize_institution_column(raw_institution):
    """8.동양류 시트의 '금융기관' 원문을 3/4번 시트의 고정 열 이름으로 매핑한다.
    회사채ㆍ전단채ㆍCPㆍ증권사및기타류는 전부 '증권사 및 기타'로 묶이고(그룹차입금
    보고서 3번 시트가 실제로 그렇게 합산돼 있음 — 동양 사례로 확인: 회사채 400억이
    O열(증권사 및 기타)에 그대로 들어감), 못 알아보는 은행명은 '기타'로 둔다."""
    name = str(raw_institution or '').strip()
    name_compact = name.replace(' ', '').replace('\n', '')
    if not name_compact:
        return None
    if name_compact in _NONBANK_LITERALS or name_compact.upper() == 'CP':
        return '증권사 및 기타'
    if name_compact in _BANK_ALIASES:
        return _BANK_ALIASES[name_compact]
    return '기타'


def classify_subject_bucket(institution_raw, subject_text):
    """5.과목별(잔액) 분류용 — 은행권 일반대/한도대 vs 회사채/전단채/CP/기타."""
    inst = str(institution_raw or '').strip().replace(' ', '').replace('\n', '')
    if inst == '회사채':
        return 'bond'
    if inst == '전단채':
        return 'stnote'
    if inst.upper() == 'CP':
        return 'cp'
    subj = str(subject_text or '')
    if '한도대' in subj:
        return 'limit'
    if '일반대' in subj:
        return 'general'
    return 'other'


# 시트 머리말이 "(단위: 억원)"이라도, 일부 법인 파일은 특정 줄만 실수로 원(WON)
# 단위 그대로 입력해둔 경우가 실제로 있었다(예: 금왕에프원의 '증권사 및 기타'
# 섹션 — 4.649억원짜리 대출을 "464900000"으로 입력, 나머지 섹션은 정상적으로
# "억원" 단위였음). 대출 한 줄이 10만억원(100,000)을 넘는 건 현실적으로 불가능한
# 규모라, 그럴 땐 원 단위로 잘못 입력된 것으로 보고 1억으로 나눠 억원으로 바꾼다.
# 개별법인 제출용 서식(parse_individual_submission_sheet)은 애초에 통째로 원
# 단위라 모든 값이 이 문턱을 넘어서고, 결과적으로 항상 억원으로 정상 환산된다.
_IMPLAUSIBLE_EOK = 100_000


def _fix_unit(value):
    if value and abs(value) > _IMPLAUSIBLE_EOK:
        return value / 100_000_000, True
    return value, False


def detect_sheet_format(ws):
    """업로드된 시트가 두 서식 중 어느 쪽인지 헤더 텍스트로 판별한다.
    'consolidated' = 그룹차입금보고서 형식('8.동양' 등, "금융기관" 헤더),
    'individual' = 계열사가 개별적으로 회신하는 [참고2] 소속기업체 차입금현황 서식
    ("법인명"+"차입처" 헤더). 둘 다 없으면 None."""
    for r in range(1, min(ws.max_row, 10) + 1):
        row_keys = set()
        for c in range(1, min(ws.max_column, 15) + 1):
            v = ws.cell(r, c).value
            if v:
                row_keys.add(str(v).strip().replace('\n', '').replace(' ', ''))
        if '금융기관' in row_keys:
            return 'consolidated'
        if '법인명' in row_keys and '차입처' in row_keys:
            return 'individual'
    return None


def parse_company_sheet(ws):
    """'8.동양' 형식 시트에서 [{institution_raw, subject, limit, balance, rate}, ...]를
    뽑는다. '계' 행(전체 소계) 이후의 관계사차입 섹션은 이 분석 범위 밖이라 제외한다."""
    header_row = None
    header_col = None
    for r in range(1, min(ws.max_row, 10) + 1):
        for c in range(1, min(ws.max_column, 6) + 1):
            v = ws.cell(r, c).value
            if v and '금융기관' in str(v):
                header_row, header_col = r, c
                break
        if header_row:
            break
    if header_row is None:
        raise ValueError("'금융기관' 헤더 행을 찾을 수 없습니다 — 시트 서식을 확인해주세요.")

    col_map = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(header_row, c).value
        if v:
            col_map[str(v).strip().replace('\n', '').replace(' ', '')] = c
    inst_col = header_col
    subj_col = col_map.get('과목', inst_col + 2)
    limit_col = col_map.get('한도', inst_col + 3)
    balance_col = col_map.get('잔액', inst_col + 4)
    rate_col = col_map.get('이율') or col_map.get('금리')

    items = []
    unit_fix_count = 0
    current_inst = None
    for r in range(header_row + 1, ws.max_row + 1):
        a_val = ws.cell(r, inst_col).value
        a_text = str(a_val).strip() if a_val is not None else ''
        if a_text == '계':
            break  # 관계사차입 등 이후 섹션은 집계 범위 밖
        if a_text == '소계':
            continue  # 소계는 라인아이템에서 직접 재계산하므로 건너뜀
        if a_text:
            current_inst = a_text
        if not current_inst:
            continue

        limit_v = limit_col and ws.cell(r, limit_col).value
        balance_v = balance_col and ws.cell(r, balance_col).value
        rate_v = rate_col and ws.cell(r, rate_col).value
        limit_num = limit_v if isinstance(limit_v, (int, float)) else 0.0
        balance_num = balance_v if isinstance(balance_v, (int, float)) else 0.0
        if limit_num == 0 and balance_num == 0:
            continue  # 빈 기관 섹션의 빈 줄(값 없음)은 라인아이템으로 만들지 않음

        limit_num, fixed1 = _fix_unit(float(limit_num))
        balance_num, fixed2 = _fix_unit(float(balance_num))
        if fixed1 or fixed2:
            unit_fix_count += 1

        subj = ws.cell(r, subj_col).value if subj_col else None
        items.append({
            'institution_raw': current_inst,
            'subject': str(subj).strip() if subj else '',
            'limit': limit_num,
            'balance': balance_num,
            'rate': float(rate_v) if isinstance(rate_v, (int, float)) else None,
        })
    return items, unit_fix_count


def parse_individual_submission_sheet(ws):
    """계열사가 개별적으로 회신하는 '[참고2] 소속기업체 차입금현황' 서식(예:
    "(유진마포130)01_유진기업_차입금현황.xlsx")을 파싱한다. '8.동양' 형식(그룹차입금
    보고서)과는 헤더/구조가 전혀 다르다:
    - 헤더: 법인명ㆍ구분ㆍ차입처ㆍ차입한도ㆍ차입잔액ㆍ금리ㆍ시작일ㆍ만기일ㆍ담보 및
      신용보강ㆍ비고 ("금융기관" 헤더가 아니라 "법인명"+"차입처"로 판별)
    - '구분' 열이 은행 실무 관점(일반대/한도대)이 아니라 회계ㆍ공시 관점 카테고리
      (금융기관 차입금/사채 외/관계사 차입금/환매조건부채권매도/콜매도/리스부채/기타)
    - 단위가 항상 원(WON) — _fix_unit()이 억원으로 자동 환산해준다(값이 항상
      10만억원 문턱을 넘으므로 매번 정상적으로 변환된다)
    반환 형식은 parse_company_sheet와 동일한 [{institution_raw, subject, limit,
    balance, rate}, ...] — 그룹 차입금 분석의 institution/subject 분류 함수를
    그대로 재사용할 수 있게 맞춘다. '차입처'가 비어있는 빈 카테고리 행(서식에
    미리 인쇄된 사채 외/리스부채 등 빈 자리)은 건너뛴다."""
    header_row = None
    col_map = {}
    for r in range(1, min(ws.max_row, 10) + 1):
        row_col_map = {}
        for c in range(1, min(ws.max_column, 15) + 1):
            v = ws.cell(r, c).value
            if v:
                row_col_map[str(v).strip().replace('\n', '').replace(' ', '')] = c
        if '법인명' in row_col_map and '차입처' in row_col_map:
            header_row, col_map = r, row_col_map
            break
    if header_row is None:
        raise ValueError("'법인명/차입처' 헤더 행을 찾을 수 없습니다 — 시트 서식을 확인해주세요.")

    name_col = col_map.get('법인명')
    category_col = col_map.get('구분')
    counterparty_col = col_map.get('차입처')
    limit_col = col_map.get('차입한도')
    balance_col = col_map.get('차입잔액')
    rate_col = col_map.get('금리')

    items = []
    unit_fix_count = 0
    current_category = None
    for r in range(header_row + 1, ws.max_row + 1):
        name_v = ws.cell(r, name_col).value if name_col else None
        name_text = str(name_v).strip().replace(' ', '') if name_v is not None else ''
        if name_text == '합계':
            break  # 총계 행 이후는 비고/유의사항 텍스트라 집계 범위 밖

        cat_v = ws.cell(r, category_col).value if category_col else None
        cat_text = str(cat_v).strip() if cat_v is not None else ''
        if cat_text:
            current_category = cat_text

        counterparty_v = ws.cell(r, counterparty_col).value if counterparty_col else None
        counterparty_text = str(counterparty_v).strip() if counterparty_v is not None else ''

        limit_v = limit_col and ws.cell(r, limit_col).value
        balance_v = balance_col and ws.cell(r, balance_col).value
        rate_v = rate_col and ws.cell(r, rate_col).value
        limit_num = limit_v if isinstance(limit_v, (int, float)) else 0.0
        balance_num = balance_v if isinstance(balance_v, (int, float)) else 0.0
        if limit_num == 0 and balance_num == 0:
            continue  # 서식에 미리 인쇄된 빈 카테고리 자리(사채 외/리스부채 등)

        limit_num, fixed1 = _fix_unit(float(limit_num))
        balance_num, fixed2 = _fix_unit(float(balance_num))
        if fixed1 or fixed2:
            unit_fix_count += 1

        items.append({
            'institution_raw': counterparty_text or current_category or '기타',
            'subject': current_category or '',
            'limit': limit_num,
            'balance': balance_num,
            'rate': float(rate_v) if isinstance(rate_v, (int, float)) else None,
        })
    return items, unit_fix_count


def _normalize_sheet_name(s):
    return str(s).strip().replace(' ', '').replace('(주)', '').replace('㈜', '')


def find_company_sheet_name(wb, company_name):
    """업로드된 워크북에서 이 법인의 상세 시트를 이름으로 찾는다 — 예전엔 무조건
    '맨 마지막 시트'를 썼는데, 법인마다 파일 구조가 달라서(참고 시트가 더 붙어있거나
    시트 순서가 다르거나) 엉뚱한 시트가 파싱될 수 있었다(실제로 이 문제로 보임).
    이제 선택한 법인명으로 끝나는 시트를 우선 찾는다(예: '8.동양' → '동양'로 끝남).
    "~로 끝난다"로 비교하는 이유: '동양'을 찾을 때 '동양에너지' 시트가 부분일치로
    같이 걸리면 안 되는데, '동양에너지'는 '동양'으로 끝나지 않으므로 자연스럽게
    걸러진다. (주)/㈜ 표기 차이는 정규화해서 흡수한다."""
    name_norm = _normalize_sheet_name(company_name)
    candidates = [s for s in wb.sheetnames if _normalize_sheet_name(s).endswith(name_norm)]
    if candidates:
        return candidates[-1]
    candidates = [s for s in wb.sheetnames if name_norm in _normalize_sheet_name(s)]
    if candidates:
        return candidates[-1]
    return wb.sheetnames[-1]  # 이름으로 못 찾으면(양식이 많이 다른 경우) 마지막 시트로 폴백


def detect_company_from_workbook(wb):
    """업로드된 워크북의 시트 이름만 보고 어느 법인 파일인지 추정한다 — 관리자가
    '법인' 드롭다운을 새로 안 바꾸고 다른 회사 파일을 올려도(또는 실수로 안 바꿔서)
    엉뚱한 법인 이름으로 저장되는 걸 막기 위해, 업로드 즉시 파일 내용과 드롭다운
    선택이 맞는지 대조하는 용도다. 여러 법인이 동시에 매칭되면(흔치 않음) 이름이
    더 긴(더 구체적인) 쪽을 우선한다(예: '동양'과 '동양에너지'가 같이 매칭되면
    '동양에너지'를 우선 — 반대 방향 충돌은 endswith 특성상 애초에 안 생긴다)."""
    matches = []
    for company in GROUP_COMPANIES:
        name_norm = _normalize_sheet_name(company)
        if any(_normalize_sheet_name(s).endswith(name_norm) for s in wb.sheetnames):
            matches.append(company)
    if not matches:
        return None
    matches.sort(key=len, reverse=True)
    return matches[0]


# 일부 법인은 원(WON) 단위 실수가 흔한 100,000,000(1억) 배가 아니라 1,000,000(100만)
# 배로 부풀려 입력돼 있는 게 실제로 확인됐다(금왕에프원 — 그룹총괄표에 사람이 이미
# 확정해둔 '1,264억원'과 대조해서 확인: 시트 계산값이 100배 작았음, 회사채 등 다른
# 섹션은 전부 0이라 같은 시트 안에서 대조할 다른 기준이 없어 일반적인 자릿수 추정으로는
# 못 잡는 경우). 이런 법인은 여기 추가하고, 새로 다른 법인에서 같은 증상이 확인되면
# 똑같이 추가하면 된다 — 확실하지 않은 채로 전체에 적용하면 오히려 정상 데이터(예:
# 동양)를 망가뜨릴 수 있어 회사명으로 범위를 좁혀둔다.
_UNIT_FIX_DIVISOR_OVERRIDE = {
    '금왕에프원': 1_000_000,
}


def load_company_items_from_bytes(file_bytes, company=None):
    """업로드된 엑셀 바이트에서 회사 상세 시트를 찾아 파싱한다. company를 주면 그
    법인명으로 끝나는 시트를 우선 찾고(find_company_sheet_name), 없으면(또는 company를
    안 주면) 맨 마지막 시트로 폴백한다 — 개별법인 제출용 서식은 시트 이름이 법인명과
    무관("작성 기준"/"1")이라 항상 이 폴백(맨 마지막 시트)으로 자연스럽게 데이터
    시트가 선택된다. 시트를 찾은 뒤 detect_sheet_format()으로 두 서식 중 어느 쪽인지
    판별해 알맞은 파서로 분기한다. (items, sheet_name, unit_fix_count, sheet_format)를
    반환한다 — unit_fix_count는 원/억원 단위 혼재를 자동 보정한 항목 수(0이면 보정
    없음), sheet_format은 'consolidated'(그룹차입금보고서)/'individual'(개별법인
    제출용)."""
    import io
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    sheet_name = find_company_sheet_name(wb, company) if company else wb.sheetnames[-1]
    ws = wb[sheet_name]
    sheet_format = detect_sheet_format(ws)
    if sheet_format == 'individual':
        items, unit_fix_count = parse_individual_submission_sheet(ws)
    elif sheet_format == 'consolidated':
        items, unit_fix_count = parse_company_sheet(ws)
    else:
        raise ValueError(
            "이 시트가 어떤 서식인지 인식할 수 없습니다 — '금융기관' 헤더(그룹차입금보고서 형식) "
            "또는 '법인명'+'차입처' 헤더(개별법인 제출용 서식) 중 하나가 있어야 합니다."
        )

    override_divisor = _UNIT_FIX_DIVISOR_OVERRIDE.get(company)
    if override_divisor and unit_fix_count:
        # parse_company_sheet가 이미 1억으로 나눠뒀으니, 이 법인에 맞는 배율(100만)과의
        # 차이만큼(1억 / 100만 = 100배) 다시 곱해 되돌린다.
        rescale = 100_000_000 / override_divisor
        for it in items:
            it['limit'] = it['limit'] * rescale if it['limit'] else it['limit']
            it['balance'] = it['balance'] * rescale if it['balance'] else it['balance']

    return items, sheet_name, unit_fix_count, sheet_format


def _weighted_rate(balance_rate_pairs):
    total_balance = sum(b for b, _r in balance_rate_pairs)
    if total_balance <= 0:
        return None
    total = sum(b * r for b, r in balance_rate_pairs if r is not None)
    priced_balance = sum(b for b, r in balance_rate_pairs if r is not None)
    if priced_balance <= 0:
        return None
    return total / priced_balance


def build_balance_by_institution(company_items):
    """{company: {institution_col: balance}} + '합계' + 맨 아래 '계' 합산 행."""
    rows = {}
    for company, items in company_items.items():
        row = {col: 0.0 for col in INSTITUTION_COLUMNS}
        for it in items:
            col = normalize_institution_column(it['institution_raw'])
            if col:
                row[col] = row.get(col, 0.0) + it['balance']
        row['합계'] = sum(row[c] for c in INSTITUTION_COLUMNS)
        rows[company] = row
    total_row = {col: sum(rows[c][col] for c in rows) for col in INSTITUTION_COLUMNS}
    total_row['합계'] = sum(total_row[c] for c in INSTITUTION_COLUMNS)
    return rows, total_row


def build_rate_by_institution(company_items):
    """{company: {institution_col: 가중평균금리}} + 맨 아래 '계' 가중평균 행."""
    rows = {}
    pairs_by_col_total = {col: [] for col in INSTITUTION_COLUMNS}
    for company, items in company_items.items():
        pairs_by_col = {col: [] for col in INSTITUTION_COLUMNS}
        for it in items:
            col = normalize_institution_column(it['institution_raw'])
            if col:
                pairs_by_col[col].append((it['balance'], it['rate']))
                pairs_by_col_total[col].append((it['balance'], it['rate']))
        row = {col: _weighted_rate(pairs_by_col[col]) for col in INSTITUTION_COLUMNS}
        all_pairs = [p for lst in pairs_by_col.values() for p in lst]
        row['가중평균'] = _weighted_rate(all_pairs)
        rows[company] = row
    total_row = {col: _weighted_rate(pairs_by_col_total[col]) for col in INSTITUTION_COLUMNS}
    all_pairs_total = [p for lst in pairs_by_col_total.values() for p in lst]
    total_row['가중평균'] = _weighted_rate(all_pairs_total)
    return rows, total_row


SUBJECT_COLUMNS = ['일반대', '한도대(한도)', '한도대(잔액)', '회사채', '전단채', 'CP', '기타']


def build_balance_by_subject(company_items):
    """{company: {subject_col: 금액}} + '합계' + 맨 아래 '계' 합산 행. 합계는 한도대
    '한도'는 빼고(잔액 항목만) 합산한다 — 실제 그룹 보고서 5번 시트와 동일한 방식."""
    rows = {}
    for company, items in company_items.items():
        row = {col: 0.0 for col in SUBJECT_COLUMNS}
        for it in items:
            bucket = classify_subject_bucket(it['institution_raw'], it['subject'])
            if bucket == 'general':
                row['일반대'] += it['balance']
            elif bucket == 'limit':
                row['한도대(한도)'] += it['limit']
                row['한도대(잔액)'] += it['balance']
            elif bucket == 'bond':
                row['회사채'] += it['balance']
            elif bucket == 'stnote':
                row['전단채'] += it['balance']
            elif bucket == 'cp':
                row['CP'] += it['balance']
            else:
                row['기타'] += it['balance']
        row['합계'] = (
            row['일반대'] + row['한도대(잔액)'] + row['회사채'] + row['전단채'] + row['CP'] + row['기타']
        )
        rows[company] = row
    total_row = {col: sum(rows[c][col] for c in rows) for col in SUBJECT_COLUMNS}
    total_row['합계'] = sum(rows[c]['합계'] for c in rows)
    return rows, total_row


def build_summary(company_items, prev_company_items=None):
    """1.그룹총괄표 형태 — 법인별 당월 총한도ㆍ잔액ㆍ가중평균금리, (있으면) 전월 잔액과
    전월대비 증감. prev_company_items가 없으면(직전월 데이터 미등록) 전월 관련 값은
    None으로 비워둔다."""
    rows = {}
    for company, items in company_items.items():
        total_limit = sum(it['limit'] for it in items)
        total_balance = sum(it['balance'] for it in items)
        rate = _weighted_rate([(it['balance'], it['rate']) for it in items])
        prev_balance = None
        if prev_company_items and company in prev_company_items:
            prev_balance = sum(it['balance'] for it in prev_company_items[company])
        rows[company] = {
            '총한도': total_limit,
            '잔액': total_balance,
            '가중평균금리': rate,
            '전월잔액': prev_balance,
            '전월대비': (total_balance - prev_balance) if prev_balance is not None else None,
        }
    total_limit_all = sum(r['총한도'] for r in rows.values())
    total_balance_all = sum(r['잔액'] for r in rows.values())
    total_rate_all = _weighted_rate([
        (it['balance'], it['rate']) for items in company_items.values() for it in items
    ])
    prev_total = None
    if prev_company_items:
        prev_total = sum(sum(it['balance'] for it in items) for items in prev_company_items.values())
    total_row = {
        '총한도': total_limit_all,
        '잔액': total_balance_all,
        '가중평균금리': total_rate_all,
        '전월잔액': prev_total,
        '전월대비': (total_balance_all - prev_total) if prev_total is not None else None,
    }
    return rows, total_row
